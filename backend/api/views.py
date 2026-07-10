from django.http import JsonResponse, StreamingHttpResponse
from django.db import connections, models
from django.utils import timezone
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.cache import cache
from django.conf import settings as django_settings
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authentication import TokenAuthentication  # noqa: F401 (kept for compatibility)
from rest_framework.authtoken.models import Token
from .authentication import ExpiringTokenAuthentication
from rest_framework import status
from datetime import date, datetime
import decimal
import logging
import re as _re
import io
import openpyxl
from openpyxl.styles import Font
from functools import wraps

from .models import UserProfile, Reporte

logger = logging.getLogger(__name__)




# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPERS â€" VALIDACION INPUT / SEGURIDAD
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _safe_int(val, default):
    """Convierte val a int de forma segura; retorna default si no es parseable."""
    try:
        return int(val)
    except (TypeError, ValueError):
        return int(default)


_SAFE_STR_RE = _re.compile(r"[^\w \-\.&]", flags=_re.UNICODE)

def _safe_str(val, max_len=100):
    """Limpia y trunca un string de entrada; elimina caracteres no esperados."""
    if val is None:
        return ''
    return _SAFE_STR_RE.sub('', str(val).strip())[:max_len]


def _validate_anho_mes(anho, mes=None):
    """Retorna JsonResponse 400 si los parÃ¡metros de perÃ­odo estÃ¡n fuera de rango, None si son vÃ¡lidos."""
    if anho < 2020 or anho > 2100:
        return JsonResponse({'success': False, 'error': 'AÃ±o fuera de rango'}, status=400)
    if mes is not None and (mes < 1 or mes > 12):
        return JsonResponse({'success': False, 'error': 'Mes fuera de rango (1-12)'}, status=400)
    return None


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPERS â€" BRUTE FORCE LOGIN
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_MAX_ATTEMPTS    = getattr(django_settings, 'LOGIN_MAX_ATTEMPTS',    5)
_LOCKOUT_SECONDS = getattr(django_settings, 'LOGIN_LOCKOUT_SECONDS', 900)


def _login_key(username, ip):
    return 'login_fail:' + str(username)[:50] + ':' + str(ip)[:45]

def _login_key_user(username):
    # Clave independiente de IP â€" previene bypass rotando X-Forwarded-For
    return 'login_fail_u:' + str(username)[:50]


def _is_locked_out(username, ip):
    # Bloquea si el contador por (usuario+IP) O por usuario-solo alcanza el lÃ­mite
    return (
        cache.get(_login_key(username, ip), 0) >= _MAX_ATTEMPTS
        or cache.get(_login_key_user(username), 0) >= _MAX_ATTEMPTS
    )


def _record_failed_login(username, ip):
    for key in (_login_key(username, ip), _login_key_user(username)):
        cache.set(key, cache.get(key, 0) + 1, _LOCKOUT_SECONDS)


def _clear_failed_logins(username, ip):
    cache.delete(_login_key(username, ip))
    cache.delete(_login_key_user(username))


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPER â€" PERMISOS DE DASHBOARD
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_ADMIN_CARGOS = frozenset(['Administrador de Sistema', 'Subadministrador de Sistemas'])

# Permisos de dashboard por defecto segÃºn cargo (espejo del frontend adminConstants)
_PERMISOS_POR_CARGO: dict[str, list[str]] = {
    'Gerente General':    ['nacional', 'regionales', 'canales', 'supervisores', 'preventas-realizadas',
                           'avances-ventas', 'unidades-vendidas', 'unidades-supervisores',
                           'informacion-rutas', 'tendencia-estacional', 'ticket-promedio',
                           'ficha-sku', 'margen-bruto', 'matriz', 'descargas',
                           'pepsico', 'softys', 'dmujer', 'apego', 'colher'],
    'Gerente de Ventas':  ['nacional', 'regionales', 'canales', 'supervisores', 'unidades-vendidas',
                           'unidades-supervisores', 'informacion-rutas', 'tendencia-estacional',
                           'ticket-promedio', 'ficha-sku', 'margen-bruto'],
    'Gerente Regional':   ['regionales', 'canales', 'supervisores', 'preventas-realizadas',
                           'avances-ventas', 'unidades-vendidas', 'unidades-supervisores',
                           'informacion-rutas', 'tendencia-estacional'],
    'Supervisor':         ['canales', 'supervisores', 'preventas-realizadas', 'avances-ventas',
                           'unidades-supervisores', 'informacion-rutas'],
    'Vendedor':           ['preventas-realizadas', 'avances-ventas'],
    'Proveedor':          ['lista-precios', 'pepsico', 'softys', 'dmujer', 'apego', 'colher'],
    'Analista de Datos':  ['nacional', 'regionales', 'canales', 'supervisores', 'preventas-realizadas',
                           'avances-ventas', 'unidades-vendidas', 'unidades-supervisores',
                           'informacion-rutas', 'tendencia-estacional', 'ticket-promedio',
                           'ficha-sku', 'margen-bruto', 'matriz', 'descargas'],
}


def _has_dashboard_perm(user, perm_id):
    """True si el usuario tiene acceso al dashboard perm_id."""
    if user.is_staff or user.is_superuser:
        return True
    try:
        if user.profile.cargo in _ADMIN_CARGOS:
            return True
        return perm_id in (user.profile.dashboard_permissions or [])
    except Exception:
        logger.exception("Error verificando permiso user=%s perm=%s",
                         getattr(user, 'username', '?'), perm_id)
        return False


def _require_perm(perm_id):
    """Decorator que verifica permiso de dashboard antes de ejecutar la vista."""
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            if not _has_dashboard_perm(request.user, perm_id):
                logger.warning("ACCESS_DENIED user=%s perm=%s path=%s",
                               getattr(request.user, 'username', '?'), perm_id, request.path)
                return JsonResponse(
                    {'success': False, 'error': 'Sin acceso a este dashboard'},
                    status=403
                )
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


def _require_any_perm(*perm_ids):
    """Decorator que acepta cualquiera de los permisos listados."""
    def decorator(func):
        @wraps(func)
        def wrapper(request, *args, **kwargs):
            if not any(_has_dashboard_perm(request.user, p) for p in perm_ids):
                logger.warning("ACCESS_DENIED user=%s perms=%s path=%s",
                               getattr(request.user, 'username', '?'), perm_ids, request.path)
                return JsonResponse(
                    {'success': False, 'error': 'Sin acceso a este dashboard'},
                    status=403
                )
            return func(request, *args, **kwargs)
        return wrapper
    return decorator


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  AUTH
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username', '').strip()[:150]
    password = request.data.get('password', '')
    ip = (
        request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
        or request.META.get('REMOTE_ADDR', 'unknown')
    )

    if not username or not password:
        return JsonResponse(
            {'success': False, 'error': 'Usuario y contraseÃ±a requeridos'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if _is_locked_out(username, ip):
        logger.warning("AUTH_LOCKOUT user=%s ip=%s", username, ip)
        return JsonResponse(
            {'success': False, 'error': 'Cuenta bloqueada temporalmente. IntentÃ¡ en 15 minutos.'},
            status=status.HTTP_429_TOO_MANY_REQUESTS
        )

    user = authenticate(username=username, password=password)
    if not user:
        _record_failed_login(username, ip)
        logger.warning("AUTH_FAIL user=%s ip=%s", username, ip)
        return JsonResponse(
            {'success': False, 'error': 'Credenciales invÃ¡lidas'},
            status=status.HTTP_401_UNAUTHORIZED
        )

    _clear_failed_logins(username, ip)
    logger.warning("AUTH_OK user=%s ip=%s", username, ip)
    Token.objects.filter(user=user).delete()
    token = Token.objects.create(user=user)
    groups = list(user.groups.values_list('name', flat=True))

    return JsonResponse({
        'success': True,
        'token': token.key,
        'user': {
            'id': user.id,
            'username': user.username,
            'full_name': user.get_full_name(),
            'email': user.email,
            'groups': groups,
            'is_staff': user.is_staff,
        }
    })


@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def logout(request):
    try:
        request.user.auth_token.delete()
        return JsonResponse({'success': True, 'message': 'SesiÃ³n cerrada'})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def me(request):
    user = request.user
    return JsonResponse({'success': True, 'user': _serialize_user(user)})


@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def token_refresh(request):
    """Renueva el token del usuario autenticado (resetea el contador de expiraciÃ³n)."""
    user = request.user
    Token.objects.filter(user=user).delete()
    new_token = Token.objects.create(user=user)
    return JsonResponse({'success': True, 'token': new_token.key})


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPERS â€" USUARIOS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _get_or_create_profile(user):
    """Devuelve el perfil del usuario, creÃ¡ndolo si no existe."""
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile


def _serialize_user(user):
    """Serializa un User + su UserProfile a dict."""
    profile = _get_or_create_profile(user)
    return {
        'id':                    user.id,
        'username':              user.username,
        'first_name':            user.first_name,
        'last_name':             user.last_name,
        'full_name':             user.get_full_name(),
        'email':                 user.email,
        'cargo':                 profile.cargo,
        'regional':              profile.regional,
        'canal':                 profile.canal,
        'is_active':             user.is_active,
        'is_staff':              user.is_staff,
        'dashboard_permissions': profile.dashboard_permissions,
        'groups':                list(user.groups.values_list('name', flat=True)),
        'date_joined':           user.date_joined.isoformat() if user.date_joined else None,
        'last_seen':             profile.last_seen.isoformat() if profile.last_seen else None,
    }


_ADMIN_CARGOS_FULL = frozenset([
    'Administrador de Sistema',
    'Subadministrador de Sistemas',
    'Gerente General',
    'Gerente de Ventas',
    'Analista de Datos',
])

def _is_admin(user):
    """True si el usuario puede ver todos los filtros (regionales, canales)."""
    if user.is_staff or user.is_superuser:
        return True
    try:
        return user.profile.cargo in _ADMIN_CARGOS_FULL
    except UserProfile.DoesNotExist:
        return False


def _is_user_manager(user):
    """True si el usuario puede gestionar cuentas (crear, editar, resetear contraseÃ±as).
    Restringido a administradores de sistema â€" excluye GG/GV/Analista que solo tienen
    acceso ampliado a filtros."""
    if user.is_superuser:
        return True
    try:
        return user.profile.cargo in _ADMIN_CARGOS
    except UserProfile.DoesNotExist:
        return False


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPER: ejecutar SQL en el DW
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _run_dw_query(sql, params=None):
    """Ejecuta una consulta en la BD del DW y retorna columnas + filas como lista de dicts."""
    with connections['dw'].cursor() as cursor:
        cursor.execute(sql, params or [])
        columns = [col[0] for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                if val is None:
                    row_dict[col] = None
                elif isinstance(val, (date, datetime)):
                    row_dict[col] = val.isoformat()
                elif isinstance(val, decimal.Decimal):
                    row_dict[col] = float(val)
                elif isinstance(val, bytes):
                    row_dict[col] = val.decode('utf-8')
                else:
                    row_dict[col] = val
            rows.append(row_dict)
    return columns, rows


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD â€" VENTAS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_ventas_kpis(request):
    """KPIs principales: total ventas, cantidad pedidos, ticket promedio del mes actual."""
    try:
        sql = """
            SELECT
                COUNT(DISTINCT fv.numero_venta)          AS total_pedidos,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta,
                COALESCE(AVG(fv.venta_neta), 0)          AS ticket_promedio,
                COUNT(DISTINCT fv.cliente_sk)            AS clientes_activos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.mes_actual = TRUE
        """
        _, rows = _run_dw_query(sql)
        return JsonResponse({'success': True, 'data': rows[0] if rows else {}})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_ventas_por_mes(request):
    """Ventas netas agrupadas por mes (Ãºltimos 12 meses)."""
    try:
        sql = """
            SELECT
                df.anho_mes                              AS periodo,
                df.mes_nombre                            AS mes,
                df.anho                                  AS anho,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta,
                COUNT(DISTINCT fv.numero_venta)          AS total_pedidos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.fecha_completa >= CURRENT_DATE - INTERVAL '12 months'
            GROUP BY df.anho_mes, df.mes_nombre, df.anho
            ORDER BY df.anho_mes
        """
        _, rows = _run_dw_query(sql)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_ventas_por_canal(request):
    """Ventas netas por canal de vendedor (mes actual)."""
    try:
        sql = """
            SELECT
                dv.canal                                 AS canal,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta,
                COUNT(DISTINCT fv.numero_venta)          AS total_pedidos,
                COUNT(DISTINCT fv.vendedor_sk)           AS vendedores
            FROM dw.fact_ventas fv
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.mes_actual = TRUE
              AND dv.es_vendedor_actual = TRUE
            GROUP BY dv.canal
            ORDER BY total_venta_neta DESC
        """
        _, rows = _run_dw_query(sql)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD â€" VENDEDORES
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_vendedores_ranking(request):
    """Ranking de vendedores por venta neta del mes actual."""
    try:
        limit = min(_safe_int(request.GET.get('limit'), 20), 100)
        sql = """
            SELECT
                dv.vendedor_nombre                       AS vendedor,
                dv.canal                                 AS canal,
                dv.ciudad                                AS ciudad,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta,
                COUNT(DISTINCT fv.numero_venta)          AS total_pedidos,
                COUNT(DISTINCT fv.cliente_sk)            AS clientes_atendidos
            FROM dw.fact_ventas fv
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.mes_actual = TRUE
              AND dv.es_vendedor_actual = TRUE
            GROUP BY dv.vendedor_nombre, dv.canal, dv.ciudad
            ORDER BY total_venta_neta DESC
            LIMIT %s
        """
        _, rows = _run_dw_query(sql, [limit])
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD â€" PRODUCTOS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_productos_top(request):
    """Top productos por venta neta del mes actual."""
    try:
        limit = min(_safe_int(request.GET.get('limit'), 20), 100)
        sql = """
            SELECT
                dp.producto_nombre                       AS producto,
                dp.grupo_descripcion                    AS grupo,
                dp.subgrupo_descripcion                 AS subgrupo,
                COALESCE(SUM(fv.cantidad), 0)            AS total_cantidad,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.mes_actual = TRUE
              AND dp.es_producto_actual = TRUE
            GROUP BY dp.producto_nombre, dp.grupo_descripcion, dp.subgrupo_descripcion
            ORDER BY total_venta_neta DESC
            LIMIT %s
        """
        _, rows = _run_dw_query(sql, [limit])
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_productos_por_grupo(request):
    """Ventas por grupo de producto del mes actual."""
    try:
        sql = """
            SELECT
                dp.grupo_descripcion                    AS grupo,
                COUNT(DISTINCT dp.producto_sk)           AS productos,
                COALESCE(SUM(fv.cantidad), 0)            AS total_cantidad,
                COALESCE(SUM(fv.venta_neta), 0)          AS total_venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            WHERE df.mes_actual = TRUE
              AND dp.es_producto_actual = TRUE
            GROUP BY dp.grupo_descripcion
            ORDER BY total_venta_neta DESC
        """
        _, rows = _run_dw_query(sql)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD NACIONAL
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

CIUDADES = {
    'santa_cruz': ['SCZ'],
    'cochabamba': ['CBA'],
    'la_paz':     ['LPZ', 'EAL'],   # La Paz + El Alto
}

# Prefijos de ruta en dim_cliente_dual para filtrar por regional
_RUTA_PREFIJOS = {
    'santa_cruz': ["ruta LIKE 'SC%%'"],
    'cochabamba': ["ruta LIKE 'CB%%'", "ruta LIKE 'CBA%%'", "ruta LIKE 'ZONA-SUPERMERCADO CB%%'"],
    'la_paz':     ["ruta LIKE 'LP%%'", "ruta LIKE 'EA%%'"],
}

def _ruta_regional_cond(regional_key: str) -> str:
    """CondiciÃ³n SQL para filtrar dim_cliente_dual.ruta por regional."""
    prefijos = _RUTA_PREFIJOS.get(regional_key)
    if not prefijos:
        return '1=1'
    return '(' + ' OR '.join(prefijos) + ')'

CIUDAD_LABELS = {
    'santa_cruz': 'Santa Cruz',
    'cochabamba': 'Cochabamba',
    'la_paz':     'La Paz',
}


def _ciudad_case(campo: str, ciudad_key: str) -> str:
    """CondiciÃ³n SQL para filtrar por ciudad usando cÃ³digos exactos."""
    vals = CIUDADES[ciudad_key]
    if len(vals) == 1:
        return f"{campo} = '{vals[0]}'"
    quoted = ', '.join(f"'{v}'" for v in vals)
    return f"{campo} IN ({quoted})"


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_nacional_periodos(request):
    """Annos y meses disponibles en fact_ventas para los selectores."""
    cached = cache.get('periodos_disponibles')
    if cached is not None:
        return JsonResponse({'success': True, 'data': cached})
    try:
        sql = """
            SELECT DISTINCT df.anho, df.mes_numero, df.mes_nombre
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha df ON fv.fecha_sk = df.fecha_sk
            ORDER BY df.anho DESC, df.mes_numero DESC
            LIMIT 36
        """
        _, rows = _run_dw_query(sql)
        cache.set('periodos_disponibles', rows, 600)  # cache 10 min
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_nacional_kpis(request):
    """KPIs: total nacional + Santa Cruz + Cochabamba + La Paz. Params: anho, mes + filtros opcionales."""
    try:
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err

        canal       = _safe_str(request.GET.get('canal', ''))
        categorias  = [s for s in request.GET.getlist('categoria') if s]
        proveedores = [s for s in request.GET.getlist('proveedor') if s]
        subgrupos   = [s for s in request.GET.getlist('subgrupo')  if s]
        marcas      = [s for s in request.GET.getlist('marca')     if s]

        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        canal_param = [canal] if canal else []
        cat_cond,  cat_params  = _multi_cat_cond(categorias)
        prov_cond, prov_params = _multi_prov_cond(proveedores)
        sub_cond,  sub_params  = _multi_sub_cond(subgrupos)
        marc_cond, marc_params = _multi_marc_cond(marcas)
        prod_cond  = f"{cat_cond} {prov_cond} {sub_cond} {marc_cond}"
        prod_params = cat_params + prov_params + sub_params + marc_params
        has_prod   = bool(categorias or proveedores or subgrupos or marcas)
        prod_join  = "JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk" if has_prod else ""

        scz  = _ciudad_case('dv.ciudad', 'santa_cruz')
        cbba = _ciudad_case('dv.ciudad', 'cochabamba')
        lpz  = _ciudad_case('dv.ciudad', 'la_paz')

        # Ventas reales por regional (usando ciudad del vendedor)
        sql_ventas = f"""
            SELECT
                COALESCE(SUM(fv.venta_neta), 0)                                        AS total_nacional,
                COALESCE(SUM(CASE WHEN {scz}  THEN fv.venta_neta END), 0)              AS santa_cruz,
                COALESCE(SUM(CASE WHEN {cbba} THEN fv.venta_neta END), 0)              AS cochabamba,
                COALESCE(SUM(CASE WHEN {lpz}  THEN fv.venta_neta END), 0)              AS la_paz,
                COALESCE(SUM(fv.cantidad), 0)                                           AS cantidad_total,
                COALESCE(SUM(CASE WHEN {scz}  THEN fv.cantidad END), 0)                AS cantidad_santa_cruz,
                COALESCE(SUM(CASE WHEN {cbba} THEN fv.cantidad END), 0)                AS cantidad_cochabamba,
                COALESCE(SUM(CASE WHEN {lpz}  THEN fv.cantidad END), 0)                AS cantidad_la_paz,
                COUNT(DISTINCT fv.cliente_sk)                                           AS cobertura_total,
                COUNT(DISTINCT CASE WHEN {scz}  THEN fv.cliente_sk END)                AS cobertura_santa_cruz,
                COUNT(DISTINCT CASE WHEN {cbba} THEN fv.cliente_sk END)                AS cobertura_cochabamba,
                COUNT(DISTINCT CASE WHEN {lpz}  THEN fv.cliente_sk END)                AS cobertura_la_paz,
                MAX(df.fecha_completa)                                                  AS fecha_corte
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            {prod_join}
            WHERE df.anho = %s AND df.mes_numero = %s
              {canal_cond} {prod_cond}
        """
        _, rows = _run_dw_query(sql_ventas, [anho, mes] + canal_param + prod_params)
        data = rows[0] if rows else {}

        # Presupuesto desde fact_presupuesto (versión activa), respeta filtros de producto
        ppto_prod_join = "JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk" if has_prod else ""
        sql_ppto = f"""
            SELECT
                COALESCE(SUM(fp.venta_neta_presupuestada), 0)                            AS total,
                COALESCE(SUM(CASE WHEN {scz}  THEN fp.venta_neta_presupuestada END), 0)  AS santa_cruz,
                COALESCE(SUM(CASE WHEN {cbba} THEN fp.venta_neta_presupuestada END), 0)  AS cochabamba,
                COALESCE(SUM(CASE WHEN {lpz}  THEN fp.venta_neta_presupuestada END), 0)  AS la_paz
            FROM dw.fact_presupuesto fp
            JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
            {ppto_prod_join}
            WHERE fp.anho = %s AND fp.mes = %s
              {canal_cond} {prod_cond}
              AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
        """
        presupuestos = {'total': 0, 'santa_cruz': 0, 'cochabamba': 0, 'la_paz': 0}
        try:
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes] + canal_param + prod_params + [anho, mes])
            if ppto_rows:
                presupuestos = {k: v or 0 for k, v in ppto_rows[0].items()}
        except Exception:
            pass

        data['presupuesto'] = presupuestos
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_nacional_tendencia(request):
    """Avance diario acumulado + presupuesto acumulado + proyeccion lineal. Params: anho, mes + filtros."""
    try:
        import calendar
        from datetime import date as _date
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err
        hoy  = datetime.now().date()

        regional    = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        canal       = _safe_str(request.GET.get('canal', ''))
        categorias  = [s for s in request.GET.getlist('categoria') if s]
        proveedores = [s for s in request.GET.getlist('proveedor') if s]
        subgrupos   = [s for s in request.GET.getlist('subgrupo')  if s]
        marcas      = [s for s in request.GET.getlist('marca')     if s]
        if regional not in REGIONALES_VALID:
            regional = 'nacional'

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        canal_param = [canal] if canal else []
        cat_cond,  cat_params  = _multi_cat_cond(categorias)
        prov_cond, prov_params = _multi_prov_cond(proveedores)
        sub_cond,  sub_params  = _multi_sub_cond(subgrupos)
        marc_cond, marc_params = _multi_marc_cond(marcas)
        prod_cond  = f"{cat_cond} {prov_cond} {sub_cond} {marc_cond}"
        prod_params = cat_params + prov_params + sub_params + marc_params
        has_filters = bool(regional != 'nacional' or canal or categorias or proveedores or subgrupos or marcas)

        if has_filters:
            # Subquery para pre-filtrar fact_ventas; LEFT JOIN exterior preserva todos los días
            sql_avance = f"""
                WITH dias AS (
                    SELECT df.dia_numero,
                           COALESCE(SUM(fv_f.venta_neta), 0) AS venta_dia
                    FROM dw.dim_fecha df
                    LEFT JOIN (
                        SELECT fv.fecha_sk, fv.venta_neta
                        FROM dw.fact_ventas fv
                        JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
                        JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
                        WHERE ({ciudad_cond}) {canal_cond} {prod_cond}
                    ) fv_f ON fv_f.fecha_sk = df.fecha_sk
                    WHERE df.anho = %s AND df.mes_numero = %s
                      AND df.fecha_completa <= CURRENT_DATE
                    GROUP BY df.dia_numero
                )
                SELECT dia_numero AS dia,
                       venta_dia,
                       SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
                FROM dias ORDER BY dia_numero
            """
            avance_params = canal_param + prod_params + [anho, mes]
        else:
            sql_avance = """
                WITH dias AS (
                    SELECT df.dia_numero,
                           COALESCE(SUM(fv.venta_neta), 0) AS venta_dia
                    FROM dw.dim_fecha df
                    LEFT JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                    WHERE df.anho = %s AND df.mes_numero = %s
                      AND df.fecha_completa <= CURRENT_DATE
                    GROUP BY df.dia_numero
                )
                SELECT dia_numero AS dia,
                       venta_dia,
                       SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
                FROM dias ORDER BY dia_numero
            """
            avance_params = [anho, mes]

        _, avance_rows = _run_dw_query(sql_avance, avance_params)

        ppto_prod_join = "JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk" if (categorias or proveedores or subgrupos or marcas) else ""
        presupuesto_mes = 0.0
        try:
            _, p = _run_dw_query(
                f"""SELECT COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS total
                   FROM dw.fact_presupuesto fp
                   JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                   {ppto_prod_join}
                   WHERE fp.anho = %s AND fp.mes = %s
                     AND ({ciudad_cond}) {canal_cond} {prod_cond}
                     AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)""",
                [anho, mes] + canal_param + prod_params + [anho, mes]
            )
            presupuesto_mes = float(p[0]['total']) if p else 0.0
        except Exception:
            pass

        # Último día con ventas reales: maneja desfase de 1 día y ausencia de domingos
        # Nota: _run_dw_query convierte fechas a str ISO, por eso se parsea con fromisoformat
        fecha_corte = None
        try:
            fc_prod_join = "JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk" if (categorias or proveedores or subgrupos or marcas) else ""
            _, fc_rows = _run_dw_query(
                f"""SELECT MAX(df.fecha_completa) AS fc
                   FROM dw.dim_fecha df
                   JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                   JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
                   {fc_prod_join}
                   WHERE df.anho = %s AND df.mes_numero = %s
                     AND ({ciudad_cond}) {canal_cond} {prod_cond}""",
                [anho, mes] + canal_param + prod_params
            )
            if fc_rows and fc_rows[0].get('fc'):
                fecha_corte = _date.fromisoformat(str(fc_rows[0]['fc'])[:10])
        except Exception:
            pass

        # Fallback: si el query de fecha_corte falla, derivar del último día con ventas reales
        if fecha_corte is None and avance_rows:
            dias_con_ventas = [int(r['dia']) for r in avance_rows if float(r.get('venta_dia') or 0) > 0]
            if dias_con_ventas:
                fecha_corte = _date(anho, mes, max(dias_con_ventas))

        dias_en_mes       = calendar.monthrange(anho, mes)[1]
        es_periodo_actual = (anho == hoy.year and mes == hoy.month)

        # Para la proyección: arrancar desde fecha_corte (no hoy) para manejar el desfase
        if fecha_corte:
            dias_transcurridos = fecha_corte.day
        elif es_periodo_actual:
            dias_transcurridos = hoy.day
        else:
            dias_transcurridos = dias_en_mes

        # Presupuesto distribuido solo entre días laborables (Lun–Sáb) del mes
        dias_lab_mes     = sum(1 for d in range(1, dias_en_mes + 1) if _date(anho, mes, d).weekday() != 6)
        ppto_por_dia_lab = presupuesto_mes / dias_lab_mes if dias_lab_mes > 0 and presupuesto_mes > 0 else 0

        avance_por_dia = {int(r['dia']): r['avance_acumulado'] for r in avance_rows}
        avance_total   = float(avance_rows[-1]['avance_acumulado']) if avance_rows else 0.0
        tasa_diaria    = avance_total / dias_transcurridos if dias_transcurridos > 0 else 0

        result       = []
        dias_lab_ac  = 0  # días laborables acumulados para el cálculo del presupuesto
        for dia in range(1, dias_en_mes + 1):
            fecha_dia  = _date(anho, mes, dia)
            es_domingo = fecha_dia.weekday() == 6

            if not es_domingo:
                dias_lab_ac += 1

            # Presupuesto: None en domingos y a partir del día siguiente a fecha_corte
            if es_domingo or (es_periodo_actual and fecha_corte and fecha_dia > fecha_corte):
                ppto_ac = None
            elif ppto_por_dia_lab > 0:
                ppto_ac = round(ppto_por_dia_lab * dias_lab_ac, 2)
            else:
                ppto_ac = None

            # Proyección: solo días futuros del período actual, excluir domingos
            proyeccion = None
            if es_periodo_actual and not es_domingo and dia > dias_transcurridos:
                proyeccion = round(avance_total + tasa_diaria * (dia - dias_transcurridos), 2)

            # Avance: None en domingos y días posteriores a fecha_corte
            if es_domingo or (fecha_corte and fecha_dia > fecha_corte):
                av_ac = None
            else:
                av_ac = avance_por_dia.get(dia)

            result.append({
                'dia':                   dia,
                'avance_acumulado':      av_ac,
                'presupuesto_acumulado': ppto_ac,
                'proyeccion_acumulada':  proyeccion,
            })

        return JsonResponse({
            'success': True, 'data': result,
            'es_periodo_actual': es_periodo_actual,
            'presupuesto_mes': presupuesto_mes,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_nacional_por_regional(request):
    """Presupuesto vs avance por regional. Params: anho, mes."""
    try:
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)

        scz  = _ciudad_case('dv.ciudad', 'santa_cruz')
        cbba = _ciudad_case('dv.ciudad', 'cochabamba')
        lpz  = _ciudad_case('dv.ciudad', 'la_paz')

        sql = f"""
            SELECT
                CASE
                    WHEN {scz}  THEN 'Santa Cruz'
                    WHEN {cbba} THEN 'Cochabamba'
                    WHEN {lpz}  THEN 'La Paz'
                    ELSE 'Otras'
                END                             AS regional,
                COALESCE(SUM(fv.venta_neta), 0) AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
            GROUP BY regional ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes])

        ppto_map: dict = {}
        try:
            sql_ppto = f"""
                SELECT
                    CASE
                        WHEN {_ciudad_case('dv.ciudad', 'santa_cruz')} THEN 'Santa Cruz'
                        WHEN {_ciudad_case('dv.ciudad', 'cochabamba')} THEN 'Cochabamba'
                        WHEN {_ciudad_case('dv.ciudad', 'la_paz')}     THEN 'La Paz'
                        ELSE 'Otras'
                    END                                                   AS regional,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)         AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv              ON fp.vendedor_sk = dv.vendedor_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY regional
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            ppto_map = {r['regional']: r['presupuesto'] for r in ppto_rows}
        except Exception:
            pass

        for row in rows:
            ppto = ppto_map.get(row['regional'], 0)
            row['presupuesto'] = ppto
            row['porcentaje']  = round(row['avance'] / ppto * 100, 1) if ppto > 0 else None

        principales = [r for r in rows if r['regional'] in ('Santa Cruz', 'Cochabamba', 'La Paz')]
        return JsonResponse({'success': True, 'data': principales})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(["GET"])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_nacional_por_canal(request):
    """Presupuesto vs avance por canal de venta (DTS, WHS, INST, LICORES, SPM, PROV). Params: anho, mes."""
    try:
        anho = _safe_int(request.GET.get("anho"), datetime.now().year)
        mes  = _safe_int(request.GET.get("mes"),  datetime.now().month)

        sql = """
            SELECT
                dv.canal_rrhh                        AS canal,
                COALESCE(SUM(fv.venta_neta), 0)      AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL
              AND dv.es_vendedor_actual = TRUE
            GROUP BY dv.canal_rrhh
            ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes])

        # Presupuesto por canal consolidado desde fact_presupuesto
        ppto_map = {}
        try:
            sql_ppto = """
                SELECT
                    CASE
                        WHEN dv.canal_rrhh IN ('DTS', 'DTS-LP', 'DTS-EA') THEN 'DTS'
                        WHEN dv.canal_rrhh IN ('WHS', 'WHS-LP', 'WHS-EA') THEN 'WHS'
                        ELSE dv.canal_rrhh
                    END                                               AS canal,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)     AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv              ON fp.vendedor_sk = dv.vendedor_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY CASE
                        WHEN dv.canal_rrhh IN ('DTS', 'DTS-LP', 'DTS-EA') THEN 'DTS'
                        WHEN dv.canal_rrhh IN ('WHS', 'WHS-LP', 'WHS-EA') THEN 'WHS'
                        ELSE dv.canal_rrhh
                    END
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            ppto_map = {r["canal"]: r["presupuesto"] for r in ppto_rows}
        except Exception:
            pass

        # Consolidar canales DTS y WHS en ventas tambiÃ©n
        canal_consolidado: dict = {}
        for row in rows:
            canal_orig = row["canal"]
            canal_key = (
                'DTS' if canal_orig in ('DTS', 'DTS-LP', 'DTS-EA') else
                'WHS' if canal_orig in ('WHS', 'WHS-LP', 'WHS-EA') else
                canal_orig
            )
            if canal_key not in canal_consolidado:
                canal_consolidado[canal_key] = {'canal': canal_key, 'avance': 0.0}
            canal_consolidado[canal_key]['avance'] += float(row['avance'] or 0)

        result = []
        for canal_key, row in canal_consolidado.items():
            ppto = ppto_map.get(canal_key, 0)
            result.append({
                'canal':       canal_key,
                'avance':      row['avance'],
                'presupuesto': ppto,
                'porcentaje':  round(row['avance'] / ppto * 100, 1) if ppto > 0 else None,
            })
        result.sort(key=lambda r: r['avance'], reverse=True)

        return JsonResponse({"success": True, "data": result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({"success": False, "error": "Error interno del servidor"}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  AUTH â€" CAMBIAR CONTRASEÃ‘A PROPIA
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def auth_change_password(request):
    user             = request.user
    current_password = request.data.get('current_password', '')
    new_password     = request.data.get('new_password', '')

    if not current_password or not new_password:
        return JsonResponse(
            {'success': False, 'error': 'Campos requeridos'},
            status=status.HTTP_400_BAD_REQUEST
        )
    if not user.check_password(current_password):
        return JsonResponse(
            {'success': False, 'error': 'La contraseÃ±a actual es incorrecta'},
            status=status.HTTP_400_BAD_REQUEST
        )
    if len(new_password) < 6:
        return JsonResponse(
            {'success': False, 'error': 'La nueva contraseÃ±a debe tener al menos 6 caracteres'},
            status=status.HTTP_400_BAD_REQUEST
        )

    user.set_password(new_password)
    user.save()
    # Invalida token actual y genera uno nuevo
    user.auth_token.delete()
    new_token, _ = Token.objects.get_or_create(user=user)
    return JsonResponse({'success': True, 'token': new_token.key})


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  ADMIN â€" GESTIÃ"N DE USUARIOS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

ADMIN_CARGOS = {'Administrador de Sistema', 'Subadministrador de Sistemas'}


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def admin_list_users(request):
    """Lista todos los usuarios del sistema (excluye superusuarios)."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    users = (
        User.objects
        .filter(is_superuser=False)
        .order_by('first_name', 'last_name', 'username')
    )
    return JsonResponse([_serialize_user(u) for u in users], safe=False)


@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def admin_create_user(request):
    """Crea un nuevo usuario con su perfil."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    data                  = request.data
    username              = data.get('username', '').strip()
    first_name            = data.get('first_name', '').strip()
    last_name             = data.get('last_name', '').strip()
    email                 = data.get('email', '').strip()
    cargo                 = data.get('cargo', '')
    regional              = data.get('regional', '')
    canal                 = data.get('canal', '')
    password              = data.get('password', '')
    dashboard_permissions = data.get('dashboard_permissions', [])
    if not isinstance(dashboard_permissions, list) or len(dashboard_permissions) == 0:
        dashboard_permissions = _PERMISOS_POR_CARGO.get(cargo, [])

    if not username:
        return JsonResponse({'success': False, 'error': 'El nombre de usuario es requerido'}, status=400)
    if not password or len(password) < 6:
        return JsonResponse({'success': False, 'error': 'La contraseÃ±a debe tener al menos 6 caracteres'}, status=400)
    if User.objects.filter(username=username).exists():
        return JsonResponse(
            {'success': False, 'error': f'El usuario "{username}" ya existe'},
            status=400
        )

    new_user = User.objects.create_user(
        username   = username,
        first_name = first_name,
        last_name  = last_name,
        email      = email,
        password   = password,
        is_staff   = (cargo in ADMIN_CARGOS),
    )
    UserProfile.objects.create(
        user                  = new_user,
        cargo                 = cargo,
        regional              = regional,
        canal                 = canal,
        dashboard_permissions = dashboard_permissions,
    )
    logger.warning("ADMIN_CREATE_USER actor=%s new_user=%s cargo=%s", request.user.username, username, cargo)
    return JsonResponse({'success': True, 'user': _serialize_user(new_user)}, status=201)


@api_view(['PATCH'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def admin_update_user(request, user_id):
    """Actualiza datos bÃ¡sicos + cargo/regional de un usuario."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    if request.user.pk == user_id:
        return JsonResponse({'success': False, 'error': 'No puedes modificar tu propia cuenta'}, status=403)

    try:
        target = User.objects.get(pk=user_id, is_superuser=False)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    data = request.data

    # Username: verificar unicidad si cambiÃ³
    new_username = data.get('username', '').strip()
    if new_username and new_username != target.username:
        if User.objects.filter(username=new_username).exclude(pk=target.pk).exists():
            return JsonResponse(
                {'success': False, 'error': f'El usuario "{new_username}" ya estÃ¡ en uso'},
                status=400
            )
        target.username = new_username

    if 'first_name' in data:
        target.first_name = data['first_name'].strip()
    if 'last_name' in data:
        target.last_name = data['last_name'].strip()
    if 'email' in data:
        target.email = data['email'].strip()
    if 'is_active' in data:
        target.is_active = bool(data['is_active'])

    cargo = data.get('cargo')
    if cargo is not None:
        target.is_staff = cargo in ADMIN_CARGOS

    target.save()

    profile = _get_or_create_profile(target)
    if cargo is not None:
        profile.cargo = cargo
    if 'regional' in data:
        profile.regional = data['regional']
    if 'canal' in data:
        profile.canal = data['canal']
    profile.save()

    return JsonResponse({'success': True, 'user': _serialize_user(target)})


@api_view(['PATCH'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def admin_update_permissions(request, user_id):
    """Actualiza los permisos de dashboards de un usuario."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    if request.user.pk == user_id:
        return JsonResponse({'success': False, 'error': 'No puedes modificar tu propia cuenta'}, status=403)

    try:
        target = User.objects.get(pk=user_id, is_superuser=False)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    perms = request.data.get('dashboard_permissions', [])
    if not isinstance(perms, list):
        return JsonResponse({'success': False, 'error': 'dashboard_permissions debe ser una lista'}, status=400)

    profile = _get_or_create_profile(target)
    profile.dashboard_permissions = perms
    profile.save()
    logger.warning("ADMIN_UPDATE_PERMS actor=%s target=%s perms=%s", request.user.username, target.username, perms)

    return JsonResponse({'success': True, 'user': _serialize_user(target)})


@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def admin_set_password(request, user_id):
    """Establece nueva contraseÃ±a para cualquier usuario (sin requerir la actual)."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    if request.user.pk == user_id:
        return JsonResponse({'success': False, 'error': 'Usa /auth/change-password/ para cambiar tu propia contraseÃ±a'}, status=403)

    try:
        target = User.objects.get(pk=user_id, is_superuser=False)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    new_password = request.data.get('new_password', '')
    if not new_password or len(new_password) < 6:
        return JsonResponse(
            {'success': False, 'error': 'La contraseÃ±a debe tener al menos 6 caracteres'},
            status=400
        )

    target.set_password(new_password)
    target.save()
    # Invalida sesiones activas del usuario afectado
    Token.objects.filter(user=target).delete()
    logger.warning("ADMIN_SET_PASSWORD actor=%s target=%s", request.user.username, target.username)
    return JsonResponse({'success': True, 'message': 'ContraseÃ±a actualizada correctamente'})


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD NACIONAL â€" POR CATEGORÃA
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('nacional')
def dashboard_nacional_por_categoria(request):
    """Ventas vs presupuesto por grupo de categorÃ­a a nivel nacional (4 categorÃ­as principales, excluyendo Exhibidores). Params: anho, mes."""
    try:
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err

        sql = """
            SELECT
                CASE dp.linea
                    WHEN 'ALIMENTOS'            THEN 'Alimentos'
                    WHEN 'APEGO'                THEN 'Apego'
                    WHEN 'BEBIDAS ALC'          THEN 'Licores'
                    WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
                    ELSE 'Sin clasificar'
                END AS categoria,
                COALESCE(SUM(fv.venta_neta), 0) AS venta_neta,
                COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                COUNT(DISTINCT fv.producto_sk)  AS productos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dp.linea IS NOT NULL
            GROUP BY dp.linea
            ORDER BY venta_neta DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes])

        # Presupuesto por categoria consolidado desde fact_presupuesto
        ppto_map = {}
        try:
            sql_ppto = """
                SELECT
                    CASE dp.linea
                        WHEN 'ALIMENTOS'            THEN 'Alimentos'
                        WHEN 'APEGO'                THEN 'Apego'
                        WHEN 'BEBIDAS ALC'          THEN 'Licores'
                        WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
                        ELSE 'Sin clasificar'
                    END AS categoria,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                  AND dp.linea IS NOT NULL
                GROUP BY dp.linea
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            ppto_map = {r["categoria"]: r["presupuesto"] for r in ppto_rows}
        except Exception:
            pass

        # Consolidar resultados con presupuesto
        result = []
        for row in rows:
            categoria = row["categoria"]
            venta_neta = float(row['venta_neta'] or 0)
            ppto = ppto_map.get(categoria, 0)
            result.append({
                'categoria': categoria,
                'venta_neta': venta_neta,
                'cantidad': int(row['cantidad'] or 0),
                'productos': int(row['productos'] or 0),
                'presupuesto': ppto,
                'porcentaje': round(venta_neta / ppto * 100, 1) if ppto > 0 else None,
            })
        result.sort(key=lambda r: r['venta_neta'], reverse=True)

        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  HELPERS REGIONALES
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

REGIONALES_VALID = {'santa_cruz', 'cochabamba', 'la_paz', 'nacional'}


def _regional_filter(regional_key, campo='dv.ciudad'):
    if not regional_key or regional_key == 'nacional':
        return '1=1'
    if regional_key not in CIUDADES:
        return '1=1'
    return _ciudad_case(campo, regional_key)


def _ppto_by_regional(anho, mes, ciudad_cond, canal_filter='', params_extra=None):
    """Retorna dict canal_rrhh->presupuesto para la regional dada."""
    sql = f"""
        SELECT dv.canal_rrhh AS canal, COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto
        FROM dw.fact_presupuesto fp
        JOIN dw.dim_vendedor dv             ON fp.vendedor_sk = dv.vendedor_sk
        WHERE fp.anho = %s AND fp.mes = %s
          AND ({ciudad_cond})
          AND dv.canal_rrhh IS NOT NULL
          {canal_filter}
          AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
        GROUP BY dv.canal_rrhh
    """
    base = [anho, mes] + (params_extra or []) + [anho, mes]
    try:
        _, rows = _run_dw_query(sql, base)
        return {r['canal']: float(r['presupuesto'] or 0) for r in rows}
    except Exception:
        return {}


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD REGIONALES
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('regionales')
def dashboard_regionales_kpis(request):
    """KPIs: total regional + canales desglosados. Params: regional, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err      = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)

        sql = f"""
            SELECT
                dv.canal_rrhh                               AS canal,
                COALESCE(SUM(fv.venta_neta), 0)             AS avance,
                MAX(df.fecha_completa)                      AS fecha_corte
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL
              AND ({ciudad_cond})
            GROUP BY dv.canal_rrhh
            ORDER BY avance DESC
        """
        _, ventas_rows = _run_dw_query(sql, [anho, mes])
        ppto_map = _ppto_by_regional(anho, mes, ciudad_cond)
        ppto_trim = ppto_map  # _ppto_by_regional ya devuelve canal_rrhh

        total_avance  = sum(float(r['avance'] or 0) for r in ventas_rows)
        total_ppto    = sum(ppto_trim.values())
        fecha_corte   = max((r['fecha_corte'] for r in ventas_rows if r.get('fecha_corte')), default=None)

        canales = []
        for row in ventas_rows:
            canal = row['canal']
            ppto  = ppto_trim.get(canal, 0)
            canales.append({'nombre': canal, 'avance': float(row['avance'] or 0), 'objetivo': ppto})

        return JsonResponse({'success': True, 'data': {
            'total':          total_avance,
            'objetivo_total': total_ppto,
            'canales':        canales,
            'fecha_corte':    fecha_corte,
        }})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('regionales')
def dashboard_regionales_tendencia(request):
    """Avance diario acumulado para una regional. Params: regional, anho, mes."""
    try:
        import calendar
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        hoy      = datetime.now().date()
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond  = _regional_filter(regional)
        ciudad_cond2 = _regional_filter(regional, campo='dv2.ciudad')

        sql_avance = f"""
            WITH dias AS (
                SELECT df.dia_numero,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_dia
                FROM dw.dim_fecha df
                LEFT JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_vendedor dv2
                        WHERE dv2.vendedor_sk = fv.vendedor_sk AND ({ciudad_cond2})
                    )
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND df.fecha_completa <= CURRENT_DATE
                GROUP BY df.dia_numero
            )
            SELECT dia_numero AS dia, venta_dia,
                   SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
            FROM dias ORDER BY dia_numero
        """
        _, avance_rows = _run_dw_query(sql_avance, [anho, mes])

        sql_ppto = f"""
            SELECT COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS total
            FROM dw.fact_presupuesto fp
            JOIN dw.dim_vendedor dv             ON fp.vendedor_sk = dv.vendedor_sk
            WHERE fp.anho = %s AND fp.mes = %s
              AND ({ciudad_cond})
              AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
        """
        presupuesto_mes = 0.0
        try:
            _, p = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            presupuesto_mes = float(p[0]['total']) if p else 0.0
        except Exception:
            pass

        dias_en_mes        = calendar.monthrange(anho, mes)[1]
        ppto_diario        = presupuesto_mes / dias_en_mes if dias_en_mes > 0 else 0
        es_periodo_actual  = (anho == hoy.year and mes == hoy.month)
        dias_transcurridos = hoy.day if es_periodo_actual else dias_en_mes
        avance_por_dia     = {int(r['dia']): r['avance_acumulado'] for r in avance_rows}
        avance_total       = float(avance_rows[-1]['avance_acumulado']) if avance_rows else 0.0
        tasa_diaria        = avance_total / dias_transcurridos if dias_transcurridos > 0 else 0

        result = []
        for dia in range(1, dias_en_mes + 1):
            proyeccion = None
            if es_periodo_actual and dia > dias_transcurridos:
                proyeccion = round(avance_total + tasa_diaria * (dia - dias_transcurridos), 2)
            result.append({
                'dia':                   dia,
                'avance_acumulado':      avance_por_dia.get(dia),
                'presupuesto_acumulado': round(ppto_diario * dia, 2) if presupuesto_mes > 0 else None,
                'proyeccion_acumulada':  proyeccion,
            })
        return JsonResponse({'success': True, 'data': result,
                             'presupuesto_mes': presupuesto_mes,
                             'es_periodo_actual': es_periodo_actual})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('regionales')
def dashboard_regionales_por_canal(request):
    """Avance vs presupuesto por canal de una regional. Params: regional, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        sql = f"""
            SELECT dv.canal_rrhh AS canal, COALESCE(SUM(fv.venta_neta), 0) AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL AND ({ciudad_cond})
            GROUP BY dv.canal_rrhh ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes])
        ppto_map = _ppto_by_regional(anho, mes, ciudad_cond)
        for row in rows:
            ppto = ppto_map.get(row['canal'], 0)
            row['presupuesto'] = ppto
            row['porcentaje']  = round(row['avance'] / ppto * 100, 1) if ppto > 0 else None
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


_CATEGORIA_CASE = """
    CASE dp.linea
        WHEN 'ALIMENTOS'            THEN 'Alimentos'
        WHEN 'APEGO'                THEN 'Apego'
        WHEN 'BEBIDAS ALC'          THEN 'Licores'
        WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
        ELSE 'Sin clasificar'
    END
"""


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('regionales')
def dashboard_regionales_por_categoria(request):
    """Ventas vs presupuesto por categorÃ­a consolidada de una regional. Params: regional, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err      = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)

        sql = f"""
            SELECT
                {_CATEGORIA_CASE}                                AS categoria,
                COALESCE(SUM(fv.venta_neta), 0)                  AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s AND ({ciudad_cond})
              AND dp.linea IS NOT NULL AND dp.linea != 'SIN LINEA'
            GROUP BY {_CATEGORIA_CASE}
            ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes])

        # Presupuesto por categorÃ­a consolidado con filtro regional
        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT
                    {_CATEGORIA_CASE}                                        AS categoria,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)             AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv              ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp              ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond})
                  AND dp.grupo_descripcion != 'EXHIBIDORES'
                  AND dp.grupo_descripcion IS NOT NULL
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY {_CATEGORIA_CASE}
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            ppto_map = {r['categoria']: float(r['presupuesto'] or 0) for r in ppto_rows}
        except Exception:
            pass

        result = []
        for row in rows:
            cat   = row['categoria']
            av    = float(row['avance'] or 0)
            ppto  = ppto_map.get(cat, 0)
            result.append({
                'categoria':   cat,
                'avance':      av,
                'presupuesto': ppto,
                'porcentaje':  round(av / ppto * 100, 1) if ppto > 0 else None,
            })

        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD CANALES / REGIONAL
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('canales')
def dashboard_canales_kpis(request):
    """KPI cards por canal con filtro opcional. Params: regional, canal, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err      = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond  = _regional_filter(regional)
        canal_cond   = "AND dv.canal_rrhh = %s" if canal else ""
        params       = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT
                dv.canal_rrhh                            AS canal,
                COALESCE(SUM(fv.venta_neta), 0)          AS avance,
                COUNT(DISTINCT fv.numero_venta)           AS pedidos,
                COUNT(DISTINCT fv.cliente_sk)             AS clientes
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL AND ({ciudad_cond})
              {canal_cond}
            GROUP BY dv.canal_rrhh ORDER BY avance DESC
        """
        _, ventas_rows = _run_dw_query(sql, params)
        canal_filter_ppto = "AND dv.canal_rrhh = %s" if canal else ""
        ppto_map = _ppto_by_regional(anho, mes, ciudad_cond, canal_filter_ppto, [canal] if canal else None)

        result = []
        for row in ventas_rows:
            ppto = ppto_map.get(row['canal'], 0)
            result.append({**row, 'presupuesto': ppto,
                           'porcentaje': round(row['avance'] / ppto * 100, 1) if ppto > 0 else None})
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('canales')
def dashboard_canales_tendencia(request):
    """Tendencia diaria para canal+regional. Params: regional, canal, anho, mes."""
    try:
        import calendar
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        hoy      = datetime.now().date()
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond2 = _regional_filter(regional, campo='dv2.ciudad')
        canal_cond2  = "AND dv2.canal_rrhh = %s" if canal else ""
        params_avance = ([canal] if canal else []) + [anho, mes]

        sql_avance = f"""
            WITH dias AS (
                SELECT df.dia_numero,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_dia
                FROM dw.dim_fecha df
                LEFT JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_vendedor dv2
                        WHERE dv2.vendedor_sk = fv.vendedor_sk
                          AND ({ciudad_cond2})
                          {canal_cond2}
                    )
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND df.fecha_completa <= CURRENT_DATE
                GROUP BY df.dia_numero
            )
            SELECT dia_numero AS dia, venta_dia,
                   SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
            FROM dias ORDER BY dia_numero
        """
        _, avance_rows = _run_dw_query(sql_avance, params_avance)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        ppto_map = _ppto_by_regional(anho, mes, ciudad_cond, canal_cond, [canal] if canal else None)
        presupuesto_mes = sum(ppto_map.values())

        dias_en_mes        = calendar.monthrange(anho, mes)[1]
        ppto_diario        = presupuesto_mes / dias_en_mes if dias_en_mes > 0 else 0
        es_periodo_actual  = (anho == hoy.year and mes == hoy.month)
        dias_transcurridos = hoy.day if es_periodo_actual else dias_en_mes
        avance_por_dia     = {int(r['dia']): r['avance_acumulado'] for r in avance_rows}
        avance_total       = float(avance_rows[-1]['avance_acumulado']) if avance_rows else 0.0
        tasa_diaria        = avance_total / dias_transcurridos if dias_transcurridos > 0 else 0

        result = []
        for dia in range(1, dias_en_mes + 1):
            proyeccion = None
            if es_periodo_actual and dia > dias_transcurridos:
                proyeccion = round(avance_total + tasa_diaria * (dia - dias_transcurridos), 2)
            result.append({
                'dia':                   dia,
                'avance_acumulado':      avance_por_dia.get(dia),
                'presupuesto_acumulado': round(ppto_diario * dia, 2) if presupuesto_mes > 0 else None,
                'proyeccion_acumulada':  proyeccion,
            })
        return JsonResponse({'success': True, 'data': result,
                             'presupuesto_mes': presupuesto_mes,
                             'es_periodo_actual': es_periodo_actual})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('canales')
def dashboard_canales_por_categoria(request):
    """Ventas vs presupuesto por categorÃ­a consolidada para canal+regional. Params: regional, canal, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params      = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT
                {_CATEGORIA_CASE}                                AS categoria,
                COALESCE(SUM(fv.venta_neta), 0)                  AS avance,
                COALESCE(SUM(fv.cantidad), 0)                    AS cantidad,
                COUNT(DISTINCT fv.producto_sk)                   AS productos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond}
              AND dp.linea IS NOT NULL AND dp.linea != 'SIN LINEA'
            GROUP BY {_CATEGORIA_CASE}
            ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, params)

        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT
                    {_CATEGORIA_CASE}                                        AS categoria,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)             AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv              ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp              ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond}
                  AND dp.grupo_descripcion != 'EXHIBIDORES'
                  AND dp.grupo_descripcion IS NOT NULL
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY {_CATEGORIA_CASE}
            """
            _, ppto_rows = _run_dw_query(sql_ppto, params + [anho, mes])
            ppto_map = {r['categoria']: float(r['presupuesto'] or 0) for r in ppto_rows}
        except Exception:
            pass

        result = []
        for row in rows:
            cat  = row['categoria']
            av   = float(row['avance'] or 0)
            ppto = ppto_map.get(cat, 0)
            result.append({
                'categoria':  cat,
                'avance':     av,
                'cantidad':   int(row['cantidad'] or 0),
                'productos':  int(row['productos'] or 0),
                'presupuesto': ppto,
                'porcentaje': round(av / ppto * 100, 1) if ppto > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('canales')
def dashboard_canales_por_sku(request):
    """
    Top SKUs para canal+categorÃ­a+regional.
    Params: regional, canal, categoria, anho, mes, limit
    """
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        categoria = _safe_str(request.GET.get('categoria', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err       = _validate_anho_mes(anho, mes)
        if err: return err
        limit     = min(_safe_int(request.GET.get('limit'), 500), 1000)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""

        params = [anho, mes]
        if canal:
            params.append(canal)

        cat_cond = ""
        if categoria and categoria in _UNIDADES_CAT_LINEA:
            cat_cond = "AND dp.linea = %s"
            params.append(_UNIDADES_CAT_LINEA[categoria])

        params_ventas = list(params) + [limit]

        sql = f"""
            SELECT
                dp.producto_codigo_erp                           AS codigo,
                dp.producto_nombre                               AS producto,
                COALESCE(dp.linea, 'Sin LÃ­nea')                  AS categoria,
                COALESCE(dp.subgrupo_descripcion, '')            AS subgrupo,
                COALESCE(SUM(fv.cantidad), 0)                    AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0)                  AS venta_neta,
                COUNT(DISTINCT fv.cliente_sk)                    AS clientes
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre,
                     dp.linea, dp.subgrupo_descripcion
            ORDER BY venta_neta DESC
            LIMIT %s
        """
        _, rows = _run_dw_query(sql, params_ventas)

        # Presupuesto por producto (Bs + Uds)
        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT dp.producto_codigo_erp                              AS codigo,
                       COALESCE(SUM(fp.venta_neta_presupuestada), 0)       AS presupuesto,
                       COALESCE(SUM(fp.cantidad_presupuestada), 0)         AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv              ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp              ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {cat_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.producto_codigo_erp
            """
            _, ppto_rows = _run_dw_query(sql_ppto, params + [anho, mes])
            ppto_map = {r['codigo']: r for r in ppto_rows}
        except Exception:
            pass

        result = []
        for row in rows:
            vn       = float(row['venta_neta'] or 0)
            cant     = int(row['cantidad'] or 0)
            p        = ppto_map.get(row['codigo'], {})
            ppto_bs  = float(p.get('presupuesto',     0) or 0)
            ppto_uds = float(p.get('presupuesto_uds', 0) or 0)
            result.append({
                **row,
                'venta_neta':    vn,
                'presupuesto':   ppto_bs,
                'presupuesto_uds': int(ppto_uds),
                'porcentaje':    round(vn   / ppto_bs  * 100, 1) if ppto_bs  > 0 else None,
                'porcentaje_uds': round(cant / ppto_uds * 100, 1) if ppto_uds > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD SOFTYS â€" CANALES / REGIONAL
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_SOFTYS_COND     = "(UPPER(dp.proveedor)  = 'SOFTYS' OR UPPER(dp.cat_comercial)  = 'SOFTYS')"
_SOFTYS_COND_DP2 = "(UPPER(dp2.proveedor) = 'SOFTYS' OR UPPER(dp2.cat_comercial) = 'SOFTYS')"

_SOFTYS_GRUPO_CASE = """
    CASE
        WHEN dp.clase_descripcion = 'BABYSEC' AND UPPER(dp.producto_nombre) LIKE '%%PACKETON%%' THEN 'Pañales Packeton'
        WHEN dp.clase_descripcion = 'BABYSEC'    THEN 'Pañales Babysec'
        WHEN dp.clase_descripcion = 'COTIDIAN'   THEN 'Pañales para Adultos'
        WHEN dp.clase_descripcion = 'LADYSOFT'   THEN 'Toallas Femeninas'
        WHEN dp.clase_descripcion = 'PAPEL NOVA' THEN 'Toallas de Papel'
        WHEN UPPER(dp.producto_nombre) LIKE '%%PAPEL HIG%%' OR UPPER(dp.producto_nombre) LIKE '%%PAP. HIG%%' OR UPPER(dp.producto_nombre) LIKE '%%PAP.HIG%%' THEN 'Papel Higiénico'
        WHEN UPPER(dp.producto_nombre) LIKE '%%PANUELO%%'   THEN 'Pañuelos'
        WHEN UPPER(dp.producto_nombre) LIKE '%%SERVILLET%%'  THEN 'Servilletas'
        ELSE 'Otros'
    END
"""

_SOFTYS_GRUPO_PATTERN: dict[str, dict] = {
    'Pañales':              {'clase': 'BABYSEC'},
    'Pañales Babysec':      {'clase': 'BABYSEC', 'name_not_ilike': '%packeton%'},
    'Pañales Packeton':     {'clase': 'BABYSEC', 'name_ilike': '%packeton%'},
    'Pañales para Adultos': {'clase': 'COTIDIAN'},
    'Papel Higiénico':      {'name_like_any': ['%PAPEL HIG%', '%PAP. HIG%', '%PAP.HIG%', '%CONFORT%']},
    'Toallas Femeninas':    {'clase': 'LADYSOFT'},
    'Pañuelos':             {'name_ilike': '%panuelo%'},
    'Toallas de Papel':     {'clase': 'PAPEL NOVA'},
    'Servilletas':          {'name_ilike': '%servillet%'},
}


def _grupo_sql_cond(grupo: str, params: list, field: str = "dp.producto_nombre") -> str:
    """Filtra por grupo Softys. clase_descripcion para BABYSEC/COTIDIAN/LADYSOFT/PAPEL NOVA;
    nombre de producto (LIKE) para Papel Higienico, Panuelos y Servilletas (todos bajo clase ELITE)."""
    if not grupo:
        return ""
    cfg = _SOFTYS_GRUPO_PATTERN.get(grupo, {})
    if not cfg:
        # Grupo no reconocido: devolver condicion imposible para no filtrar datos incorrectos
        logger.warning("_grupo_sql_cond: grupo no reconocido '%s' — no se mostraran resultados", grupo)
        return "AND false"
    parts = []
    if 'clase' in cfg:
        parts.append("dp.clase_descripcion = %s")
        params.append(cfg['clase'])
    if 'name_ilike' in cfg:
        parts.append("UPPER(dp.producto_nombre) LIKE UPPER(%s)")
        params.append(cfg['name_ilike'])
    if 'name_not_ilike' in cfg:
        parts.append("UPPER(dp.producto_nombre) NOT LIKE UPPER(%s)")
        params.append(cfg['name_not_ilike'])
    if 'name_like_any' in cfg:
        patterns = cfg['name_like_any']
        clauses = ' OR '.join(['UPPER(dp.producto_nombre) LIKE %s'] * len(patterns))
        parts.append(f'({clauses})')
        for p in patterns:
            params.append(p.upper())
    if not parts:
        return ""
    return "AND " + " AND ".join(parts)


def _ppto_softys_by_canal(anho, mes, ciudad_cond, canal_filter='', params_extra=None):
    """Presupuesto por canal filtrando solo productos Softys."""
    sql = f"""
        SELECT dv.canal_rrhh AS canal, COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto
        FROM dw.fact_presupuesto fp
        JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
        JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
        WHERE fp.anho = %s AND fp.mes = %s
          AND ({ciudad_cond})
          AND dv.canal_rrhh IS NOT NULL
          {canal_filter}
          AND {_SOFTYS_COND}
          AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
        GROUP BY dv.canal_rrhh
    """
    base = [anho, mes] + (params_extra or []) + [anho, mes]
    try:
        _, rows = _run_dw_query(sql, base)
        return {r['canal']: float(r['presupuesto'] or 0) for r in rows}
    except Exception:
        return {}


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_kpis(request):
    """KPI cards por canal para productos Softys. Params: regional, canal, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        dia  = _safe_int(request.GET.get('dia'),  0)
        dia_cond = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params      = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT
                dv.canal_rrhh                            AS canal,
                COALESCE(SUM(fv.venta_neta), 0)          AS avance,
                COUNT(DISTINCT fv.numero_venta)           AS pedidos,
                COUNT(DISTINCT fv.cliente_sk)             AS clientes
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL AND ({ciudad_cond})
              {canal_cond} {dia_cond}
              AND {_SOFTYS_COND}
            GROUP BY dv.canal_rrhh ORDER BY avance DESC
        """
        _, ventas_rows = _run_dw_query(sql, params)
        # Max day with data (for the date picker)
        sql_maxdia = f"""
            SELECT MAX(df.dia_numero) AS max_dia
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dv.canal_rrhh IS NOT NULL AND ({ciudad_cond})
              AND {_SOFTYS_COND}
        """
        _, maxdia_rows = _run_dw_query(sql_maxdia, [anho, mes])
        max_dia = int(maxdia_rows[0]['max_dia'] or 0) if maxdia_rows else 0
        canal_filter_ppto = "AND dv.canal_rrhh = %s" if canal else ""
        ppto_map = _ppto_softys_by_canal(anho, mes, ciudad_cond, canal_filter_ppto, [canal] if canal else None)

        # Universe: active clients from dim_cliente_dual per canal (real client portfolio)
        # Uses ruta prefix to filter by regional â€" no join to dim_planificacion so no clients excluded
        # canal_rrhh (dim_vendedor) â†’ cd.canal (dim_cliente_dual) mapping:
        #   WHS, WHS-LIC, WHS-* â†’ WHS  |  DTS, DTS-* â†’ DTS  |  others: exact match
        def _rrhh_to_dual(rrhh: str) -> str:
            if rrhh.startswith('WHS'): return 'WHS'
            if rrhh.startswith('DTS'): return 'DTS'
            return rrhh

        ruta_cond = _ruta_regional_cond(regional)
        # When filtering by a specific canal, translate canal_rrhh â†’ cd.canal
        dual_canal_val = _rrhh_to_dual(canal) if canal else None
        canal_dual     = "AND cd.canal = %s" if dual_canal_val else ""
        sql_universo = f"""
            SELECT cd.canal, COUNT(DISTINCT cd.id_cliente) AS universo
            FROM dual.dim_cliente_dual cd
            WHERE cd.es_actual = true
              AND ({ruta_cond})
              {canal_dual}
            GROUP BY cd.canal
        """
        _, universo_rows = _run_dw_query(sql_universo, ([dual_canal_val] if dual_canal_val else []))
        universo_raw = {r['canal']: int(r['universo'] or 0) for r in universo_rows}

        # Build lookup: canal_rrhh â†’ universe size (using the rrhhâ†’dual mapping)
        def _universo_for(rrhh: str) -> int:
            return universo_raw.get(_rrhh_to_dual(rrhh), 0)

        result = []
        for row in ventas_rows:
            ppto     = ppto_map.get(row['canal'], 0)
            universo = _universo_for(row['canal'])
            clientes = int(row['clientes'] or 0)
            cobertura = round(clientes / universo * 100, 1) if universo > 0 else None
            result.append({
                **row, 'presupuesto': ppto,
                'porcentaje': round(float(row['avance']) / ppto * 100, 1) if ppto > 0 else None,
                'universo': universo, 'cobertura': cobertura,
            })

        universo_total  = sum(_universo_for(r['canal']) for r in ventas_rows)
        clientes_total  = sum(int(r['clientes'] or 0)   for r in ventas_rows)
        cobertura_total = round(clientes_total / universo_total * 100, 1) if universo_total > 0 else None

        return JsonResponse({
            'success': True, 'data': result, 'max_dia': max_dia,
            'universo_total': universo_total, 'cobertura_total': cobertura_total,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_tendencia(request):
    """Tendencia diaria Softys para canal+regional. Params: regional, canal, anho, mes."""
    try:
        import calendar
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        hoy  = datetime.now().date()
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond2 = _regional_filter(regional, campo='dv2.ciudad')
        canal_cond2  = "AND dv2.canal_rrhh = %s" if canal else ""
        params_avance = ([canal] if canal else []) + [anho, mes]

        sql_avance = f"""
            WITH dias AS (
                SELECT df.dia_numero,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_dia
                FROM dw.dim_fecha df
                LEFT JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_vendedor dv2
                        WHERE dv2.vendedor_sk = fv.vendedor_sk
                          AND ({ciudad_cond2})
                          {canal_cond2}
                    )
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_producto dp2
                        WHERE dp2.producto_sk = fv.producto_sk
                          AND {_SOFTYS_COND_DP2}
                    )
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND df.fecha_completa <= CURRENT_DATE
                GROUP BY df.dia_numero
            )
            SELECT dia_numero AS dia, venta_dia,
                   SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
            FROM dias ORDER BY dia_numero
        """
        _, avance_rows = _run_dw_query(sql_avance, params_avance)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        ppto_map = _ppto_softys_by_canal(anho, mes, ciudad_cond, canal_cond, [canal] if canal else None)
        presupuesto_mes = sum(ppto_map.values())

        dias_en_mes        = calendar.monthrange(anho, mes)[1]
        ppto_diario        = presupuesto_mes / dias_en_mes if dias_en_mes > 0 else 0
        es_periodo_actual  = (anho == hoy.year and mes == hoy.month)
        dias_transcurridos = hoy.day if es_periodo_actual else dias_en_mes
        avance_por_dia     = {int(r['dia']): r['avance_acumulado'] for r in avance_rows}
        avance_total       = float(avance_rows[-1]['avance_acumulado']) if avance_rows else 0.0
        tasa_diaria        = avance_total / dias_transcurridos if dias_transcurridos > 0 else 0

        result = []
        for dia in range(1, dias_en_mes + 1):
            proyeccion = None
            if es_periodo_actual and dia > dias_transcurridos:
                proyeccion = round(avance_total + tasa_diaria * (dia - dias_transcurridos), 2)
            result.append({
                'dia':                   dia,
                'avance_acumulado':      avance_por_dia.get(dia),
                'presupuesto_acumulado': round(ppto_diario * dia, 2) if presupuesto_mes > 0 else None,
                'proyeccion_acumulada':  proyeccion,
            })
        return JsonResponse({'success': True, 'data': result,
                             'presupuesto_mes': presupuesto_mes,
                             'es_periodo_actual': es_periodo_actual})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_por_categoria(request):
    """Ventas Softys por categorÃ­a para canal+regional. Params: regional, canal, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params      = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT
                {_CATEGORIA_CASE}                                AS categoria,
                COALESCE(SUM(fv.venta_neta), 0)                  AS avance,
                COALESCE(SUM(fv.cantidad), 0)                    AS cantidad,
                COUNT(DISTINCT fv.producto_sk)                   AS productos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond}
              AND dp.linea IS NOT NULL AND dp.linea != 'SIN LINEA'
              AND {_SOFTYS_COND}
            GROUP BY {_CATEGORIA_CASE}
            ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, params)

        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT
                    {_CATEGORIA_CASE}                                        AS categoria,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)             AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond}
                  AND dp.grupo_descripcion != 'EXHIBIDORES'
                  AND dp.grupo_descripcion IS NOT NULL
                  AND {_SOFTYS_COND}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY {_CATEGORIA_CASE}
            """
            _, ppto_rows = _run_dw_query(sql_ppto, params + [anho, mes])
            ppto_map = {r['categoria']: float(r['presupuesto'] or 0) for r in ppto_rows}
        except Exception:
            pass

        result = []
        for row in rows:
            cat  = row['categoria']
            av   = float(row['avance'] or 0)
            ppto = ppto_map.get(cat, 0)
            result.append({
                'categoria':   cat,
                'avance':      av,
                'cantidad':    int(row['cantidad'] or 0),
                'productos':   int(row['productos'] or 0),
                'presupuesto': ppto,
                'porcentaje':  round(av / ppto * 100, 1) if ppto > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_por_sku(request):
    """Top SKUs Softys para canal+grupo+regional. Params: regional, canal, grupo, anho, mes."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()
        grupo = _safe_str(request.GET.get('grupo', ''))
        anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err   = _validate_anho_mes(anho, mes)
        if err: return err
        dia  = _safe_int(request.GET.get('dia'),  0)
        dia_cond = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""
        limit = min(_safe_int(request.GET.get('limit'), 500), 1000)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""

        # sub_params: filtros de la subconsulta fact_ventas (período + regional + canal)
        sub_params = [anho, mes]
        if canal:
            sub_params.append(canal)

        # cat_params: filtros de dim_producto (grupo/subcategoría)
        cat_params: list = []
        cat_cond = _grupo_sql_cond(grupo, cat_params)

        params_ventas = sub_params + cat_params + [limit]

        sql = f"""
            SELECT
                dp.producto_codigo_erp                           AS codigo,
                dp.producto_nombre                               AS producto,
                COALESCE(dp.linea, '')                           AS categoria,
                COALESCE(dp.subgrupo_descripcion, '')            AS subgrupo,
                COALESCE(SUM(v.cantidad), 0)                     AS cantidad,
                COALESCE(SUM(v.venta_neta), 0)                   AS venta_neta,
                COUNT(DISTINCT v.cliente_sk)                     AS clientes
            FROM dw.dim_producto dp
            LEFT JOIN (
                SELECT fv.producto_sk, fv.cantidad, fv.venta_neta, fv.cliente_sk
                FROM dw.fact_ventas fv
                JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
                JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND ({ciudad_cond}) {canal_cond} {dia_cond}
            ) v ON v.producto_sk = dp.producto_sk
            WHERE dp.es_producto_actual = true
              AND {_SOFTYS_COND} {cat_cond}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre,
                     dp.linea, dp.subgrupo_descripcion
            ORDER BY venta_neta DESC
            LIMIT %s
        """
        _, rows = _run_dw_query(sql, params_ventas)

        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT dp.producto_codigo_erp                        AS codigo,
                       COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto,
                       COALESCE(SUM(fp.cantidad_presupuestada), 0)   AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {cat_cond}
                  AND {_SOFTYS_COND}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.producto_codigo_erp
            """
            _, ppto_rows = _run_dw_query(sql_ppto, sub_params + cat_params + [anho, mes])
            ppto_map = {r['codigo']: r for r in ppto_rows}
        except Exception:
            pass

        result = []
        for row in rows:
            vn       = float(row['venta_neta'] or 0)
            cant     = int(row['cantidad'] or 0)
            p        = ppto_map.get(row['codigo'], {})
            ppto_bs  = float(p.get('presupuesto',     0) or 0)
            ppto_uds = float(p.get('presupuesto_uds', 0) or 0)
            result.append({
                **row,
                'venta_neta':      vn,
                'presupuesto':     ppto_bs,
                'presupuesto_uds': int(ppto_uds),
                'porcentaje':      round(vn   / ppto_bs  * 100, 1) if ppto_bs  > 0 else None,
                'porcentaje_uds':  round(cant / ppto_uds * 100, 1) if ppto_uds > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_por_regional(request):
    """Ventas Softys desglosadas por regional (para la vista Nacional). Params: anho, mes, canal."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            canal = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            canal = _safe_str(request.GET.get('canal', ''))
        else:
            canal = (profile.canal or '').strip()
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err
        dia  = _safe_int(request.GET.get('dia'),  0)
        dia_cond = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""

        scz  = _ciudad_case('dv.ciudad', 'santa_cruz')
        cbba = _ciudad_case('dv.ciudad', 'cochabamba')
        lpz  = _ciudad_case('dv.ciudad', 'la_paz')
        canal_cond = "AND dv.canal_rrhh = %s" if canal else ""

        sql = f"""
            SELECT
                CASE
                    WHEN {scz}  THEN 'Santa Cruz'
                    WHEN {cbba} THEN 'Cochabamba'
                    WHEN {lpz}  THEN 'La Paz'
                    ELSE 'Otras'
                END                               AS regional,
                COALESCE(SUM(fv.venta_neta), 0)   AS avance,
                COUNT(DISTINCT fv.cliente_sk)      AS clientes,
                COUNT(DISTINCT fv.numero_venta)    AS pedidos
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              {canal_cond} {dia_cond}
              AND {_SOFTYS_COND}
            GROUP BY regional ORDER BY avance DESC
        """
        _, ventas_rows = _run_dw_query(sql, [anho, mes] + ([canal] if canal else []))

        # Budget by regional
        ppto_scz  = _ppto_softys_by_canal(anho, mes, _ciudad_case('dv.ciudad', 'santa_cruz'),  canal_cond, [canal] if canal else None)
        ppto_cbba = _ppto_softys_by_canal(anho, mes, _ciudad_case('dv.ciudad', 'cochabamba'),   canal_cond, [canal] if canal else None)
        ppto_lpz  = _ppto_softys_by_canal(anho, mes, _ciudad_case('dv.ciudad', 'la_paz'),       canal_cond, [canal] if canal else None)
        ppto_map = {
            'Santa Cruz': sum(ppto_scz.values()),
            'Cochabamba': sum(ppto_cbba.values()),
            'La Paz':     sum(ppto_lpz.values()),
        }

        result = []
        for row in ventas_rows:
            reg  = row['regional']
            av   = float(row['avance'] or 0)
            ppto = ppto_map.get(reg, 0)
            result.append({
                'regional':   reg,
                'avance':     av,
                'presupuesto': ppto,
                'clientes':   int(row['clientes'] or 0),
                'pedidos':    int(row['pedidos'] or 0),
                'porcentaje': round(av / ppto * 100, 1) if ppto > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€ Helpers histÃ³rico â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_MESES_SHORT = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def _periodos_anteriores(anho, mes, n):
    """Lista de (anho, mes) de los Ãºltimos n meses en orden cronolÃ³gico."""
    periodos = []
    a, m = anho, mes
    for _ in range(n):
        periodos.append((a, m))
        m -= 1
        if m == 0:
            m = 12
            a -= 1
    return list(reversed(periodos))

def _mes_label(anho, mes):
    return f"{_MESES_SHORT[mes - 1]} {str(anho)[2:]}"

def _month_keys_placeholders(periodos):
    keys = [a * 100 + m for a, m in periodos]
    placeholders = ', '.join(['%s'] * len(keys))
    return keys, placeholders

def _auth_regional_canal(request):
    """Extrae regional_key y canal segÃºn rol. Retorna (regional_key, canal)."""
    is_admin = _is_admin(request.user)
    profile  = _get_or_create_profile(request.user)
    cargo    = (profile.cargo or '').strip()
    if is_admin:
        regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal    = _safe_str(request.GET.get('canal', ''))
    elif cargo == 'Gerente Regional':
        regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        canal    = _safe_str(request.GET.get('canal', ''))
    elif cargo == 'Proveedor':
        regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        canal    = (profile.canal or '').strip()
    else:
        regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
        canal    = (profile.canal or '').strip()
    return regional, canal


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_canales_por_grupo(request):
    """Ventas Softys agrupadas por lÃ­nea de producto (PaÃ±ales, Papel HigiÃ©nico, etc.). Params: regional, canal, anho, mes, dia."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        dia  = _safe_int(request.GET.get('dia'),  0)
        err  = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        dia_cond    = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""
        params      = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT
                {_SOFTYS_GRUPO_CASE}            AS grupo,
                COALESCE(SUM(fv.venta_neta), 0) AS avance,
                COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                COUNT(DISTINCT fv.cliente_sk)   AS clientes
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {dia_cond}
              AND {_SOFTYS_COND}
            GROUP BY {_SOFTYS_GRUPO_CASE}
            ORDER BY avance DESC
        """
        _, rows = _run_dw_query(sql, params)

        ppto_map = {}
        try:
            sql_ppto = f"""
                SELECT
                    {_SOFTYS_GRUPO_CASE}                                AS grupo,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)        AS presupuesto
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond}
                  AND {_SOFTYS_COND}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY {_SOFTYS_GRUPO_CASE}
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes] + ([canal] if canal else []) + [anho, mes])
            ppto_map = {r['grupo']: float(r['presupuesto'] or 0) for r in ppto_rows}
        except Exception:
            pass

        _ORDEN_G = ["PaÃ±ales", "PaÃ±ales para Adultos", "Papel HigiÃ©nico", "Toallas Femeninas", "PaÃ±uelos", "Toallas de Papel", "Otros"]
        result = []
        for row in rows:
            g    = row['grupo']
            av   = float(row['avance'] or 0)
            ppto = ppto_map.get(g, 0)
            result.append({
                'grupo':       g,
                'avance':      av,
                'presupuesto': ppto,
                'cantidad':    int(row['cantidad'] or 0),
                'clientes':    int(row['clientes'] or 0),
                'porcentaje':  round(av / ppto * 100, 1) if ppto > 0 else None,
            })
        result.sort(key=lambda x: _ORDEN_G.index(x['grupo']) if x['grupo'] in _ORDEN_G else 99)
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_sku_tendencia(request):
    """Tendencia diaria de ventas para un SKU Softys especÃ­fico. Params: regional, canal, anho, mes, sku."""
    try:
        import calendar
        regional, canal = _auth_regional_canal(request)
        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        sku  = _safe_str(request.GET.get('sku', ''))
        if not sku:
            return JsonResponse({'success': False, 'error': 'ParÃ¡metro sku requerido'}, status=400)
        err = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond  = _regional_filter(regional)
        ciudad_cond2 = _regional_filter(regional, campo='dv2.ciudad')
        canal_cond   = "AND dv.canal_rrhh = %s" if canal else ""
        canal_cond2  = "AND dv2.canal_rrhh = %s" if canal else ""

        params_avance = ([canal] if canal else []) + [sku, anho, mes]

        sql_avance = f"""
            WITH dias AS (
                SELECT df.dia_numero,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_dia
                FROM dw.dim_fecha df
                LEFT JOIN dw.fact_ventas fv ON fv.fecha_sk = df.fecha_sk
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_vendedor dv2
                        WHERE dv2.vendedor_sk = fv.vendedor_sk
                          AND ({ciudad_cond2})
                          {canal_cond2}
                    )
                    AND EXISTS (
                        SELECT 1 FROM dw.dim_producto dp2
                        WHERE dp2.producto_sk = fv.producto_sk
                          AND {_SOFTYS_COND_DP2}
                          AND dp2.producto_codigo_erp = %s
                    )
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND df.fecha_completa <= CURRENT_DATE
                GROUP BY df.dia_numero
            )
            SELECT dia_numero AS dia, venta_dia,
                   SUM(venta_dia) OVER (ORDER BY dia_numero) AS avance_acumulado
            FROM dias ORDER BY dia_numero
        """
        _, avance_rows = _run_dw_query(sql_avance, params_avance)

        ppto_total = 0
        producto_nombre = ''
        try:
            sql_info = f"""
                SELECT COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto,
                       MAX(dp.producto_nombre)                        AS nombre
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond}
                  AND dp.producto_codigo_erp = %s
                  AND {_SOFTYS_COND}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
            """
            _, info_rows = _run_dw_query(sql_info, [anho, mes] + ([canal] if canal else []) + [sku, anho, mes])
            if info_rows:
                ppto_total = float(info_rows[0]['presupuesto'] or 0)
                producto_nombre = info_rows[0]['nombre'] or ''
        except Exception:
            pass

        if not producto_nombre:
            try:
                _, nombre_rows = _run_dw_query(
                    "SELECT producto_nombre FROM dw.dim_producto WHERE producto_codigo_erp = %s LIMIT 1", [sku])
                if nombre_rows:
                    producto_nombre = nombre_rows[0]['producto_nombre'] or ''
            except Exception:
                pass

        hoy = datetime.now().date()
        dias_en_mes = calendar.monthrange(anho, mes)[1]
        es_periodo_actual = (anho == hoy.year and mes == hoy.month)
        dias_transcurridos = hoy.day if es_periodo_actual else dias_en_mes
        ppto_diario = ppto_total / dias_en_mes if dias_en_mes > 0 and ppto_total > 0 else 0
        avance_por_dia = {int(r['dia']): float(r['avance_acumulado'] or 0) for r in avance_rows}
        avance_total = float(avance_rows[-1]['avance_acumulado']) if avance_rows else 0.0
        tasa_diaria = avance_total / dias_transcurridos if dias_transcurridos > 0 else 0

        result = []
        for d in range(1, dias_en_mes + 1):
            proyeccion = None
            if es_periodo_actual and d > dias_transcurridos:
                proyeccion = round(avance_total + tasa_diaria * (d - dias_transcurridos), 2)
            result.append({
                'dia':                   d,
                'avance_acumulado':      avance_por_dia.get(d),
                'presupuesto_acumulado': round(ppto_diario * d, 2) if ppto_diario > 0 else None,
                'proyeccion_acumulada':  proyeccion,
            })
        return JsonResponse({
            'success': True, 'data': result,
            'producto_nombre': producto_nombre,
            'presupuesto_total': ppto_total,
            'es_periodo_actual': es_periodo_actual,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_historico_canales(request):
    """Ventas Softys por canal mes a mes. Params: regional, canal, anho, mes, meses."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho    = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes     = _safe_int(request.GET.get('mes'),  datetime.now().month)
        meses_n = min(_safe_int(request.GET.get('meses'), 6), 12)
        modo    = request.GET.get('modo', 'completo')
        dia_ref = _safe_int(request.GET.get('dia_ref'), 0)
        dia_hist_cond = f"AND df.dia_numero <= {dia_ref}" if modo in ('mismo_rango', 'personalizado') and 1 <= dia_ref <= 31 else ""
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        periodos = _periodos_anteriores(anho, mes, meses_n)
        keys, placeholders = _month_keys_placeholders(periodos)
        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params = keys + ([canal] if canal else [])

        sql = f"""
            SELECT df.anho, df.mes_numero,
                   dv.canal_rrhh                   AS nombre,
                   COALESCE(SUM(fv.venta_neta), 0) AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE (df.anho * 100 + df.mes_numero) IN ({placeholders})
              AND ({ciudad_cond}) {canal_cond}
              AND dv.canal_rrhh IS NOT NULL
              {dia_hist_cond}
              AND {_SOFTYS_COND}
            GROUP BY df.anho, df.mes_numero, dv.canal_rrhh
            ORDER BY df.anho, df.mes_numero, avance DESC
        """
        _, rows = _run_dw_query(sql, params)

        # Pivot
        nombres = list(dict.fromkeys(r['nombre'] for r in rows))
        grid = {n: {p: 0.0 for p in periodos} for n in nombres}
        for r in rows:
            grid[r['nombre']][(r['anho'], r['mes_numero'])] = float(r['avance'] or 0)

        periodos_out = [{'anho': a, 'mes': m, 'label': _mes_label(a, m)} for a, m in periodos]
        series = [{'nombre': n, 'valores': [grid[n][p] for p in periodos]} for n in nombres]
        return JsonResponse({'success': True, 'periodos': periodos_out, 'series': series})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_historico_grupos(request):
    """Ventas Softys por lÃ­nea de producto mes a mes. Params: regional, canal, anho, mes, meses, modo, dia_ref."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho    = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes     = _safe_int(request.GET.get('mes'),  datetime.now().month)
        meses_n = min(_safe_int(request.GET.get('meses'), 6), 12)
        modo    = request.GET.get('modo', 'completo')
        dia_ref = _safe_int(request.GET.get('dia_ref'), 0)
        dia_hist_cond = f"AND df.dia_numero <= {dia_ref}" if modo in ('mismo_rango', 'personalizado') and 1 <= dia_ref <= 31 else ""
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        periodos = _periodos_anteriores(anho, mes, meses_n)
        keys, placeholders = _month_keys_placeholders(periodos)
        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params = keys + ([canal] if canal else [])

        # Agrupa por nombre de producto y categoriza en Python para evitar LIKE con % en GROUP BY
        sql = f"""
            SELECT df.anho, df.mes_numero,
                   dp.producto_nombre              AS producto,
                   COALESCE(SUM(fv.venta_neta), 0) AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE (df.anho * 100 + df.mes_numero) IN ({placeholders})
              AND ({ciudad_cond}) {canal_cond}
              {dia_hist_cond}
              AND {_SOFTYS_COND}
            GROUP BY df.anho, df.mes_numero, dp.producto_nombre
            ORDER BY df.anho, df.mes_numero
        """
        _, rows = _run_dw_query(sql, params)

        def _grupo(nombre):
            n = (nombre or '').upper()
            if 'PAN BABYSEC'         in n: return 'PaÃ±ales'
            if 'PAN COTIDIAN'        in n: return 'PaÃ±ales para Adultos'
            if 'PAPEL HIG.'          in n: return 'Papel HigiÃ©nico'
            if 'LADYSOFT'            in n: return 'Toallas Femeninas'
            if 'PANUELO ELITE'       in n: return 'PaÃ±uelos'
            if 'TOALLAS DE PAPEL NOVA' in n: return 'Toallas de Papel'
            return None  # excluir "Otros" de la vista de categorÃ­as

        _ORDEN_G = ['PaÃ±ales', 'PaÃ±ales para Adultos', 'Papel HigiÃ©nico', 'Toallas Femeninas', 'PaÃ±uelos', 'Toallas de Papel']
        grid = {g: {p: 0.0 for p in periodos} for g in _ORDEN_G}

        for r in rows:
            g = _grupo(r['producto'])
            if g and g in grid:
                key = (int(r['anho']), int(r['mes_numero']))
                if key in grid[g]:
                    grid[g][key] += float(r['avance'] or 0)

        periodos_out = [{'anho': a, 'mes': m, 'label': _mes_label(a, m)} for a, m in periodos]
        series = [
            {'nombre': g, 'valores': [grid[g][p] for p in periodos]}
            for g in _ORDEN_G
            if any(v > 0 for v in grid[g].values())
        ]
        return JsonResponse({'success': True, 'periodos': periodos_out, 'series': series})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_historico_skus(request):
    """Top 10 SKUs Softys mes a mes. Params: regional, canal, anho, mes, meses, grupo."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho    = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes     = _safe_int(request.GET.get('mes'),  datetime.now().month)
        meses_n = min(_safe_int(request.GET.get('meses'), 6), 12)
        modo    = request.GET.get('modo', 'completo')
        dia_ref = _safe_int(request.GET.get('dia_ref'), 0)
        dia_hist_cond = f"AND df.dia_numero <= {dia_ref}" if modo in ('mismo_rango', 'personalizado') and 1 <= dia_ref <= 31 else ""
        grupo   = _safe_str(request.GET.get('grupo', ''))
        sku_codigo = _safe_str(request.GET.get('sku', ''))
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        periodos = _periodos_anteriores(anho, mes, meses_n)
        keys, placeholders = _month_keys_placeholders(periodos)
        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params = keys + ([canal] if canal else [])

        grupo_cond = _grupo_sql_cond(grupo, params)

        sku_cond = ""
        if sku_codigo:
            sku_cond = "AND dp.producto_codigo_erp = %s"
            params.append(sku_codigo)

        sql = f"""
            SELECT df.anho, df.mes_numero,
                   dp.producto_codigo_erp          AS codigo,
                   dp.producto_nombre              AS producto,
                   COALESCE(SUM(fv.venta_neta), 0) AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE (df.anho * 100 + df.mes_numero) IN ({placeholders})
              AND ({ciudad_cond}) {canal_cond} {grupo_cond} {sku_cond}
              {dia_hist_cond}
              AND {_SOFTYS_COND}
            GROUP BY df.anho, df.mes_numero, dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY df.anho, df.mes_numero, avance DESC
        """
        _, rows = _run_dw_query(sql, params)

        # Acumular totales por SKU para rankear top 10
        from collections import defaultdict
        sku_total = defaultdict(float)
        sku_nombre = {}
        sku_grid = defaultdict(lambda: {p: 0.0 for p in periodos})
        for r in rows:
            cod = r['codigo']
            av  = float(r['avance'] or 0)
            sku_total[cod] += av
            sku_nombre[cod] = r['producto']
            sku_grid[cod][(r['anho'], r['mes_numero'])] = av

        top10 = sorted(sku_total, key=lambda c: -sku_total[c])[:10]
        periodos_out = [{'anho': a, 'mes': m, 'label': _mes_label(a, m)} for a, m in periodos]
        skus_out = [
            {'codigo': c, 'producto': sku_nombre[c], 'total': round(sku_total[c], 2),
             'valores': [sku_grid[c][p] for p in periodos]}
            for c in top10
        ]
        return JsonResponse({'success': True, 'periodos': periodos_out, 'skus': skus_out})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  SOFTYS â€" CLIENTES POR VENDEDOR
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_vendedores(request):
    """Lista vendedores Softys del periodo. Params: regional, canal, anho, mes, dia, grupo."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)
        dia   = _safe_int(request.GET.get('dia'),  0)
        grupo = _safe_str(request.GET.get('grupo', ''))
        err   = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        dia_cond    = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""
        params      = [anho, mes] + ([canal] if canal else [])

        grupo_cond = _grupo_sql_cond(grupo, params)

        sql = f"""
            SELECT dv.vendedor_nombre,
                   COUNT(DISTINCT fv.cliente_sk)   AS clientes,
                   COALESCE(SUM(fv.venta_neta), 0) AS total
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond}
              {dia_cond} {grupo_cond}
              AND {_SOFTYS_COND}
              AND dv.vendedor_nombre IS NOT NULL
            GROUP BY dv.vendedor_nombre
            ORDER BY total DESC
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': [
            {'vendedor': r['vendedor_nombre'], 'clientes': int(r['clientes'] or 0), 'total': float(r['total'] or 0)}
            for r in rows
        ]})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_clientes_semana(request):
    """Clientes de un vendedor con ventas Softys por semana. Params: regional, canal, anho, mes, dia, vendedor, grupo."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho     = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes      = _safe_int(request.GET.get('mes'),  datetime.now().month)
        dia      = _safe_int(request.GET.get('dia'),  0)
        vendedor = _safe_str(request.GET.get('vendedor', ''))
        grupo    = _safe_str(request.GET.get('grupo', ''))
        err      = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond   = _regional_filter(regional)
        canal_cond    = "AND dv.canal_rrhh = %s" if canal else ""
        vendedor_cond = "AND dv.vendedor_nombre = %s" if vendedor else ""
        dia_cond      = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""
        params        = [anho, mes] + ([canal] if canal else []) + ([vendedor] if vendedor else [])

        grupo_cond = _grupo_sql_cond(grupo, params)

        sql = f"""
            SELECT
                dc.cliente_codigo_erp                                                               AS codigo,
                COALESCE(dc.cliente_nombre, dc.cliente_codigo_erp)                                 AS nombre,
                COALESCE(SUM(CASE WHEN df.dia_numero BETWEEN  1 AND  7 THEN fv.venta_neta END), 0) AS sem1,
                COALESCE(SUM(CASE WHEN df.dia_numero BETWEEN  8 AND 14 THEN fv.venta_neta END), 0) AS sem2,
                COALESCE(SUM(CASE WHEN df.dia_numero BETWEEN 15 AND 21 THEN fv.venta_neta END), 0) AS sem3,
                COALESCE(SUM(CASE WHEN df.dia_numero BETWEEN 22 AND 28 THEN fv.venta_neta END), 0) AS sem4,
                COALESCE(SUM(CASE WHEN df.dia_numero >= 29             THEN fv.venta_neta END), 0) AS sem5,
                COALESCE(SUM(fv.venta_neta), 0)                                                    AS total
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_cliente  dc ON fv.cliente_sk  = dc.cliente_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {vendedor_cond}
              {dia_cond} {grupo_cond}
              AND {_SOFTYS_COND}
            GROUP BY dc.cliente_codigo_erp, dc.cliente_nombre
            ORDER BY total DESC
        """
        _, rows = _run_dw_query(sql, params)

        clientes = [
            {
                'codigo': r['codigo'], 'nombre': r['nombre'],
                'sem1': float(r['sem1'] or 0), 'sem2': float(r['sem2'] or 0),
                'sem3': float(r['sem3'] or 0), 'sem4': float(r['sem4'] or 0),
                'sem5': float(r['sem5'] or 0), 'total': float(r['total'] or 0),
            }
            for r in rows
        ]
        tiene_sem5 = any(c['sem5'] > 0 for c in clientes)
        totales = {
            'sem1': sum(c['sem1'] for c in clientes), 'sem2': sum(c['sem2'] for c in clientes),
            'sem3': sum(c['sem3'] for c in clientes), 'sem4': sum(c['sem4'] for c in clientes),
            'sem5': sum(c['sem5'] for c in clientes), 'total': sum(c['total'] for c in clientes),
        }
        return JsonResponse({'success': True, 'clientes': clientes, 'totales': totales, 'tiene_sem5': tiene_sem5})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_sku_por_cliente(request):
    """SKUs Softys de un cliente especÃ­fico. Params: regional, canal, anho, mes, dia, cliente, semana (1-5|0=todo), grupo, vendedor, meses (>0 = rango histÃ³rico)."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho           = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes            = _safe_int(request.GET.get('mes'),  datetime.now().month)
        dia            = _safe_int(request.GET.get('dia'),  0)
        meses          = min(_safe_int(request.GET.get('meses'), 0), 24)
        cliente_codigo = _safe_str(request.GET.get('cliente', ''))
        semana         = _safe_int(request.GET.get('semana'), 0)
        grupo          = _safe_str(request.GET.get('grupo', ''))
        vendedor       = _safe_str(request.GET.get('vendedor', ''))
        err            = _validate_anho_mes(anho, mes)
        if err: return err
        if not cliente_codigo:
            return JsonResponse({'success': False, 'error': 'ParÃ¡metro cliente requerido'}, status=400)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond   = _regional_filter(regional)
        canal_cond    = "AND dv.canal_rrhh = %s" if canal else ""
        vendedor_cond = "AND dv.vendedor_nombre = %s" if vendedor else ""

        if meses > 0:
            # Range mode: aggregate across last N months
            periodos = _periodos_anteriores(anho, mes, meses)
            keys, placeholders = _month_keys_placeholders(periodos)
            date_cond = f"(df.anho * 100 + df.mes_numero) IN ({placeholders})"
            params = keys + [cliente_codigo] + ([canal] if canal else []) + ([vendedor] if vendedor else [])
            dia_cond    = ""
            semana_cond = ""
        else:
            # Single month mode
            date_cond = "df.anho = %s AND df.mes_numero = %s"
            params    = [anho, mes, cliente_codigo] + ([canal] if canal else []) + ([vendedor] if vendedor else [])
            dia_cond  = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""
            _SEM_RANGES = {1: (1, 7), 2: (8, 14), 3: (15, 21), 4: (22, 28), 5: (29, 31)}
            if semana in _SEM_RANGES:
                s, e = _SEM_RANGES[semana]
                semana_cond = f"AND df.dia_numero BETWEEN {s} AND {e}"
            else:
                semana_cond = ""

        grupo_cond = _grupo_sql_cond(grupo, params)

        sql = f"""
            SELECT
                dp.producto_codigo_erp              AS codigo,
                dp.producto_nombre                  AS producto,
                COALESCE(SUM(fv.cantidad), 0)       AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0)     AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_cliente  dc ON fv.cliente_sk  = dc.cliente_sk
            WHERE {date_cond}
              AND dc.cliente_codigo_erp = %s
              AND ({ciudad_cond}) {canal_cond} {vendedor_cond}
              {dia_cond} {semana_cond} {grupo_cond}
              AND {_SOFTYS_COND}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY venta_neta DESC
        """
        _, rows = _run_dw_query(sql, params)

        nombre_sql = "SELECT COALESCE(cliente_nombre, cliente_codigo_erp) AS nombre FROM dw.dim_cliente WHERE cliente_codigo_erp = %s LIMIT 1"
        _, nombre_rows = _run_dw_query(nombre_sql, [cliente_codigo])
        cliente_nombre = nombre_rows[0]['nombre'] if nombre_rows else cliente_codigo

        skus = [
            {'codigo': r['codigo'], 'producto': r['producto'],
             'cantidad': float(r['cantidad'] or 0), 'venta_neta': float(r['venta_neta'] or 0)}
            for r in rows
        ]
        return JsonResponse({
            'success': True, 'cliente_nombre': cliente_nombre, 'semana': semana,
            'skus': skus, 'total_uds': sum(s['cantidad'] for s in skus),
            'total_bs': sum(s['venta_neta'] for s in skus),
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_clientes_mes(request):
    """Clientes Softys con ventas por mes (comparativo). Params: regional, canal, anho, mes, meses, modo, dia_ref, vendedor, grupo."""
    try:
        regional, canal = _auth_regional_canal(request)
        anho    = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes     = _safe_int(request.GET.get('mes'),  datetime.now().month)
        meses_n = min(_safe_int(request.GET.get('meses'), 6), 24)
        modo    = request.GET.get('modo', 'completo')
        dia_ref = _safe_int(request.GET.get('dia_ref'), 0)
        vendedor = _safe_str(request.GET.get('vendedor', ''))
        grupo   = _safe_str(request.GET.get('grupo', ''))
        dia_hist_cond = f"AND df.dia_numero <= {dia_ref}" if modo in ('mismo_rango', 'personalizado') and 1 <= dia_ref <= 31 else ""
        err = _validate_anho_mes(anho, mes)
        if err: return err
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        periodos = _periodos_anteriores(anho, mes, meses_n)
        keys, placeholders = _month_keys_placeholders(periodos)
        ciudad_cond   = _regional_filter(regional)
        canal_cond    = "AND dv.canal_rrhh = %s" if canal else ""
        vendedor_cond = "AND dv.vendedor_nombre = %s" if vendedor else ""
        params = keys + ([canal] if canal else []) + ([vendedor] if vendedor else [])

        grupo_cond = _grupo_sql_cond(grupo, params)

        sql = f"""
            SELECT
                dc.cliente_codigo_erp                               AS codigo,
                COALESCE(dc.cliente_nombre, dc.cliente_codigo_erp)  AS nombre,
                df.anho, df.mes_numero,
                COALESCE(SUM(fv.venta_neta), 0)                     AS avance
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_cliente  dc ON fv.cliente_sk  = dc.cliente_sk
            WHERE (df.anho * 100 + df.mes_numero) IN ({placeholders})
              AND ({ciudad_cond}) {canal_cond} {vendedor_cond}
              {dia_hist_cond} {grupo_cond}
              AND {_SOFTYS_COND}
            GROUP BY dc.cliente_codigo_erp, dc.cliente_nombre, df.anho, df.mes_numero
        """
        _, rows = _run_dw_query(sql, params)

        periodos_out = [{'anho': a, 'mes': m, 'label': _mes_label(a, m)} for a, m in periodos]

        # Pivot: {(codigo, nombre): {(anho, mes): float}}
        client_grid = {}
        client_totals = {}
        for r in rows:
            key = (r['codigo'], r['nombre'])
            period_key = (int(r['anho']), int(r['mes_numero']))
            if key not in client_grid:
                client_grid[key] = {p: 0.0 for p in periodos}
                client_totals[key] = 0.0
            v = float(r['avance'] or 0)
            if period_key in client_grid[key]:
                client_grid[key][period_key] += v
                client_totals[key] += v

        sorted_clients = sorted(client_totals.items(), key=lambda x: x[1], reverse=True)[:500]
        clientes_out = [
            {
                'codigo': k[0], 'nombre': k[1],
                'total': client_totals[k],
                'valores': [client_grid[k][p] for p in periodos],
            }
            for k, _ in sorted_clients
        ]
        return JsonResponse({'success': True, 'periodos': periodos_out, 'clientes': clientes_out})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('softys', 'softys-nuevo')
def dashboard_softys_export(request):
    """Exporta detalle plano de ventas Softys (una fila por lÃ­nea de pedido).
    Params: regional, canal, grupo, anho, mes, dia.
    Pensado para descarga Excel â€" limitado a 150 000 filas."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Proveedor':
            regional = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal    = (profile.canal or '').strip()
        else:
            regional = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal    = (profile.canal or '').strip()

        grupo = _safe_str(request.GET.get('grupo', ''))
        anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err   = _validate_anho_mes(anho, mes)
        if err: return err
        dia   = _safe_int(request.GET.get('dia'), 0)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        dia_cond    = f"AND df.dia_numero <= {dia}" if 1 <= dia <= 31 else ""

        params = [anho, mes]
        if canal:
            params.append(canal)
        cat_cond = _grupo_sql_cond(grupo, params)

        sql = f"""
            SELECT
                MAKE_DATE(df.anho, df.mes_numero, df.dia_numero)          AS fecha,
                df.anho                                                    AS anho,
                df.mes_numero                                              AS mes,
                df.mes_nombre                                              AS mes_nombre,
                df.dia_numero                                              AS dia,
                CASE
                    WHEN dv.ciudad IN ('SCZ')       THEN 'Santa Cruz'
                    WHEN dv.ciudad IN ('CBA')       THEN 'Cochabamba'
                    WHEN dv.ciudad IN ('LPZ','EAL') THEN 'La Paz'
                    ELSE COALESCE(dv.ciudad, 'â€"')
                END                                                        AS regional,
                dv.canal_rrhh                                              AS canal,
                INITCAP(dv.supervisor)                                     AS supervisor,
                dv.vendedor_nombre                                         AS vendedor,
                dc.cliente_codigo_erp                                      AS cod_cliente,
                COALESCE(dc.cliente_nombre, dc.cliente_codigo_erp)         AS cliente,
                dp.producto_codigo_erp                                     AS cod_producto,
                dp.producto_nombre                                         AS producto,
                {_SOFTYS_GRUPO_CASE}                                       AS linea_softys,
                COALESCE(dp.subgrupo_descripcion, 'â€"')                    AS subcategoria,
                fv.cantidad                                                AS cantidad,
                fv.venta_neta                                              AS venta_neta,
                fv.numero_venta                                            AS nro_pedido
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            JOIN dw.dim_cliente  dc ON fv.cliente_sk  = dc.cliente_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {dia_cond}
              AND {_SOFTYS_COND}
            ORDER BY fecha, nro_pedido
            LIMIT 150000
        """
        _, rows = _run_dw_query(sql, params)
        for r in rows:
            if r.get('fecha'):
                r['fecha'] = str(r['fecha'])

        # Presupuesto por SKU + canal (todos los SKUs Softys, tengan ppto o no)
        ppto_params = [anho, mes, anho, mes]
        if canal:
            ppto_canal_cond = "AND dv.canal_rrhh = %s"
            ppto_params.append(canal)
        else:
            ppto_canal_cond = ""
        sql_ppto = f"""
            SELECT
                dp.producto_codigo_erp                              AS cod_producto,
                dp.producto_nombre                                  AS producto,
                CASE
                    WHEN dv.ciudad IN ('SCZ')       THEN 'Santa Cruz'
                    WHEN dv.ciudad IN ('CBA')       THEN 'Cochabamba'
                    WHEN dv.ciudad IN ('LPZ','EAL') THEN 'La Paz'
                    ELSE ''
                END                                                 AS regional,
                COALESCE(dv.canal_rrhh, '')                        AS canal,
                COALESCE(SUM(fp.venta_neta_presupuestada), 0)      AS presupuesto
            FROM dw.dim_producto dp
            LEFT JOIN dw.fact_presupuesto fp
                ON fp.producto_sk = dp.producto_sk
                AND fp.anho = %s AND fp.mes = %s
                AND fp.version_sk = (
                    SELECT MAX(version_sk) FROM dw.dim_presupuesto_version
                    WHERE anho = %s AND mes = %s
                )
            LEFT JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fp.vendedor_sk
                AND ({ciudad_cond}) {ppto_canal_cond}
            WHERE {_SOFTYS_COND_DP2.replace('dp2', 'dp')}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre, dv.ciudad, dv.canal_rrhh
            ORDER BY dp.producto_nombre, dv.ciudad, dv.canal_rrhh NULLS LAST
        """
        _, ppto_rows = _run_dw_query(sql_ppto, ppto_params)

        return JsonResponse({'success': True, 'data': rows, 'total': len(rows), 'presupuesto_por_sku': ppto_rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD SUPERVISORES
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_REGIONAL_NAME_TO_KEY = {
    'Santa Cruz': 'santa_cruz',
    'Cochabamba': 'cochabamba',
    'La Paz':     'la_paz',
    'Nacional':   'nacional',
}

# Mapeo categorÃ­a â†’ valor de dp.linea en el DW
_LINEA_ALIMENTOS = 'ALIMENTOS'
_LINEA_APEGO     = 'APEGO'
_LINEA_LICORES   = 'BEBIDAS ALC'
_LINEA_HPC       = 'HOME Y PERSONAL CARE'


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('supervisores', 'unidades-supervisores')
def dashboard_supervisores_vendedores(request):
    """
    Avance por vendedor desglosado por categorÃ­a.
    - Admins: filtran por regional/canal/supervisor via query params.
    - Gerente Regional: regional+canal del perfil, supervisor libre via query param.
    - Supervisores: regional+canal del perfil, supervisor = su propio nombre.
    """
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()

        if is_admin:
            regional_key      = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal             = _safe_str(request.GET.get('canal', ''))
            supervisor_filter = _safe_str(request.GET.get('supervisor', ''))
        elif cargo == 'Gerente Regional':
            regional_key      = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal             = _safe_str(request.GET.get('canal', ''), 30)
            supervisor_filter = _safe_str(request.GET.get('supervisor', ''))
        elif 'supervisor' in cargo.lower():
            regional_key      = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal             = (profile.canal or '').strip()
            full_name         = f"{profile.user.first_name} {profile.user.last_name}".strip()
            supervisor_filter = full_name
        else:
            return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err

        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond     = _regional_filter(regional_key)
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor_filter else ""
        params_base     = [anho, mes] + ([canal] if canal else []) + ([supervisor_filter] if supervisor_filter else [])

        # â"€â"€ Ventas por vendedor y categorÃ­a (CASE WHEN pivot) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        # Agrupamos por vendedor_nombre para consolidar vendedores con mÃºltiples
        # SKs histÃ³ricos (SCD2). El SK devuelto es el actual (es_vendedor_actual=true);
        # si no existe uno actual se usa el mayor SK disponible (fallback).
        sql_ventas = f"""
            SELECT
                COALESCE(
                    MAX(CASE WHEN dv.es_vendedor_actual THEN dv.vendedor_sk ELSE NULL END),
                    MAX(dv.vendedor_sk)
                )                                                                            AS vendedor_sk,
                dv.vendedor_nombre                                                          AS vendedor,
                COALESCE(SUM(CASE WHEN dp.linea = 'ALIMENTOS'            THEN fv.venta_neta ELSE 0 END), 0) AS alimentos,
                COALESCE(SUM(CASE WHEN dp.linea = 'APEGO'                THEN fv.venta_neta ELSE 0 END), 0) AS apego,
                COALESCE(SUM(CASE WHEN dp.linea = 'BEBIDAS ALC'          THEN fv.venta_neta ELSE 0 END), 0) AS licores,
                COALESCE(SUM(CASE WHEN dp.linea = 'HOME Y PERSONAL CARE' THEN fv.venta_neta ELSE 0 END), 0) AS hpc,
                COALESCE(SUM(CASE WHEN dp.linea = 'SIN LINEA' OR dp.linea IS NULL THEN fv.venta_neta ELSE 0 END), 0) AS sin_clasificar,
                COALESCE(SUM(fv.venta_neta), 0)                                                              AS total,
                COALESCE(SUM(CASE WHEN dp.linea = 'ALIMENTOS'            THEN fv.cantidad ELSE 0 END), 0)   AS alimentos_cant,
                COALESCE(SUM(CASE WHEN dp.linea = 'APEGO'                THEN fv.cantidad ELSE 0 END), 0)   AS apego_cant,
                COALESCE(SUM(CASE WHEN dp.linea = 'BEBIDAS ALC'          THEN fv.cantidad ELSE 0 END), 0)   AS licores_cant,
                COALESCE(SUM(CASE WHEN dp.linea = 'HOME Y PERSONAL CARE' THEN fv.cantidad ELSE 0 END), 0)   AS hpc_cant,
                COALESCE(SUM(CASE WHEN dp.linea = 'SIN LINEA' OR dp.linea IS NULL THEN fv.cantidad ELSE 0 END), 0) AS sin_clasificar_cant,
                COALESCE(SUM(fv.cantidad), 0)                                                                AS total_cant
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
            GROUP BY dv.vendedor_nombre
            ORDER BY total DESC
        """
        _, ventas_rows = _run_dw_query(sql_ventas, params_base)

        # â"€â"€ Presupuesto por vendedor y categorÃ­a â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        ppto_map = {}
        ppto_rows = []
        try:
            sql_ppto = f"""
                SELECT
                    COALESCE(
                        MAX(CASE WHEN dv.es_vendedor_actual THEN dv.vendedor_sk ELSE NULL END),
                        MAX(dv.vendedor_sk)
                    )                                                                                                          AS vendedor_sk,
                    COALESCE(SUM(CASE WHEN dp.linea = 'ALIMENTOS'            THEN fp.venta_neta_presupuestada ELSE 0 END), 0) AS alimentos_ppto,
                    COALESCE(SUM(CASE WHEN dp.linea = 'APEGO'                THEN fp.venta_neta_presupuestada ELSE 0 END), 0) AS apego_ppto,
                    COALESCE(SUM(CASE WHEN dp.linea = 'BEBIDAS ALC'          THEN fp.venta_neta_presupuestada ELSE 0 END), 0) AS licores_ppto,
                    COALESCE(SUM(CASE WHEN dp.linea = 'HOME Y PERSONAL CARE' THEN fp.venta_neta_presupuestada ELSE 0 END), 0) AS hpc_ppto,
                    COALESCE(SUM(CASE WHEN dp.linea = 'SIN LINEA' OR dp.linea IS NULL THEN fp.venta_neta_presupuestada ELSE 0 END), 0) AS sin_clasificar_ppto,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)                                                              AS total_ppto,
                    COALESCE(SUM(CASE WHEN dp.linea = 'ALIMENTOS'            THEN fp.cantidad_presupuestada ELSE 0 END), 0)   AS alimentos_ppto_uds,
                    COALESCE(SUM(CASE WHEN dp.linea = 'APEGO'                THEN fp.cantidad_presupuestada ELSE 0 END), 0)   AS apego_ppto_uds,
                    COALESCE(SUM(CASE WHEN dp.linea = 'BEBIDAS ALC'          THEN fp.cantidad_presupuestada ELSE 0 END), 0)   AS licores_ppto_uds,
                    COALESCE(SUM(CASE WHEN dp.linea = 'HOME Y PERSONAL CARE' THEN fp.cantidad_presupuestada ELSE 0 END), 0)   AS hpc_ppto_uds,
                    COALESCE(SUM(CASE WHEN dp.linea = 'SIN LINEA' OR dp.linea IS NULL THEN fp.cantidad_presupuestada ELSE 0 END), 0) AS sin_clasificar_ppto_uds,
                    COALESCE(SUM(fp.cantidad_presupuestada), 0)                                                                AS total_ppto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor             dv  ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dv.vendedor_nombre
            """
            _, ppto_rows = _run_dw_query(sql_ppto, params_base + [anho, mes])
            ppto_map = {r['vendedor_sk']: r for r in ppto_rows}
        except Exception:
            pass

        def _pct(avance, ppto):
            a, p = float(avance or 0), float(ppto or 0)
            return round(a / p * 100, 1) if p > 0 else None

        def _pct_uds(cant, ppto_uds):
            c, pu = int(cant or 0), float(ppto_uds or 0)
            return round(c / pu * 100, 1) if pu > 0 else None

        result = []
        for row in ventas_rows:
            sk   = row['vendedor_sk']
            p    = ppto_map.get(sk, {})
            a_a  = float(row['alimentos']       or 0);  ca_a = int(row['alimentos_cant']       or 0)
            a_e  = float(row['apego']           or 0);  ca_e = int(row['apego_cant']           or 0)
            a_l  = float(row['licores']         or 0);  ca_l = int(row['licores_cant']         or 0)
            a_h  = float(row['hpc']             or 0);  ca_h = int(row['hpc_cant']             or 0)
            a_sc = float(row['sin_clasificar']  or 0);  ca_sc= int(row['sin_clasificar_cant']  or 0)
            a_t  = float(row['total']           or 0);  ca_t = int(row['total_cant']           or 0)
            p_a  = float(p.get('alimentos_ppto')        or 0);  pu_a  = float(p.get('alimentos_ppto_uds')        or 0)
            p_e  = float(p.get('apego_ppto')            or 0);  pu_e  = float(p.get('apego_ppto_uds')            or 0)
            p_l  = float(p.get('licores_ppto')          or 0);  pu_l  = float(p.get('licores_ppto_uds')          or 0)
            p_h  = float(p.get('hpc_ppto')              or 0);  pu_h  = float(p.get('hpc_ppto_uds')              or 0)
            p_sc = float(p.get('sin_clasificar_ppto')   or 0);  pu_sc = float(p.get('sin_clasificar_ppto_uds')   or 0)
            p_t  = float(p.get('total_ppto')            or 0);  pu_t  = float(p.get('total_ppto_uds')            or 0)
            result.append({
                'vendedor_sk':    sk,
                'vendedor':       row['vendedor'],
                'alimentos':      a_a,  'alimentos_ppto':      p_a,  'alimentos_pct':      _pct(a_a,  p_a),  'alimentos_cant':      ca_a,  'alimentos_ppto_uds':      int(pu_a),  'alimentos_pct_uds':      _pct_uds(ca_a,  pu_a),
                'apego':          a_e,  'apego_ppto':          p_e,  'apego_pct':          _pct(a_e,  p_e),  'apego_cant':          ca_e,  'apego_ppto_uds':          int(pu_e),  'apego_pct_uds':          _pct_uds(ca_e,  pu_e),
                'licores':        a_l,  'licores_ppto':        p_l,  'licores_pct':        _pct(a_l,  p_l),  'licores_cant':        ca_l,  'licores_ppto_uds':        int(pu_l),  'licores_pct_uds':        _pct_uds(ca_l,  pu_l),
                'hpc':            a_h,  'hpc_ppto':            p_h,  'hpc_pct':            _pct(a_h,  p_h),  'hpc_cant':            ca_h,  'hpc_ppto_uds':            int(pu_h),  'hpc_pct_uds':            _pct_uds(ca_h,  pu_h),
                'sin_clasificar': a_sc, 'sin_clasificar_ppto': p_sc, 'sin_clasificar_pct': _pct(a_sc, p_sc), 'sin_clasificar_cant': ca_sc, 'sin_clasificar_ppto_uds': int(pu_sc), 'sin_clasificar_pct_uds': _pct_uds(ca_sc, pu_sc),
                'total':          a_t,  'total_ppto':          p_t,  'total_pct':          _pct(a_t,  p_t),  'total_cant':          ca_t,  'total_ppto_uds':          int(pu_t),  'total_pct_uds':          _pct_uds(ca_t,  pu_t),
            })

        total_avance = sum(r['total'] for r in result)
        # Sumar presupuesto de TODOS los vendedores con presupuesto (no solo los que vendieron)
        total_ppto   = sum(float(r.get('total_ppto') or 0) for r in ppto_rows)

        sql_fc = f"""
            SELECT MAX(df.fecha_completa) AS fc
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
        """
        _, fc_rows  = _run_dw_query(sql_fc, params_base)
        fecha_corte = str(fc_rows[0]['fc']) if fc_rows and fc_rows[0]['fc'] else None

        return JsonResponse({
            'success':      True,
            'regional':     regional_key,
            'canal':        canal,
            'supervisor':   supervisor_filter,
            'total_avance': total_avance,
            'total_ppto':   total_ppto,
            'total_pct':    round(total_avance / total_ppto * 100, 1) if total_ppto > 0 else None,
            'fecha_corte':  fecha_corte,
            'vendedores':   result,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('supervisores', 'unidades-supervisores')
def dashboard_supervisores_liquidaciones(request):
    """
    Venta neta diaria por vendedor, excluyendo domingos.
    Retorna { fechas: [str], rows: [{ vendedor_sk, vendedor, ventas: {fecha: monto} }] }
    """
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()

        if is_admin:
            regional_key      = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal             = _safe_str(request.GET.get('canal', ''))
            supervisor_filter = _safe_str(request.GET.get('supervisor', ''))
        elif cargo == 'Gerente Regional':
            regional_key      = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal             = _safe_str(request.GET.get('canal', ''), 30)
            supervisor_filter = _safe_str(request.GET.get('supervisor', ''))
        elif 'supervisor' in cargo.lower():
            regional_key      = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal             = (profile.canal or '').strip()
            full_name         = f"{profile.user.first_name} {profile.user.last_name}".strip()
            supervisor_filter = full_name
        else:
            return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err

        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond     = _regional_filter(regional_key)
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor_filter else ""
        params          = [anho, mes] + ([canal] if canal else []) + ([supervisor_filter] if supervisor_filter else [])

        sql = f"""
            SELECT
                dv.vendedor_sk,
                dv.vendedor_nombre                               AS vendedor,
                df.fecha_completa::date                          AS fecha,
                COALESCE(SUM(fv.venta_neta), 0)                 AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND EXTRACT(DOW FROM df.fecha_completa) != 0
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
            GROUP BY dv.vendedor_sk, dv.vendedor_nombre, df.fecha_completa::date
            ORDER BY dv.vendedor_nombre, df.fecha_completa::date
        """
        _, rows = _run_dw_query(sql, params)

        vendedores: dict = {}
        fechas_set: set  = set()
        for r in rows:
            sk    = r['vendedor_sk']
            fecha = str(r['fecha'])
            venta = float(r['venta_neta'] or 0)
            fechas_set.add(fecha)
            if sk not in vendedores:
                vendedores[sk] = {'vendedor_sk': sk, 'vendedor': r['vendedor'], 'ventas': {}}
            vendedores[sk]['ventas'][fecha] = venta

        fechas = sorted(fechas_set)
        return JsonResponse({'success': True, 'fechas': fechas, 'rows': list(vendedores.values())})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('supervisores', 'tendencia-estacional', 'preventas-realizadas',
                   'unidades-supervisores', 'informacion-rutas', 'canales', 'regionales')
def dashboard_supervisores_supervisor_lista(request):
    """Retorna lista de supervisores distintos para el regional/canal/aÃ±o/mes dado."""
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()

        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
        elif cargo == 'Gerente Regional':
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
        else:
            return JsonResponse({'success': True, 'data': []})

        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err  = _validate_anho_mes(anho, mes)
        if err: return err

        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional_key)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params      = [anho, mes] + ([canal] if canal else [])

        sql = f"""
            SELECT DISTINCT INITCAP(dv.supervisor) AS supervisor
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond}
              AND dv.supervisor IS NOT NULL AND dv.supervisor != ''
            ORDER BY supervisor
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': [r['supervisor'] for r in rows]})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD PREVENTAS REALIZADAS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

import calendar as _cal_mod

def _preventas_fecha_rango(request):
    """Returns (fecha_desde, fecha_hasta) strings 'YYYY-MM-DD' for SQL filter.
    Accepts fecha_desde/fecha_hasta params; falls back to anho+mes."""
    fd = request.GET.get('fecha_desde', '').strip()
    fh = request.GET.get('fecha_hasta', '').strip()
    if fd and fh:
        try:
            date.fromisoformat(fd)
            date.fromisoformat(fh)
            return fd, fh
        except ValueError:
            pass
    now  = datetime.now()
    anho = _safe_int(request.GET.get('anho'), now.year)
    mes  = _safe_int(request.GET.get('mes'),  now.month)
    last = _cal_mod.monthrange(anho, mes)[1]
    return f"{anho}-{mes:02d}-01", f"{anho}-{mes:02d}-{last:02d}"

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('preventas-realizadas')
def dashboard_preventas_kpis(request):
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif cargo == 'Gerente Regional':
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif 'supervisor' in cargo.lower():
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = f"{profile.user.first_name} {profile.user.last_name}".strip()
        else:
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        fecha_desde, fecha_hasta = _preventas_fecha_rango(request)
        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        ciudad_cond     = _regional_filter(regional_key, campo='dv.ciudad')
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor else ""
        params = [fecha_desde, fecha_hasta] + ([canal] if canal else []) + ([supervisor] if supervisor else [])
        sql = f"""
            SELECT
                COUNT(DISTINCT dp.nro_transaccion)                             AS total_pedidos,
                ROUND(COALESCE(SUM(dp.importe_total), 0)::NUMERIC, 2)          AS total_importe,
                MAX(dp.fecha_actualizacion)                                     AS ultima_actualizacion
            FROM dual.dim_preventa dp
            JOIN dw.dim_vendedor dv ON dv.vendedor_codigo_erp = dp.codigo_usuario
                AND dv.es_vendedor_actual = TRUE
            WHERE dp.fecha_transaccion::date BETWEEN %s AND %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
        """
        _, rows = _run_dw_query(sql, params)
        row = rows[0] if rows else {}
        ua  = row.get('ultima_actualizacion')
        if ua is not None and hasattr(ua, 'isoformat'):
            ua_str = ua.isoformat()
        elif ua is not None:
            ua_str = str(ua)
        else:
            ua_str = None
        return JsonResponse({
            'success':               True,
            'total_pedidos':         row.get('total_pedidos', 0),
            'total_importe':         float(row.get('total_importe', 0) or 0),
            'ultima_actualizacion':  ua_str,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('preventas-realizadas')
def dashboard_preventas_por_canal(request):
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif cargo == 'Gerente Regional':
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif 'supervisor' in cargo.lower():
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = f"{profile.user.first_name} {profile.user.last_name}".strip()
        else:
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        fecha_desde, fecha_hasta = _preventas_fecha_rango(request)
        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        ciudad_cond     = _regional_filter(regional_key, campo='dv.ciudad')
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor else ""
        # supervisor activo â†’ agrupa por vendedor; canal activo â†’ por supervisor; sin filtros â†’ por canal
        if supervisor:
            grupo_col        = "dp.nombre_usuario"
            agrupado_por_val = "vendedor"
        elif canal:
            grupo_col        = "dv.supervisor"
            agrupado_por_val = "supervisor"
        else:
            grupo_col        = "dv.canal_rrhh"
            agrupado_por_val = "canal"
        params = [fecha_desde, fecha_hasta] + ([canal] if canal else []) + ([supervisor] if supervisor else [])
        sql = f"""
            SELECT
                COALESCE({grupo_col}, 'Sin Asignar')                           AS grupo,
                COUNT(DISTINCT dp.nro_transaccion)                             AS pedidos,
                ROUND(COALESCE(SUM(dp.importe_total), 0)::NUMERIC, 2)          AS monto
            FROM dual.dim_preventa dp
            JOIN dw.dim_vendedor dv ON dv.vendedor_codigo_erp = dp.codigo_usuario
                AND dv.es_vendedor_actual = TRUE
            WHERE dp.fecha_transaccion::date BETWEEN %s AND %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
            GROUP BY {grupo_col}
            ORDER BY monto DESC
            LIMIT 20
        """
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'grupo':   r['grupo'],
                'pedidos': r['pedidos'],
                'monto':   float(r['monto'] or 0),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data, 'agrupado_por': agrupado_por_val})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('preventas-realizadas')
def dashboard_preventas_por_vendedor(request):
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()
        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif cargo == 'Gerente Regional':
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        elif 'supervisor' in cargo.lower():
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = f"{profile.user.first_name} {profile.user.last_name}".strip()
        else:
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
            supervisor   = _safe_str(request.GET.get('supervisor', ''))
        fecha_desde, fecha_hasta = _preventas_fecha_rango(request)
        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        ciudad_cond     = _regional_filter(regional_key, campo='dv.ciudad')
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor else ""
        # params: [fd,fh] for rutas CTE, [fd,fh] for clientes CTE inner, [fd,fh] for main WHERE + filters
        params = (
            [fecha_desde, fecha_hasta]
            + [fecha_desde, fecha_hasta]
            + [fecha_desde, fecha_hasta]
            + ([canal] if canal else [])
            + ([supervisor] if supervisor else [])
        )
        sql = f"""
            WITH ruta_clientes AS (
                SELECT ruta, COUNT(*) AS total_clientes
                FROM dual.dim_cliente_dual
                WHERE es_actual = true
                GROUP BY ruta
            ),
            vendedor_rutas AS (
                SELECT
                    nombre_usuario,
                    STRING_AGG(DISTINCT ruta, ' / ' ORDER BY ruta) AS rutas_concat
                FROM dual.dim_preventa
                WHERE fecha_transaccion::date BETWEEN %s AND %s
                GROUP BY nombre_usuario
            ),
            vendedor_clientes_total AS (
                SELECT
                    sub.nombre_usuario,
                    COALESCE(SUM(rc.total_clientes), 0) AS total_clientes
                FROM (
                    SELECT DISTINCT nombre_usuario, ruta
                    FROM dual.dim_preventa
                    WHERE fecha_transaccion::date BETWEEN %s AND %s
                ) sub
                LEFT JOIN ruta_clientes rc ON rc.ruta = sub.ruta
                GROUP BY sub.nombre_usuario
            )
            SELECT
                dp.nombre_usuario                                                       AS vendedor,
                vr.rutas_concat                                                         AS ruta,
                MAX(dv.supervisor)                                                      AS supervisor,
                COALESCE(vtc.total_clientes, 0)                                         AS total_clientes,
                COUNT(DISTINCT dp.cod_cliente)                                          AS pedidos,
                ROUND(
                    COUNT(DISTINCT dp.cod_cliente)::NUMERIC
                    / NULLIF(COALESCE(vtc.total_clientes, 0), 0) * 100
                , 1)                                                                    AS pct_efectividad,
                ROUND(COALESCE(SUM(dp.importe_total), 0)::NUMERIC, 2)                  AS monto_total,
                TO_CHAR(MIN(dp.fecha_transaccion), 'HH24:MI')                          AS hora_inicio,
                TO_CHAR(MAX(dp.fecha_transaccion), 'HH24:MI')                          AS hora_ultimo,
                ROUND(EXTRACT(EPOCH FROM (MAX(dp.fecha_transaccion)
                    - MIN(dp.fecha_transaccion))) / 60)                                AS minutos_trabajados
            FROM dual.dim_preventa dp
            JOIN dw.dim_vendedor dv ON dv.vendedor_codigo_erp = dp.codigo_usuario
                AND dv.es_vendedor_actual = TRUE
            LEFT JOIN vendedor_rutas vr ON vr.nombre_usuario = dp.nombre_usuario
            LEFT JOIN vendedor_clientes_total vtc ON vtc.nombre_usuario = dp.nombre_usuario
            WHERE dp.fecha_transaccion::date BETWEEN %s AND %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
            GROUP BY dp.nombre_usuario, vr.rutas_concat, vtc.total_clientes
            ORDER BY monto_total DESC
        """
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'vendedor':           r['vendedor'],
                'ruta':               r['ruta'],
                'supervisor':         r['supervisor'],
                'total_clientes':     r['total_clientes'] or 0,
                'pedidos':            r['pedidos'],
                'pct_efectividad':    float(r['pct_efectividad']) if r.get('pct_efectividad') is not None else None,
                'monto_total':        float(r['monto_total'] or 0),
                'hora_inicio':        r.get('hora_inicio'),
                'hora_ultimo':        r.get('hora_ultimo'),
                'minutos_trabajados': int(r['minutos_trabajados']) if r.get('minutos_trabajados') is not None else None,
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('preventas-realizadas')
def dashboard_preventas_top_faltantes(request):
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
        else:
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
        fecha_desde, fecha_hasta = _preventas_fecha_rango(request)
        supervisor = _safe_str(request.GET.get('supervisor', ''))
        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        ciudad_cond     = _regional_filter(regional_key, campo='dv.ciudad')
        canal_cond      = "AND dv.canal_rrhh = %s" if canal else ""
        supervisor_cond = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor else ""
        params = [fecha_desde, fecha_hasta] + ([canal] if canal else []) + ([supervisor] if supervisor else [])
        sql = f"""
            SELECT
                dp.cod_producto                                                    AS codigo,
                dp.nombre_producto                                                 AS producto,
                COALESCE(SUM(dp.cantidad), 0)                                      AS cant_pedida,
                COALESCE(SUM(CASE WHEN dp.estado = true  THEN dp.cantidad ELSE 0 END), 0) AS cant_atendida,
                COALESCE(SUM(CASE WHEN dp.estado = false THEN dp.cantidad ELSE 0 END), 0) AS cant_faltante,
                ROUND(
                    SUM(CASE WHEN dp.estado = true THEN dp.cantidad ELSE 0 END)
                    / NULLIF(SUM(dp.cantidad), 0) * 100
                , 1) AS ns_pct
            FROM dual.dim_preventa dp
            JOIN dw.dim_vendedor dv ON dv.vendedor_codigo_erp = dp.codigo_usuario
                AND dv.es_vendedor_actual = TRUE
            WHERE dp.fecha_transaccion::date BETWEEN %s AND %s
              AND ({ciudad_cond}) {canal_cond} {supervisor_cond}
            GROUP BY dp.cod_producto, dp.nombre_producto
            HAVING SUM(CASE WHEN dp.estado = false THEN dp.cantidad ELSE 0 END) > 0
            ORDER BY cant_faltante DESC
            LIMIT 20
        """
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'codigo':       r['codigo'],
                'producto':     r['producto'],
                'linea':        '',
                'cant_pedida':  float(r['cant_pedida'] or 0),
                'cant_atendida':float(r['cant_atendida'] or 0),
                'cant_faltante':float(r['cant_faltante'] or 0),
                'ns_pct':       float(r['ns_pct']) if r.get('ns_pct') is not None else None,
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('preventas-realizadas')
def dashboard_preventas_supervisores_lista(request):
    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        if is_admin:
            regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''))
        else:
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
        fecha_desde, fecha_hasta = _preventas_fecha_rango(request)
        if regional_key not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        ciudad_cond = _regional_filter(regional_key, campo='dv.ciudad')
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        params      = [fecha_desde, fecha_hasta] + ([canal] if canal else [])
        sql = f"""
            SELECT DISTINCT dv.supervisor
            FROM dual.dim_preventa dp
            JOIN dw.dim_vendedor dv ON dv.vendedor_codigo_erp = dp.codigo_usuario
                AND dv.es_vendedor_actual = TRUE
            WHERE dp.fecha_transaccion::date BETWEEN %s AND %s
              AND ({ciudad_cond}) {canal_cond}
              AND dv.supervisor IS NOT NULL AND dv.supervisor != ''
            ORDER BY dv.supervisor
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': [r['supervisor'] for r in rows]})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD UNIDADES VENDIDAS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_UNIDADES_CAT_LINEA = {
    'Alimentos':            'ALIMENTOS',
    'Apego':                'APEGO',
    'Licores':              'BEBIDAS ALC',
    'Home & Personal Care': 'HOME Y PERSONAL CARE',
    'Sin Clasificar':       None,   # linea IS NULL OR linea = 'SIN LINEA'
}

_SIN_CLASIFICAR_COND = "(dp.linea = 'SIN LINEA' OR dp.linea IS NULL)"


def _unidades_cat_params(categoria, base_params):
    """Returns (cat_cond_sql, params_extended)."""
    if categoria and categoria in _UNIDADES_CAT_LINEA:
        if categoria == 'Sin Clasificar':
            return f"AND {_SIN_CLASIFICAR_COND}", list(base_params)
        return "AND dp.linea = %s", list(base_params) + [_UNIDADES_CAT_LINEA[categoria]]
    return "", list(base_params)


def _multi_cat_cond(categorias):
    """Multi-select categoría → OR of linea conditions."""
    if not categorias:
        return "", []
    parts, params = [], []
    for cat in categorias:
        if cat == 'Sin Clasificar':
            parts.append(_SIN_CLASIFICAR_COND)
        elif cat in _UNIDADES_CAT_LINEA and _UNIDADES_CAT_LINEA[cat]:
            parts.append("dp.linea = %s")
            params.append(_UNIDADES_CAT_LINEA[cat])
    if not parts:
        return "", []
    return f"AND ({' OR '.join(parts)})", params


def _multi_prov_cond(proveedores):
    """Multi-select proveedor → UPPER(dp.proveedor) IN (...)."""
    if not proveedores:
        return "", []
    phs = ", ".join(["%s"] * len(proveedores))
    return f"AND UPPER(dp.proveedor) IN ({phs})", [p.upper() for p in proveedores]


def _multi_sub_cond(subgrupos):
    """Multi-select subgrupo_descripcion → IN (...)."""
    if not subgrupos:
        return "", []
    phs = ", ".join(["%s"] * len(subgrupos))
    return f"AND dp.subgrupo_descripcion IN ({phs})", list(subgrupos)


def _multi_marc_cond(marcas):
    """Multi-select marca → IN (...)."""
    if not marcas:
        return "", []
    phs = ", ".join(["%s"] * len(marcas))
    return f"AND dp.marca IN ({phs})", list(marcas)


def _prev_period(anho, mes):
    return (anho - 1, 12) if mes == 1 else (anho, mes - 1)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('unidades-vendidas')
def dashboard_unidades_kpis(request):
    """KPI totales: cantidad, venta_neta, presupuesto. Params: regional, canal, categoria, proveedor, marca, anho, mes."""
    try:
        regional  = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal     = _safe_str(request.GET.get('canal', ''))
        categoria = _safe_str(request.GET.get('categoria', ''))
        proveedor = _safe_str(request.GET.get('proveedor', ''))
        marca     = _safe_str(request.GET.get('marca', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond    = _regional_filter(regional)
        canal_cond     = "AND dv.canal_rrhh = %s" if canal else ""
        proveedor_cond = "AND UPPER(dp.proveedor) = UPPER(%s)" if proveedor else ""
        marca_cond     = "AND dp.marca = %s" if marca else ""
        base_params    = [anho, mes] + ([canal] if canal else [])
        cat_cond, cat_params = _unidades_cat_params(categoria, base_params)
        extra = ([proveedor] if proveedor else []) + ([marca] if marca else [])
        params_v = cat_params + extra

        sql_v = f"""
            SELECT
                COALESCE(SUM(fv.cantidad), 0)   AS total_cantidad,
                COALESCE(SUM(fv.venta_neta), 0) AS total_venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {proveedor_cond} {marca_cond}
        """
        _, v_rows = _run_dw_query(sql_v, params_v)

        ppto_cat_cond, params_pp = _unidades_cat_params(categoria, base_params)
        sql_p = f"""
            SELECT COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS total_ppto
            FROM dw.fact_presupuesto fp
            JOIN dw.dim_vendedor             dv  ON fp.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
            WHERE fp.anho = %s AND fp.mes = %s
              AND ({ciudad_cond}) {canal_cond} {ppto_cat_cond} {proveedor_cond} {marca_cond}
              AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
        """
        p_rows = []
        try:
            _, p_rows = _run_dw_query(sql_p, params_pp + extra + [anho, mes])
        except Exception:
            pass

        sql_fc = f"""
            SELECT MAX(df.fecha_completa) AS fc
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond}
        """
        _, fc_rows = _run_dw_query(sql_fc, [anho, mes] + ([canal] if canal else []))

        total_cant  = int(v_rows[0]['total_cantidad'] or 0)    if v_rows else 0
        total_vn    = float(v_rows[0]['total_venta_neta'] or 0) if v_rows else 0
        total_ppto  = float(p_rows[0]['total_ppto'] or 0)      if p_rows else 0
        fecha_corte = str(fc_rows[0]['fc']) if fc_rows and fc_rows[0]['fc'] else None

        return JsonResponse({
            'success':          True,
            'total_cantidad':   total_cant,
            'total_venta_neta': total_vn,
            'total_ppto':       total_ppto,
            'total_pct':        round(total_vn / total_ppto * 100, 1) if total_ppto > 0 else None,
            'fecha_corte':      fecha_corte,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('unidades-vendidas', 'unidades-supervisores')
def dashboard_unidades_por_subgrupo(request):
    """
    Ventas+presupuesto agrupados por subgrupo dentro de la categorÃ­a seleccionada.
    Params: regional, canal, categoria (requerido), anho, mes, proveedor
    """
    try:
        regional  = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal     = _safe_str(request.GET.get('canal', ''))
        categoria = _safe_str(request.GET.get('categoria', ''))
        proveedor = _safe_str(request.GET.get('proveedor', ''))
        marca     = _safe_str(request.GET.get('marca', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond    = _regional_filter(regional)
        canal_cond     = "AND dv.canal_rrhh = %s" if canal else ""
        proveedor_cond = "AND UPPER(dp.proveedor) = UPPER(%s)" if proveedor else ""
        marca_cond     = "AND dp.marca = %s" if marca else ""
        base_params    = [anho, mes] + ([canal] if canal else [])
        cat_cond, params = _unidades_cat_params(categoria, base_params)
        if proveedor:
            params = params + [proveedor]
        if marca:
            params = params + [marca]

        sql_v = f"""
            SELECT
                COALESCE(dp.subgrupo_descripcion, 'Sin Subgrupo') AS subgrupo,
                COALESCE(SUM(fv.cantidad), 0)                      AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0)                    AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {proveedor_cond} {marca_cond}
            GROUP BY dp.subgrupo_descripcion
            ORDER BY venta_neta DESC
        """
        _, v_rows = _run_dw_query(sql_v, params)

        ppto_map = {}
        try:
            ppto_cat_cond, params_p = _unidades_cat_params(categoria, base_params)
            if proveedor:
                params_p = params_p + [proveedor]
            if marca:
                params_p = params_p + [marca]
            sql_p = f"""
                SELECT
                    COALESCE(dp.subgrupo_descripcion, 'Sin Subgrupo') AS subgrupo,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)     AS presupuesto,
                    COALESCE(SUM(fp.cantidad_presupuestada), 0)        AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor             dv  ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {ppto_cat_cond} {proveedor_cond} {marca_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.subgrupo_descripcion
            """
            _, p_rows = _run_dw_query(sql_p, params_p + [anho, mes])
            ppto_map = {
                r['subgrupo']: (float(r['presupuesto'] or 0), float(r['presupuesto_uds'] or 0))
                for r in p_rows
            }
        except Exception:
            pass

        result = []
        for row in v_rows:
            sg        = row['subgrupo']
            vn        = float(row['venta_neta'] or 0)
            cant      = int(row['cantidad'] or 0)
            ppto, ppto_uds = ppto_map.get(sg, (0, 0))
            result.append({
                'subgrupo':       sg,
                'cantidad':       cant,
                'venta_neta':     vn,
                'presupuesto':    ppto,
                'presupuesto_uds': int(ppto_uds),
                'porcentaje':     round(vn   / ppto     * 100, 1) if ppto     > 0 else None,
                'porcentaje_uds': round(cant / ppto_uds * 100, 1) if ppto_uds > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('unidades-vendidas', 'unidades-supervisores')
def dashboard_unidades_proveedores(request):
    """
    Proveedores distintos que tienen ventas en la categorÃ­a dada.
    Params: regional, canal, categoria, anho, mes
    """
    try:
        regional  = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal     = _safe_str(request.GET.get('canal', ''))
        categoria = _safe_str(request.GET.get('categoria', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        base_params = [anho, mes] + ([canal] if canal else [])
        cat_cond, params = _unidades_cat_params(categoria, base_params)

        sql = f"""
            SELECT DISTINCT UPPER(dp.proveedor) AS proveedor
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond}
              AND dp.proveedor IS NOT NULL AND dp.proveedor <> ''
            ORDER BY proveedor
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': [r['proveedor'] for r in rows]})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('unidades-vendidas')
def dashboard_unidades_por_sku(request):
    """
    SKU-level data filtrado por categorÃ­a + subgrupo.
    Params: regional, canal, categoria, subgrupo, anho, mes, limit
    """
    try:
        regional  = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal     = _safe_str(request.GET.get('canal', ''))
        categoria = _safe_str(request.GET.get('categoria', ''))
        subgrupo  = _safe_str(request.GET.get('subgrupo', ''))
        proveedor = _safe_str(request.GET.get('proveedor', ''))
        marca     = _safe_str(request.GET.get('marca', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err       = _validate_anho_mes(anho, mes)
        if err: return err
        limit     = min(_safe_int(request.GET.get('limit'), 500), 1000)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond    = _regional_filter(regional)
        canal_cond     = "AND dv.canal_rrhh = %s" if canal else ""
        proveedor_cond = "AND UPPER(dp.proveedor) = UPPER(%s)" if proveedor else ""
        marca_cond     = "AND dp.marca = %s" if marca else ""
        base_params    = [anho, mes] + ([canal] if canal else [])
        cat_cond, cat_params = _unidades_cat_params(categoria, base_params)
        sub_cond   = "AND dp.subgrupo_descripcion = %s" if subgrupo else ""
        sub_extra  = [subgrupo] if subgrupo else []
        prov_extra = [proveedor] if proveedor else []
        marc_extra = [marca] if marca else []

        params_v    = cat_params + sub_extra + prov_extra + marc_extra + [limit]
        params_ppto = cat_params + sub_extra + prov_extra + marc_extra

        sql_v = f"""
            SELECT
                dp.producto_codigo_erp          AS codigo,
                dp.producto_nombre              AS producto,
                COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0) AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {sub_cond} {proveedor_cond} {marca_cond}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY venta_neta DESC
            LIMIT %s
        """
        _, v_rows = _run_dw_query(sql_v, params_v)

        ppto_map = {}
        try:
            sql_p = f"""
                SELECT
                    dp.producto_codigo_erp                        AS codigo,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto,
                    COALESCE(SUM(fp.cantidad_presupuestada), 0)   AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor             dv  ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {cat_cond} {sub_cond} {proveedor_cond} {marca_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.producto_codigo_erp
            """
            _, p_rows = _run_dw_query(sql_p, params_ppto + [anho, mes])
            ppto_map = {r['codigo']: r for r in p_rows}
        except Exception:
            pass

        result = []
        for row in v_rows:
            vn       = float(row['venta_neta'] or 0)
            cant     = int(row['cantidad'] or 0)
            p        = ppto_map.get(row['codigo'], {})
            ppto_bs  = float(p.get('presupuesto',     0) or 0)
            ppto_uds = float(p.get('presupuesto_uds', 0) or 0)
            result.append({
                'codigo':          row['codigo'],
                'producto':        row['producto'],
                'cantidad':        cant,
                'venta_neta':      vn,
                'presupuesto':     ppto_bs,
                'presupuesto_uds': int(ppto_uds),
                'porcentaje':      round(vn   / ppto_bs  * 100, 1) if ppto_bs  > 0 else None,
                'porcentaje_uds':  round(cant / ppto_uds * 100, 1) if ppto_uds > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('new-nacional')
def dashboard_new_nacional_opciones(request):
    """
    Opciones en cascada (multi-select) para los filtros del mock-up.
    Params (multi-value): categoria[], proveedor[], subgrupo[]
    Params: regional, canal, anho, mes
    """
    try:
        regional    = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        canal       = _safe_str(request.GET.get('canal', ''))
        categorias  = [s for s in request.GET.getlist('categoria') if s]
        proveedores = [s for s in request.GET.getlist('proveedor') if s]
        subgrupos   = [s for s in request.GET.getlist('subgrupo') if s]
        anho        = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes         = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            regional = 'nacional'

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        canal_param = [canal] if canal else []
        base_params = [anho, mes] + canal_param

        cat_cond,  cat_params  = _multi_cat_cond(categorias)
        prov_cond, prov_params = _multi_prov_cond(proveedores)
        sub_cond,  sub_params  = _multi_sub_cond(subgrupos)

        # 1) Proveedores: filtrado solo por categorías
        sql_prov = f"""
            SELECT DISTINCT UPPER(dp.proveedor) AS proveedor
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond}
              AND dp.proveedor IS NOT NULL AND dp.proveedor <> ''
            ORDER BY proveedor
        """
        _, prov_rows = _run_dw_query(sql_prov, base_params + cat_params)

        # 2) Subgrupos: filtrado por categorías + proveedores
        sql_sub = f"""
            SELECT DISTINCT dp.subgrupo_descripcion AS subgrupo
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {prov_cond}
              AND dp.subgrupo_descripcion IS NOT NULL AND dp.subgrupo_descripcion <> ''
            ORDER BY subgrupo
        """
        _, sub_rows = _run_dw_query(sql_sub, base_params + cat_params + prov_params)

        # 3) Marcas: filtrado por categorías + proveedores + subgrupos
        sql_marc = f"""
            SELECT DISTINCT dp.marca AS marca
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {prov_cond} {sub_cond}
              AND dp.marca IS NOT NULL AND dp.marca <> ''
            ORDER BY marca
        """
        _, marc_rows = _run_dw_query(sql_marc, base_params + cat_params + prov_params + sub_params)

        return JsonResponse({
            'success':     True,
            'proveedores': [r['proveedor'] for r in prov_rows],
            'subgrupos':   [r['subgrupo']  for r in sub_rows],
            'marcas':      [r['marca']     for r in marc_rows],
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('new-nacional')
def dashboard_new_nacional_comparacion(request):
    """
    Tabla comparativa por grupo con mes actual y mes anterior.
    Agrupa por el filtro más específico seleccionado: marca > subgrupo > proveedor > categoría > total.
    Params (multi-value): categoria[], proveedor[], subgrupo[], marca[]
    Params: regional, canal, anho, mes
    """
    try:
        regional    = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        canal       = _safe_str(request.GET.get('canal', ''))
        categorias  = [s for s in request.GET.getlist('categoria') if s]
        proveedores = [s for s in request.GET.getlist('proveedor') if s]
        subgrupos   = [s for s in request.GET.getlist('subgrupo') if s]
        marcas      = [s for s in request.GET.getlist('marca') if s]
        anho        = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes         = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            regional = 'nacional'

        prev_anho, prev_mes = _prev_period(anho, mes)
        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        canal_param = [canal] if canal else []

        cat_cond,  cat_params  = _multi_cat_cond(categorias)
        prov_cond, prov_params = _multi_prov_cond(proveedores)
        sub_cond,  sub_params  = _multi_sub_cond(subgrupos)
        marc_cond, marc_params = _multi_marc_cond(marcas)

        # Dimensión de agrupamiento (más específica primero)
        if marcas:
            grp_expr  = "dp.marca"
            grp_cond  = marc_cond
            grp_params = marc_params
            up_cond   = f"{cat_cond} {prov_cond} {sub_cond}"
            up_params = cat_params + prov_params + sub_params
            group_by  = "marca"
        elif subgrupos:
            grp_expr  = "dp.subgrupo_descripcion"
            grp_cond  = sub_cond
            grp_params = sub_params
            up_cond   = f"{cat_cond} {prov_cond}"
            up_params = cat_params + prov_params
            group_by  = "subgrupo"
        elif proveedores:
            grp_expr  = "UPPER(dp.proveedor)"
            grp_cond  = prov_cond
            grp_params = prov_params
            up_cond   = cat_cond
            up_params = cat_params
            group_by  = "proveedor"
        elif categorias:
            grp_expr = """CASE dp.linea
                WHEN 'ALIMENTOS'            THEN 'Alimentos'
                WHEN 'APEGO'                THEN 'Apego'
                WHEN 'BEBIDAS ALC'          THEN 'Licores'
                WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
                ELSE 'Sin Clasificar'
            END"""
            grp_cond  = cat_cond
            grp_params = cat_params
            up_cond   = ""
            up_params = []
            group_by  = "categoria"
        else:
            # Sin filtros → desglosar por categoría como vista general
            grp_expr = """CASE dp.linea
                WHEN 'ALIMENTOS'            THEN 'Alimentos'
                WHEN 'APEGO'                THEN 'Apego'
                WHEN 'BEBIDAS ALC'          THEN 'Licores'
                WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
                ELSE 'Sin Clasificar'
            END"""
            grp_cond   = ""
            grp_params = []
            up_cond    = ""
            up_params  = []
            group_by   = "categoria"

        all_cond      = f"{up_cond} {grp_cond}"
        all_params    = up_params + grp_params
        grp_by_clause = f"GROUP BY {grp_expr}"
        order_clause  = "ORDER BY venta_neta DESC"

        # Ventas mes actual
        sql_v = f"""
            SELECT {grp_expr} AS name,
                   COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                   COALESCE(SUM(fv.venta_neta), 0) AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {all_cond}
            {grp_by_clause}
            {order_clause}
        """
        _, v_rows = _run_dw_query(sql_v, [anho, mes] + canal_param + all_params)

        # Ventas mes anterior
        ant_map = {}
        try:
            sql_ant = f"""
                SELECT {grp_expr} AS name,
                       COALESCE(SUM(fv.cantidad), 0)   AS cantidad_ant,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_neta_ant
                FROM dw.fact_ventas fv
                JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
                JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND ({ciudad_cond}) {canal_cond} {all_cond}
                {grp_by_clause}
            """
            _, ant_rows = _run_dw_query(sql_ant, [prev_anho, prev_mes] + canal_param + all_params)
            ant_map = {r['name']: r for r in ant_rows}
        except Exception:
            pass

        # Presupuesto mes actual
        ppto_map = {}
        try:
            sql_p = f"""
                SELECT {grp_expr} AS name,
                       COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS ppto_bs,
                       COALESCE(SUM(fp.cantidad_presupuestada), 0)   AS ppto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {all_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                {grp_by_clause}
            """
            _, p_rows = _run_dw_query(sql_p, [anho, mes] + canal_param + all_params + [anho, mes])
            ppto_map = {r['name']: r for r in p_rows}
        except Exception:
            pass

        result = []
        for row in v_rows:
            name     = row['name'] or '—'
            cant     = int(row['cantidad'] or 0)
            vn       = float(row['venta_neta'] or 0)
            ant      = ant_map.get(name, {})
            cant_ant = int(ant.get('cantidad_ant') or 0)
            vn_ant   = float(ant.get('venta_neta_ant') or 0)
            pp       = ppto_map.get(name, {})
            ppto_bs  = float(pp.get('ppto_bs') or 0)
            ppto_uds = int(pp.get('ppto_uds') or 0)

            pct_cumpl    = round(vn / ppto_bs * 100, 1) if ppto_bs > 0 else None
            gap_bs       = round(vn - ppto_bs, 0) if ppto_bs > 0 else None
            pct_camb_bs  = round((vn - vn_ant)   / vn_ant   * 100, 1) if vn_ant   > 0 else None
            pct_camb_uds = round((cant - cant_ant)/ cant_ant * 100, 1) if cant_ant > 0 else None

            result.append({
                'name':          name,
                'cantidad':      cant,
                'venta_neta':    vn,
                'ppto_bs':       ppto_bs,
                'ppto_uds':      ppto_uds,
                'pct_cumpl':     pct_cumpl,
                'gap_bs':        gap_bs,
                'cantidad_ant':  cant_ant,
                'venta_neta_ant': vn_ant,
                'pct_camb_bs':   pct_camb_bs,
                'pct_camb_uds':  pct_camb_uds,
            })

        return JsonResponse({
            'success':   True,
            'data':      result,
            'group_by':  group_by,
            'prev_anho': prev_anho,
            'prev_mes':  prev_mes,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('new-nacional')
def dashboard_new_nacional_skus(request):
    """
    SKUs con datos del mes actual y anterior. Acepta filtros multi-valor.
    Params (multi-value): categoria[], proveedor[], subgrupo[], marca[]
    Params: regional, canal, anho, mes, limit
    """
    try:
        regional    = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        canal       = _safe_str(request.GET.get('canal', ''))
        categorias  = [s for s in request.GET.getlist('categoria') if s]
        proveedores = [s for s in request.GET.getlist('proveedor') if s]
        subgrupos   = [s for s in request.GET.getlist('subgrupo') if s]
        marcas      = [s for s in request.GET.getlist('marca') if s]
        anho        = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes         = _safe_int(request.GET.get('mes'),  datetime.now().month)
        limit       = min(_safe_int(request.GET.get('limit'), 500), 1000)
        if regional not in REGIONALES_VALID:
            regional = 'nacional'

        prev_anho, prev_mes = _prev_period(anho, mes)
        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        canal_param = [canal] if canal else []

        cat_cond,  cat_params  = _multi_cat_cond(categorias)
        prov_cond, prov_params = _multi_prov_cond(proveedores)
        sub_cond,  sub_params  = _multi_sub_cond(subgrupos)
        marc_cond, marc_params = _multi_marc_cond(marcas)
        filter_cond   = f"{cat_cond} {prov_cond} {sub_cond} {marc_cond}"
        filter_params = cat_params + prov_params + sub_params + marc_params

        # Ventas mes actual
        sql_v = f"""
            SELECT dp.producto_codigo_erp          AS codigo,
                   dp.producto_nombre              AS producto,
                   COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                   COALESCE(SUM(fv.venta_neta), 0) AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {filter_cond}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY venta_neta DESC
            LIMIT %s
        """
        _, v_rows = _run_dw_query(sql_v, [anho, mes] + canal_param + filter_params + [limit])

        # Ventas mes anterior
        ant_map = {}
        try:
            sql_ant = f"""
                SELECT dp.producto_codigo_erp          AS codigo,
                       COALESCE(SUM(fv.cantidad), 0)   AS cantidad_ant,
                       COALESCE(SUM(fv.venta_neta), 0) AS venta_neta_ant
                FROM dw.fact_ventas fv
                JOIN dw.dim_fecha    df ON fv.fecha_sk    = df.fecha_sk
                JOIN dw.dim_vendedor dv ON fv.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fv.producto_sk = dp.producto_sk
                WHERE df.anho = %s AND df.mes_numero = %s
                  AND ({ciudad_cond}) {canal_cond} {filter_cond}
                GROUP BY dp.producto_codigo_erp
            """
            _, ant_rows = _run_dw_query(sql_ant, [prev_anho, prev_mes] + canal_param + filter_params)
            ant_map = {r['codigo']: r for r in ant_rows}
        except Exception:
            pass

        # Presupuesto
        ppto_map = {}
        try:
            sql_p = f"""
                SELECT dp.producto_codigo_erp                        AS codigo,
                       COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto,
                       COALESCE(SUM(fp.cantidad_presupuestada), 0)   AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_vendedor dv ON fp.vendedor_sk = dv.vendedor_sk
                JOIN dw.dim_producto dp ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND ({ciudad_cond}) {canal_cond} {filter_cond}
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.producto_codigo_erp
            """
            _, p_rows = _run_dw_query(sql_p, [anho, mes] + canal_param + filter_params + [anho, mes])
            ppto_map = {r['codigo']: r for r in p_rows}
        except Exception:
            pass

        result = []
        for row in v_rows:
            codigo   = row['codigo']
            cant     = int(row['cantidad'] or 0)
            vn       = float(row['venta_neta'] or 0)
            ant      = ant_map.get(codigo, {})
            cant_ant = int(ant.get('cantidad_ant') or 0)
            vn_ant   = float(ant.get('venta_neta_ant') or 0)
            pp       = ppto_map.get(codigo, {})
            ppto_bs  = float(pp.get('presupuesto') or 0)
            ppto_uds = int(pp.get('presupuesto_uds') or 0)

            pct_cumpl = round(vn / ppto_bs * 100, 1) if ppto_bs > 0 else None
            gap_pct   = round((vn / ppto_bs - 1) * 100, 1) if ppto_bs > 0 else None
            pct_camb_bs  = round((vn - vn_ant)    / vn_ant   * 100, 1) if vn_ant   > 0 else None
            pct_camb_uds = round((cant - cant_ant) / cant_ant * 100, 1) if cant_ant > 0 else None

            result.append({
                'codigo':         codigo,
                'producto':       row['producto'],
                'cantidad':       cant,
                'venta_neta':     vn,
                'presupuesto':    ppto_bs,
                'presupuesto_uds': ppto_uds,
                'pct_cumpl':      pct_cumpl,
                'gap_pct':        gap_pct,
                'cantidad_ant':   cant_ant,
                'venta_neta_ant': vn_ant,
                'pct_camb_bs':    pct_camb_bs,
                'pct_camb_uds':   pct_camb_uds,
            })

        return JsonResponse({
            'success':   True,
            'data':      result,
            'prev_anho': prev_anho,
            'prev_mes':  prev_mes,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('unidades-vendidas', 'unidades-supervisores')
def dashboard_unidades_vendedor_sku(request):
    """
    SKUs vendidos por un vendedor especÃ­fico, filtrado por categorÃ­a y opcionalmente subgrupo.
    Params: regional, canal, vendedor_sk (int), categoria, subgrupo, anho, mes, limit
    """
    try:
        regional    = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal       = _safe_str(request.GET.get('canal', ''))
        vendedor_sk = _safe_str(request.GET.get('vendedor_sk', ''))
        categoria   = _safe_str(request.GET.get('categoria', ''))
        subgrupo    = _safe_str(request.GET.get('subgrupo', ''))
        proveedor   = _safe_str(request.GET.get('proveedor', ''))
        anho        = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes         = _safe_int(request.GET.get('mes'),  datetime.now().month)
        limit       = min(_safe_int(request.GET.get('limit'), 300), 500)

        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)
        if not vendedor_sk:
            return JsonResponse({'success': False, 'error': 'vendedor_sk requerido'}, status=400)

        ciudad_cond    = _regional_filter(regional)
        canal_cond     = "AND dv.canal_rrhh = %s" if canal else ""
        proveedor_cond = "AND UPPER(dp.proveedor) = UPPER(%s)" if proveedor else ""
        base_params = [anho, mes] + ([canal] if canal else [])
        cat_cond, params_v = _unidades_cat_params(categoria, base_params)
        sub_cond = ""
        if subgrupo:
            sub_cond = "AND dp.subgrupo_descripcion = %s"
            params_v = params_v + [subgrupo]
        if proveedor:
            params_v = params_v + [proveedor]
        params_v = params_v + [int(vendedor_sk), limit]

        sql_v = f"""
            SELECT
                dp.producto_codigo_erp          AS codigo,
                dp.producto_nombre              AS producto,
                COALESCE(SUM(fv.cantidad), 0)   AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0) AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {sub_cond} {proveedor_cond}
              AND dv.vendedor_sk = %s
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY cantidad DESC
            LIMIT %s
        """
        _, v_rows = _run_dw_query(sql_v, params_v)

        # Presupuesto por SKU para este vendedor
        ppto_map = {}
        try:
            sql_p = f"""
                SELECT
                    dp.producto_codigo_erp                            AS codigo,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0)     AS presupuesto,
                    COALESCE(SUM(fp.cantidad_presupuestada), 0)        AS presupuesto_uds
                FROM dw.fact_presupuesto fp
                JOIN dw.dim_producto             dp  ON fp.producto_sk = dp.producto_sk
                WHERE fp.anho = %s AND fp.mes = %s
                  AND fp.vendedor_sk = %s
                  AND fp.version_sk = (SELECT MAX(version_sk) FROM dw.dim_presupuesto_version WHERE anho = %s AND mes = %s)
                GROUP BY dp.producto_codigo_erp
            """
            _, p_rows = _run_dw_query(sql_p, [anho, mes, int(vendedor_sk), anho, mes])
            ppto_map = {
                r['codigo']: (float(r['presupuesto'] or 0), float(r['presupuesto_uds'] or 0))
                for r in p_rows
            }
        except Exception:
            pass

        result = []
        for row in v_rows:
            cod  = row['codigo']
            cant = int(row['cantidad'] or 0)
            vn   = float(row['venta_neta'] or 0)
            ppto, ppto_uds = ppto_map.get(cod, (0, 0))
            result.append({
                'codigo':          cod,
                'producto':        row['producto'],
                'cantidad':        cant,
                'venta_neta':      vn,
                'presupuesto':     ppto,
                'presupuesto_uds': int(ppto_uds),
                'porcentaje':      round(vn   / ppto     * 100, 1) if ppto     > 0 else None,
                'porcentaje_uds':  round(cant / ppto_uds * 100, 1) if ppto_uds > 0 else None,
            })
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('unidades-vendidas')
def dashboard_unidades_por_vendedor(request):
    """
    Vendedores que vendieron en una sub-categorÃ­a (o categorÃ­a) dada.
    Params: regional, canal, categoria, subgrupo, anho, mes
    """
    try:
        regional  = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
        canal     = _safe_str(request.GET.get('canal', ''))
        categoria = _safe_str(request.GET.get('categoria', ''))
        subgrupo  = _safe_str(request.GET.get('subgrupo', ''))
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        if regional not in REGIONALES_VALID:
            return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

        ciudad_cond = _regional_filter(regional)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
        base_params = [anho, mes] + ([canal] if canal else [])
        cat_cond, params = _unidades_cat_params(categoria, base_params)
        sub_cond = ""
        if subgrupo:
            sub_cond = "AND dp.subgrupo_descripcion = %s"
            params = params + [subgrupo]

        sql = f"""
            SELECT
                dv.vendedor_sk,
                dv.vendedor_nombre                 AS vendedor,
                COALESCE(SUM(fv.cantidad), 0)      AS cantidad,
                COALESCE(SUM(fv.venta_neta), 0)    AS venta_neta
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha    df   ON fv.fecha_sk    = df.fecha_sk
            JOIN dw.dim_vendedor dv   ON fv.vendedor_sk = dv.vendedor_sk
            JOIN dw.dim_producto dp   ON fv.producto_sk = dp.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond}) {canal_cond} {cat_cond} {sub_cond}
            GROUP BY dv.vendedor_sk, dv.vendedor_nombre
            ORDER BY cantidad DESC
        """
        _, rows = _run_dw_query(sql, params)
        result = [
            {
                'vendedor_sk': r['vendedor_sk'],
                'vendedor':    r['vendedor'],
                'cantidad':    int(r['cantidad'] or 0),
                'venta_neta':  float(r['venta_neta'] or 0),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': result})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD â€" PROVEEDORES
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_PROV_PERM_MAP = {
    'PEPSICO': 'pepsico',
    'SOFTYS':  'softys',
    'DMUJER':  'dmujer',
    'APEGO':   'apego',
    'COLHER':  'colher',
}


def _check_proveedor_perm(request, proveedor):
    """Verifica que el usuario tenga permiso para el dashboard del proveedor."""
    perm_id = _PROV_PERM_MAP.get(proveedor.upper())
    if not perm_id:
        return False
    return _has_dashboard_perm(request.user, perm_id)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_proveedor_kpis(request):
    """Total ventas + desglose por regional para un proveedor. Params: proveedor, anho, mes."""
    try:
        proveedor = _safe_str(request.GET.get('proveedor', ''), 50).upper()
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err       = _validate_anho_mes(anho, mes)
        if err: return err

        if not proveedor:
            return JsonResponse({'success': False, 'error': 'ParÃ¡metro proveedor requerido'}, status=400)
        if not _check_proveedor_perm(request, proveedor):
            return JsonResponse({'success': False, 'error': 'Sin acceso a este dashboard'}, status=403)

        prov_filter = "(UPPER(dp.proveedor) = %s OR UPPER(dp.cat_comercial) = %s)"

        sql_total = f"""
            SELECT COALESCE(SUM(fv.total), 0)              AS total,
                   COUNT(DISTINCT fv.numero_venta)         AS pedidos,
                   COUNT(DISTINCT fv.cliente_sk)           AS clientes
            FROM dw.fact_ventas fv
            JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
            WHERE df.anho = %s AND df.mes_numero = %s AND {prov_filter}
        """
        _, rows_total = _run_dw_query(sql_total, [anho, mes, proveedor, proveedor])

        scz  = _ciudad_case('dv.ciudad', 'santa_cruz')
        cbba = _ciudad_case('dv.ciudad', 'cochabamba')
        lpz  = _ciudad_case('dv.ciudad', 'la_paz')
        sql_reg = f"""
            SELECT
                CASE
                    WHEN {scz}  THEN 'Santa Cruz'
                    WHEN {cbba} THEN 'Cochabamba'
                    WHEN {lpz}  THEN 'La Paz'
                    ELSE 'Otras'
                END                              AS regional,
                COALESCE(SUM(fv.total), 0)       AS total
            FROM dw.fact_ventas fv
            JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
            JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
            WHERE df.anho = %s AND df.mes_numero = %s AND {prov_filter}
            GROUP BY regional ORDER BY total DESC
        """
        _, rows_reg = _run_dw_query(sql_reg, [anho, mes, proveedor, proveedor])

        kpis = rows_total[0] if rows_total else {'total': 0, 'pedidos': 0, 'clientes': 0}
        return JsonResponse({'success': True, 'data': {
            'total':      float(kpis.get('total', 0) or 0),
            'pedidos':    int(kpis.get('pedidos', 0) or 0),
            'clientes':   int(kpis.get('clientes', 0) or 0),
            'regionales': rows_reg,
        }})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_proveedor_por_marca(request):
    """Ventas por articulo (producto_nombre). Params: proveedor, anho, mes."""
    try:
        proveedor = _safe_str(request.GET.get('proveedor', ''), 50).upper()
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err       = _validate_anho_mes(anho, mes)
        if err: return err

        if not proveedor:
            return JsonResponse({'success': False, 'error': 'ParÃ¡metro proveedor requerido'}, status=400)
        if not _check_proveedor_perm(request, proveedor):
            return JsonResponse({'success': False, 'error': 'Sin acceso a este dashboard'}, status=403)

        sql = """
            SELECT cd.canal                            AS marca,
                   COALESCE(SUM(fv.total), 0)          AS total,
                   COALESCE(SUM(fv.cantidad), 0)       AS cantidad
            FROM dw.fact_ventas fv
            JOIN dw.dim_producto      dp ON dp.producto_sk    = fv.producto_sk
            JOIN dw.dim_fecha         df ON df.fecha_sk       = fv.fecha_sk
            JOIN dw.dim_cliente       dc ON dc.cliente_sk     = fv.cliente_sk
            JOIN dual.dim_cliente_dual cd ON cd.codigo_cliente = dc.cliente_codigo_erp
                                         AND cd.es_actual = TRUE
            WHERE df.anho = %s AND df.mes_numero = %s
              AND (UPPER(dp.proveedor) = %s OR UPPER(dp.cat_comercial) = %s)
            GROUP BY cd.canal
            ORDER BY total DESC
        """
        _, rows = _run_dw_query(sql, [anho, mes, proveedor, proveedor])
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_proveedor_tabla(request):
    """Detalle completo de ventas por proveedor. Params: proveedor, anho, mes."""
    try:
        proveedor = _safe_str(request.GET.get('proveedor', ''), 50).upper()
        anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)
        err       = _validate_anho_mes(anho, mes)
        if err: return err

        if not proveedor:
            return JsonResponse({'success': False, 'error': 'ParÃ¡metro proveedor requerido'}, status=400)
        if not _check_proveedor_perm(request, proveedor):
            return JsonResponse({'success': False, 'error': 'Sin acceso a este dashboard'}, status=403)

        sql = """
            SELECT
                dv.canal_rrhh AS canal,
                da.ciudad,
                df.mes_nombre,
                dp.proveedor,
                dp.cat_comercial    AS marca,
                fv.numero_venta,
                df.fecha_completa,
                dc.cliente_codigo_erp,
                dp.grupo_descripcion,
                dp.clase_descripcion,
                dp.producto_nombre,
                dp.unidad_medida,
                fv.cantidad,
                fv.total,
                dv.vendedor_nombre,
                INITCAP(dv.supervisor) AS supervisor_nombre
            FROM dw.fact_ventas fv
            JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
            JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_almacen  da ON da.almacen_sk  = fv.almacen_sk
            JOIN dw.dim_cliente  dc ON dc.cliente_sk  = fv.cliente_sk
            JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              AND (UPPER(dp.proveedor) = %s OR UPPER(dp.cat_comercial) = %s)
            ORDER BY fv.numero_venta, dp.producto_nombre
        """
        _, rows = _run_dw_query(sql, [anho, mes, proveedor, proveedor])

        response: dict = {'success': True, 'data': rows}

        if proveedor == 'SOFTYS':
            sql_ppto = """
                SELECT
                    dp.producto_codigo_erp  AS cod_producto,
                    dp.producto_nombre,
                    COALESCE(dv.canal_rrhh, '') AS canal,
                    COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto
                FROM dw.dim_producto dp
                LEFT JOIN dw.fact_presupuesto fp
                    ON fp.producto_sk = dp.producto_sk
                    AND fp.anho = %s AND fp.mes = %s
                    AND fp.version_sk = (
                        SELECT MAX(version_sk) FROM dw.dim_presupuesto_version
                        WHERE anho = %s AND mes = %s
                    )
                LEFT JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fp.vendedor_sk
                WHERE (UPPER(dp.proveedor) = 'SOFTYS' OR UPPER(dp.cat_comercial) = 'SOFTYS')
                GROUP BY dp.producto_codigo_erp, dp.producto_nombre, dv.canal_rrhh
                ORDER BY dp.producto_nombre, dv.canal_rrhh NULLS LAST
            """
            _, ppto_rows = _run_dw_query(sql_ppto, [anho, mes, anho, mes])
            response['presupuesto_por_sku'] = ppto_rows

        return JsonResponse(response)
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  EXPORTACIONES XLSX
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_XLSX_HEADERS = [
    "No.Venta", "Fecha", "Cliente", "Nombre Cliente",
    "Grupo", "Descripcion Grupo",
    "SubGrupo", "Descripcion SubGrupo",
    "Clase", "Descripcion Clase",
    "SubClase", "Descripcion SubClase",
    "Articulo", "Descripcion Articulo",
    "U/M", "Cantidad", "Precio", "SubTotal", "Descuento", "Total",
    "Local", "Descripcion Local",
    "Almacen", "Descripcion Almacen",
    "Ruta", "Descripcion Ruta",
    "Vendedor", "Descripcion Vendedor",
    "Distribuidor", "Descripcion Distribuidor",
    "Zona", "Descripcion Zona",
    "Componente", "ICE", "Venta Neta", "Pago",
    "Exhibidor", "Clase", "Ruta", "Punto Frio",
]


def exportar_ventas_combo_armado(request):
    """
    Exporta ventas a XLSX. Acepta token via header Authorization o query param _t.
    Params: fecha_desde (YYYY-MM-DD), fecha_hasta (YYYY-MM-DD)
    """
    token_str = ""
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if auth_header.startswith("Token "):
        token_str = auth_header[6:]
    else:
        token_str = request.GET.get("_t", "")

    if not token_str:
        return JsonResponse({"error": "No autorizado"}, status=401)

    try:
        from rest_framework.authtoken.models import Token as DRFToken
        token_obj = DRFToken.objects.select_related("user").get(key=token_str)
        if not token_obj.user.is_active:
            return JsonResponse({"error": "Usuario inactivo"}, status=401)
    except Exception:
        return JsonResponse({"error": "Token invÃ¡lido"}, status=401)

    if not _has_dashboard_perm(token_obj.user, 'descargas'):
        return JsonResponse({"error": "Sin acceso a descargas"}, status=403)

    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")

    if not fecha_desde or not fecha_hasta:
        return JsonResponse({"success": False, "error": "ParÃ¡metros fecha_desde y fecha_hasta requeridos"}, status=400)

    try:
        datetime.strptime(fecha_desde, "%Y-%m-%d")
        datetime.strptime(fecha_hasta, "%Y-%m-%d")
    except ValueError:
        return JsonResponse({"success": False, "error": "Formato de fecha invÃ¡lido. Use YYYY-MM-DD"}, status=400)

    sql = """
        SELECT
            fv.numero_venta,
            fv.fecha_venta,
            dc.cliente_codigo_erp,
            COALESCE(dc.cliente_nombre, ''),
            COALESCE(dp.grupo_codigo::INTEGER, NULL),
            COALESCE(dp.grupo_descripcion, ''),
            COALESCE(dp.subgrupo_codigo::INTEGER, NULL),
            COALESCE(dp.subgrupo_descripcion, ''),
            COALESCE(dp.clase_codigo::INTEGER, NULL),
            COALESCE(dp.clase_descripcion, ''),
            COALESCE(dp.subclase_codigo::INTEGER, NULL),
            COALESCE(dp.subclase_descripcion, ''),
            dp.producto_codigo_erp,
            COALESCE(dp.producto_nombre, ''),
            COALESCE(dp.unidad_medida, ''),
            fv.cantidad,
            fv.precio_unitario,
            fv.subtotal,
            fv.descuento,
            fv.total,
            COALESCE(dl.local_codigo_erp::INTEGER, NULL),
            COALESCE(dl.local_nombre, ''),
            COALESCE(da.almacen_codigo_erp::INTEGER, NULL),
            COALESCE(da.almacen_nombre, ''),
            COALESCE(fv.ruta_codigo::INTEGER, NULL),
            COALESCE(fv.ruta_descripcion, ''),
            COALESCE(dv.vendedor_codigo_erp::INTEGER, NULL),
            COALESCE(dv.vendedor_nombre, ''),
            COALESCE(ddi.distribuidor_codigo_erp::INTEGER, NULL),
            COALESCE(ddi.distribuidor_nombre, ''),
            COALESCE(dz.zona_codigo_erp::INTEGER, NULL),
            COALESCE(dz.zona_descripcion, ''),
            COALESCE(dp.componente, ''),
            fv.ice,
            fv.venta_neta,
            fv.pago
        FROM dw.fact_ventas                  fv
        JOIN      dw.dim_producto        dp  ON fv.producto_sk     = dp.producto_sk
        JOIN      dw.dim_cliente         dc  ON fv.cliente_sk      = dc.cliente_sk
        LEFT JOIN dw.dim_vendedor        dv  ON fv.vendedor_sk     = dv.vendedor_sk
        LEFT JOIN dw.dim_local           dl  ON fv.local_sk        = dl.local_sk
        LEFT JOIN dw.dim_almacen         da  ON fv.almacen_sk      = da.almacen_sk
        LEFT JOIN dw.dim_distribuidor   ddi  ON fv.distribuidor_sk = ddi.distribuidor_sk
        LEFT JOIN dw.dim_zona            dz  ON fv.zona_sk         = dz.zona_sk
        WHERE SUBSTR(fv.fecha_venta, 7, 4) || '-' || SUBSTR(fv.fecha_venta, 4, 2) || '-' || SUBSTR(fv.fecha_venta, 1, 2)
              BETWEEN %s AND %s
        ORDER BY SUBSTR(fv.fecha_venta, 7, 4) || '-' || SUBSTR(fv.fecha_venta, 4, 2) || '-' || SUBSTR(fv.fecha_venta, 1, 2),
                 fv.numero_venta
    """

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Ventas Combo Armado")

    bold = Font(bold=True)
    header_row = []
    for h in _XLSX_HEADERS:
        c = openpyxl.cell.WriteOnlyCell(ws, value=h)
        c.font = bold
        header_row.append(c)
    ws.append(header_row)

    try:
        with connections['dw'].cursor() as cursor:
            cursor.execute(sql, [fecha_desde, fecha_hasta])
            while True:
                batch = cursor.fetchmany(2000)
                if not batch:
                    break
                for row in batch:
                    ws.append(list(row) + ["", "", "", ""])
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({"success": False, "error": "Error interno del servidor"}, status=500)

    buffer = io.BytesIO()
    wb.save(buffer)
    file_size = buffer.tell()
    buffer.seek(0)

    def _chunks(buf, size=65536):
        chunk = buf.read(size)
        while chunk:
            yield chunk
            chunk = buf.read(size)

    filename = f"ventas_combo_armado_{fecha_desde}_{fecha_hasta}.xlsx"
    response = StreamingHttpResponse(
        _chunks(buffer),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Encoding"] = "identity"
    response["X-File-Size"] = str(file_size)
    return response


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  CANALES DISPONIBLES (lista para selectores)
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_canales_lista(request):
    """Devuelve la lista de canal_rrhh distintos presentes en dim_vendedor."""
    cached = cache.get('canales_lista')
    if cached is not None:
        return JsonResponse({'success': True, 'data': cached})
    try:
        _, rows = _run_dw_query(
            "SELECT DISTINCT canal_rrhh FROM dw.dim_vendedor "
            "WHERE canal_rrhh IS NOT NULL ORDER BY canal_rrhh"
        )
        canales = [r['canal_rrhh'] for r in rows]
        cache.set('canales_lista', canales, 3600)
        return JsonResponse({'success': True, 'data': canales})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD INFORMACIÃ"N RUTAS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_marcas_lista(_request):
    """Lista de marcas activas en el DW, para filtro de cobertura."""
    cached = cache.get('clase_lista_rutas')
    if cached is not None:
        return JsonResponse({'success': True, 'data': cached})
    try:
        sql = """
            SELECT DISTINCT dp.clase_descripcion
            FROM dw.dim_producto dp
            WHERE dp.clase_descripcion IS NOT NULL
              AND dp.clase_descripcion != ''
              AND dp.clase_descripcion != 'PENDIENTES'
            ORDER BY dp.clase_descripcion
        """
        _, rows = _run_dw_query(sql, [])
        data = [r['clase_descripcion'] for r in rows]
        cache.set('clase_lista_rutas', data, 1800)   # 30 min
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas(request):
    """
    Tabla de rutas con cobertura del mes.
    Params: regional, canal, supervisor, dia, marca, anho, mes
    """
    is_admin = _is_admin(request.user)
    profile  = _get_or_create_profile(request.user)
    cargo    = (profile.cargo or '').strip()

    dia   = _safe_str(request.GET.get('dia',   'Todos'), 10)
    marca = _safe_str(request.GET.get('marca', ''),       80)
    anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)

    if is_admin:
        regional_raw = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        regional     = regional_raw if regional_raw in REGIONALES_VALID else 'nacional'
        canal        = _safe_str(request.GET.get('canal',      'Todos'), 30)
        supervisor   = _safe_str(request.GET.get('supervisor', 'Todos'), 100)
    elif cargo == 'Gerente Regional':
        regional     = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'nacional')
        canal        = _safe_str(request.GET.get('canal', 'Todos'), 30)
        supervisor   = _safe_str(request.GET.get('supervisor', 'Todos'), 100)
    elif 'supervisor' in cargo.lower():
        regional     = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'nacional')
        canal        = (profile.canal or '').strip()
        supervisor   = f"{profile.user.first_name} {profile.user.last_name}".strip()
    else:
        regional_raw = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        regional     = regional_raw if regional_raw in REGIONALES_VALID else 'nacional'
        canal        = _safe_str(request.GET.get('canal',      'Todos'), 30)
        supervisor   = _safe_str(request.GET.get('supervisor', 'Todos'), 100)

    ciudad_cond   = _regional_filter(regional)
    regional_cond = "" if regional == 'nacional' else f"AND ({ciudad_cond})"
    canal_cond    = "AND dv.canal_rrhh = %s"               if canal and canal != 'Todos' else ""
    dia_cond      = "AND dp.dia = %s"                      if dia != 'Todos'             else ""
    sup_cond      = "AND UPPER(dv.supervisor) = UPPER(%s)" if supervisor and supervisor != 'Todos' else ""
    marca_cond    = "AND dprod.clase_descripcion = %s" if marca else ""

    params_main = [anho, mes]
    if marca:                          params_main.append(marca)
    if canal and canal != 'Todos':     params_main.append(canal)
    if dia != 'Todos':                 params_main.append(dia)
    if supervisor and supervisor != 'Todos': params_main.append(supervisor)

    try:
        sql = f"""
            WITH ventas_mes AS (
                SELECT DISTINCT fv.cliente_sk, dv2.canal_rrhh
                FROM dw.fact_ventas    fv
                JOIN dw.dim_fecha      df    ON df.fecha_sk     = fv.fecha_sk
                JOIN dw.dim_vendedor   dv2   ON dv2.vendedor_sk = fv.vendedor_sk
                JOIN dw.dim_producto   dprod ON dprod.producto_sk = fv.producto_sk
                WHERE df.anho = %s AND df.mes_numero = %s
                  {marca_cond}
            )
            SELECT
                dc.ruta,
                MAX(dp.vendedor)   AS vendedor,
                MAX(dv.supervisor) AS supervisor,
                (SELECT STRING_AGG(d2.dia, ', ' ORDER BY d2.dia)
                 FROM (SELECT DISTINCT dia FROM dual.dim_planificacion
                       WHERE ruta = dc.ruta AND es_actual = true) AS d2) AS dia,
                COUNT(DISTINCT dc.id_cliente)                                          AS total_clientes,
                COUNT(DISTINCT CASE WHEN vm.canal_rrhh = dv.canal_rrhh
                                    THEN dck.cliente_sk END)                           AS clientes_con_compra,
                ROUND(
                    COUNT(DISTINCT CASE WHEN vm.canal_rrhh = dv.canal_rrhh
                                        THEN dck.cliente_sk END)::NUMERIC
                    / NULLIF(COUNT(DISTINCT dc.id_cliente), 0) * 100
                , 1)                                                                   AS pct_cobertura
            FROM dual.dim_cliente_dual dc
            LEFT JOIN dual.dim_planificacion dp
                   ON dp.ruta = dc.ruta AND dp.es_actual = true
            LEFT JOIN dw.dim_vendedor dv
                   ON dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
                  AND dv.es_vendedor_actual = true
            LEFT JOIN dw.dim_cliente dck
                   ON dck.cliente_codigo_erp = dc.codigo_cliente
                  AND dck.es_cliente_actual = true
            LEFT JOIN ventas_mes vm ON vm.cliente_sk = dck.cliente_sk
            WHERE dc.es_actual = true
              {regional_cond}
              {canal_cond}
              {dia_cond}
              {sup_cond}
            GROUP BY dc.ruta
            ORDER BY pct_cobertura DESC NULLS LAST, dc.ruta
        """
        _, rows = _run_dw_query(sql, params_main)

        # Lista de supervisores para el filtro (no afectada por filtro de supervisor)
        params_sup = []
        if canal and canal != 'Todos': params_sup.append(canal)
        sql_sup = f"""
            SELECT DISTINCT dv.supervisor
            FROM dual.dim_planificacion dp
            JOIN dw.dim_vendedor dv
                   ON dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
                  AND dv.es_vendedor_actual = true
            WHERE dp.es_actual = true
              AND dv.supervisor IS NOT NULL
              {regional_cond}
              {canal_cond}
            ORDER BY dv.supervisor
        """
        _, sup_rows = _run_dw_query(sql_sup, params_sup)
        supervisores = [r['supervisor'] for r in sup_rows if r.get('supervisor')]

        return JsonResponse({'success': True, 'data': rows, 'supervisores': supervisores})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas_detalle(request):
    """
    Ventas y pedidos semanales de una ruta especÃ­fica.
    Params: ruta, canal, marca, anho, mes
    """
    ruta  = _safe_str(request.GET.get('ruta',  ''), 100)
    canal = _safe_str(request.GET.get('canal', ''), 30)
    marca = _safe_str(request.GET.get('marca', ''), 80)
    anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)

    if not ruta:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro ruta requerido'}, status=400)

    try:
        canal_cond = "AND dv.canal_rrhh = %s" if canal else ""
        marca_cond = "AND dp.clase_descripcion = %s" if marca else ""
        params_det = [anho, mes, ruta] + ([canal] if canal else []) + ([marca] if marca else [])
        sql = f"""
            SELECT
                CEIL(df.dia_numero / 7.0)::INT                      AS semana,
                COUNT(DISTINCT fv.numero_venta)                      AS pedidos,
                COALESCE(SUM(fv.venta_neta), 0)                     AS venta_neta
            FROM dw.fact_ventas       fv
            JOIN dw.dim_fecha         df  ON df.fecha_sk    = fv.fecha_sk
            JOIN dw.dim_vendedor      dv  ON dv.vendedor_sk = fv.vendedor_sk
            JOIN dw.dim_producto      dp  ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_cliente       dck ON dck.cliente_sk = fv.cliente_sk
            JOIN dual.dim_cliente_dual dcd
                   ON dcd.codigo_cliente = dck.cliente_codigo_erp
                  AND dcd.es_actual = true
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dcd.ruta = %s
              {canal_cond}
              {marca_cond}
            GROUP BY semana
            ORDER BY semana
        """
        _, rows = _run_dw_query(sql, params_det)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas_clientes(request):
    """
    Compras semanales por cliente para una ruta.
    Params: ruta, canal, marca, anho, mes
    """
    ruta  = _safe_str(request.GET.get('ruta',  ''), 100)
    canal = _safe_str(request.GET.get('canal', ''), 30)
    marca = _safe_str(request.GET.get('marca', ''), 80)
    anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)

    if not ruta:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro ruta requerido'}, status=400)

    try:
        canal_cond = "AND dv.canal_rrhh = %s" if canal else ""
        marca_cond = "AND dp.clase_descripcion = %s" if marca else ""
        # params: subquery (anho, mes, canal?, marca?) then outer WHERE (ruta)
        params = (
            [anho, mes]
            + ([canal] if canal else [])
            + ([marca] if marca else [])
            + [ruta]
        )
        sql = f"""
            SELECT
                dcd.codigo_cliente                                                  AS codigo_cliente,
                INITCAP(dck.cliente_nombre)                                         AS nombre_cliente,
                ventas.semana                                                       AS semana,
                COALESCE(ventas.bs, 0)                                             AS bs,
                COALESCE(ventas.pedidos, 0)                                        AS pedidos
            FROM dual.dim_cliente_dual   dcd
            JOIN dw.dim_cliente          dck ON dck.cliente_codigo_erp = dcd.codigo_cliente
                                            AND dck.es_cliente_actual = true
            LEFT JOIN (
                SELECT
                    fv.cliente_sk,
                    CEIL(df.dia_numero / 7.0)::INT                                  AS semana,
                    COALESCE(SUM(fv.venta_neta), 0)                                AS bs,
                    COUNT(DISTINCT fv.numero_venta)                                AS pedidos
                FROM dw.fact_ventas       fv
                JOIN dw.dim_fecha         df  ON df.fecha_sk    = fv.fecha_sk
                JOIN dw.dim_vendedor      dv  ON dv.vendedor_sk = fv.vendedor_sk
                JOIN dw.dim_producto      dp  ON dp.producto_sk = fv.producto_sk
                WHERE df.anho = %s AND df.mes_numero = %s
                  {canal_cond}
                  {marca_cond}
                GROUP BY fv.cliente_sk, semana
            ) ventas                         ON ventas.cliente_sk = dck.cliente_sk
            WHERE dcd.ruta = %s
              AND dcd.es_actual = true
            ORDER BY dck.cliente_nombre, ventas.semana NULLS LAST
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas_cliente_detalle(request):
    """
    Detalle de SKUs comprados por un cliente en una ruta (todo el mes o una semana).
    Params: ruta, codigo_cliente, anho, mes, semana (opcional)
    """
    ruta           = _safe_str(request.GET.get('ruta',           ''), 100)
    codigo_cliente = _safe_str(request.GET.get('codigo_cliente', ''), 50)
    anho           = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes            = _safe_int(request.GET.get('mes'),  datetime.now().month)
    semana_raw     = request.GET.get('semana', '')
    semana         = _safe_int(semana_raw, None) if semana_raw else None

    if not ruta or not codigo_cliente:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metros ruta y codigo_cliente requeridos'}, status=400)

    try:
        fecha_cond = ""
        params = [anho, mes, ruta, codigo_cliente]
        if semana is not None:
            fecha_cond = "AND CEIL(df.dia_numero / 7.0)::INT = %s"
            params.append(semana)

        sql = f"""
            SELECT
                dp.producto_nombre                                   AS sku,
                COALESCE(SUM(fv.venta_neta), 0)                    AS bs,
                COALESCE(SUM(fv.cantidad), 0)::INT                  AS unidades
            FROM dw.fact_ventas        fv
            JOIN dw.dim_fecha          df  ON df.fecha_sk    = fv.fecha_sk
            JOIN dw.dim_producto       dp  ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_cliente        dck ON dck.cliente_sk = fv.cliente_sk
                                          AND dck.es_cliente_actual = true
            JOIN dual.dim_cliente_dual dcd ON dcd.codigo_cliente = dck.cliente_codigo_erp
                                          AND dcd.es_actual = true
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dcd.ruta = %s
              AND dck.cliente_codigo_erp = %s
              {fecha_cond}
            GROUP BY dp.producto_nombre
            ORDER BY bs DESC
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas_categorias(request):
    """
    Ventas por categorÃ­a para una ruta, con % del total.
    Params: ruta, canal, marca, anho, mes
    """
    ruta  = _safe_str(request.GET.get('ruta',  ''), 100)
    canal = _safe_str(request.GET.get('canal', ''), 30)
    marca = _safe_str(request.GET.get('marca', ''), 80)
    anho  = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes   = _safe_int(request.GET.get('mes'),  datetime.now().month)

    if not ruta:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro ruta requerido'}, status=400)

    try:
        canal_cond = "AND dv.canal_rrhh = %s" if canal else ""
        marca_cond = "AND dp.clase_descripcion = %s" if marca else ""
        params = [anho, mes, ruta] + ([canal] if canal else []) + ([marca] if marca else [])
        sql = f"""
            SELECT
                {_CATEGORIA_CASE}                                         AS categoria,
                COUNT(DISTINCT fv.numero_venta)                           AS pedidos,
                COALESCE(SUM(fv.venta_neta), 0)                          AS venta_neta
            FROM dw.fact_ventas       fv
            JOIN dw.dim_fecha         df  ON df.fecha_sk    = fv.fecha_sk
            JOIN dw.dim_vendedor      dv  ON dv.vendedor_sk = fv.vendedor_sk
            JOIN dw.dim_producto      dp  ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_cliente       dck ON dck.cliente_sk = fv.cliente_sk
            JOIN dual.dim_cliente_dual dcd
                   ON dcd.codigo_cliente = dck.cliente_codigo_erp
                  AND dcd.es_actual = true
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dcd.ruta = %s
              {canal_cond}
              {marca_cond}
            GROUP BY {_CATEGORIA_CASE}
            ORDER BY venta_neta DESC
        """
        _, rows = _run_dw_query(sql, params)
        total_bs = sum(float(r['venta_neta']) for r in rows)
        for r in rows:
            r['pct'] = round(float(r['venta_neta']) / total_bs * 100, 1) if total_bs > 0 else 0
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('informacion-rutas')
def dashboard_informacion_rutas_skus(request):
    """
    Top SKUs de una categorÃ­a para una ruta, con % cobertura de clientes.
    Params: ruta, canal, categoria, marca, anho, mes
    """
    ruta      = _safe_str(request.GET.get('ruta',      ''), 100)
    canal     = _safe_str(request.GET.get('canal',     ''), 30)
    categoria = _safe_str(request.GET.get('categoria', ''), 50)
    marca     = _safe_str(request.GET.get('marca',     ''), 80)
    anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes       = _safe_int(request.GET.get('mes'),  datetime.now().month)

    if not ruta:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro ruta requerido'}, status=400)

    canal_cond = "AND dv.canal_rrhh = %s" if canal else ""
    marca_cond = "AND dp.marca = %s"       if marca else ""

    if categoria and categoria in _UNIDADES_CAT_LINEA:
        if categoria == 'Sin Clasificar':
            cat_cond   = f"AND {_SIN_CLASIFICAR_COND}"
            cat_params = []
        else:
            cat_cond   = "AND dp.linea = %s"
            cat_params = [_UNIDADES_CAT_LINEA[categoria]]
    else:
        cat_cond   = ""
        cat_params = []

    try:
        params = [ruta, anho, mes, ruta] + ([canal] if canal else []) + ([marca] if marca else []) + cat_params
        sql = f"""
            WITH total_ruta AS (
                SELECT COUNT(DISTINCT codigo_cliente) AS n
                FROM dual.dim_cliente_dual
                WHERE es_actual = true AND ruta = %s
            )
            SELECT
                dp.producto_codigo_erp                                     AS codigo,
                dp.producto_nombre                                         AS producto,
                COUNT(DISTINCT fv.numero_venta)                            AS pedidos,
                COALESCE(SUM(fv.venta_neta), 0)                            AS venta_neta,
                COUNT(DISTINCT dck.cliente_sk)                             AS clientes_con_sku,
                (SELECT n FROM total_ruta)                                 AS total_clientes,
                ROUND(
                    COUNT(DISTINCT dck.cliente_sk)::NUMERIC
                    / NULLIF((SELECT n FROM total_ruta), 0) * 100
                , 1)                                                       AS pct_cobertura
            FROM dw.fact_ventas       fv
            JOIN dw.dim_fecha         df  ON df.fecha_sk    = fv.fecha_sk
            JOIN dw.dim_vendedor      dv  ON dv.vendedor_sk = fv.vendedor_sk
            JOIN dw.dim_producto      dp  ON dp.producto_sk = fv.producto_sk
            JOIN dw.dim_cliente       dck ON dck.cliente_sk = fv.cliente_sk
            JOIN dual.dim_cliente_dual dcd
                   ON dcd.codigo_cliente = dck.cliente_codigo_erp
                  AND dcd.es_actual = true
            WHERE df.anho = %s AND df.mes_numero = %s
              AND dcd.ruta = %s
              {canal_cond}
              {marca_cond}
              {cat_cond}
            GROUP BY dp.producto_codigo_erp, dp.producto_nombre
            ORDER BY venta_neta DESC
            LIMIT 30
        """
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': rows})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD MATRIZ
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_any_perm('canales', 'supervisores', 'unidades-supervisores', 'informacion-rutas')
def dashboard_matriz_datos(request):
    """
    Datos granulares para la Tabla Matriz (pivot table interactiva).
    Grain: Regional Ã— Canal Ã— Supervisor Ã— Vendedor Ã— Ruta Ã— CategorÃ­a Ã— SKU Ã— PerÃ­odo
    Presupuesto asignado proporcionalmente al mix de ventas del vendedor.
    Cajas 9L = cantidad Ã— dp.u_L / 9000  (solo BEBIDAS ALC).
    Params: anho, mes, regional (role-based), canal (opt)
    """
    import calendar as _cal

    try:
        is_admin = _is_admin(request.user)
        profile  = _get_or_create_profile(request.user)
        cargo    = (profile.cargo or '').strip()

        anho = _safe_int(request.GET.get('anho'), datetime.now().year)
        mes  = _safe_int(request.GET.get('mes'),  datetime.now().month)

        if is_admin:
            regional_key = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
        elif cargo == 'Gerente Regional':
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = _safe_str(request.GET.get('canal', ''), 30)
        elif 'supervisor' in cargo.lower():
            regional_key = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'santa_cruz')
            canal        = (profile.canal or '').strip()
        else:
            regional_key = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
            canal        = _safe_str(request.GET.get('canal', ''), 30)

        if regional_key not in REGIONALES_VALID:
            regional_key = 'nacional'

        ciudad_cond = _regional_filter(regional_key)
        canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""

        # Supervisor lock for supervisor users
        sup_filter  = ''
        sup_cond    = ''
        if 'supervisor' in cargo.lower() and not is_admin:
            sup_filter = f"{profile.user.first_name} {profile.user.last_name}".strip()
            sup_cond   = "AND UPPER(dv.supervisor) = UPPER(%s)"

        dias_mes = _cal.monthrange(anho, mes)[1]

        scz  = _ciudad_case('dv.ciudad', 'santa_cruz')
        cbba = _ciudad_case('dv.ciudad', 'cochabamba')
        lpz  = _ciudad_case('dv.ciudad', 'la_paz')

        params = (
            [anho, mes, anho, mes]                            # ppto CTE
            + [anho, mes]                                     # ventas_sku fecha
            + ([canal]       if canal      else [])           # ventas_sku canal
            + ([sup_filter]  if sup_filter else [])           # ventas_sku supervisor
            + [anho, mes]                                     # dia_corte CTE
        )

        sql = f"""
        WITH
        ppto AS (
            SELECT fp.vendedor_sk,
                   COALESCE(SUM(fp.venta_neta_presupuestada), 0) AS presupuesto
            FROM dw.fact_presupuesto fp
            WHERE fp.anho = %s AND fp.mes = %s
              AND fp.version_sk = (
                  SELECT MAX(version_sk) FROM dw.dim_presupuesto_version
                  WHERE anho = %s AND mes = %s
              )
            GROUP BY fp.vendedor_sk
        ),
        ventas_sku AS (
            SELECT
                dv.vendedor_sk,
                COALESCE(dcd.ruta, '')                                             AS ruta,
                dp.producto_sk,
                dp.linea,
                COALESCE(SUM(fv.venta_neta), 0)                                    AS bs,
                COALESCE(SUM(fv.cantidad),   0)                                    AS uds,
                COALESCE(SUM(
                    CASE WHEN dp.linea = 'BEBIDAS ALC'
                    THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                    ELSE 0 END
                ), 0)                                                               AS cajas_9l
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha          df  ON df.fecha_sk      = fv.fecha_sk
            JOIN dw.dim_vendedor       dv  ON dv.vendedor_sk   = fv.vendedor_sk
            JOIN dw.dim_producto       dp  ON dp.producto_sk   = fv.producto_sk
            LEFT JOIN dw.dim_cliente     dck ON dck.cliente_sk  = fv.cliente_sk
                                           AND dck.es_cliente_actual = true
            LEFT JOIN dual.dim_cliente_dual dcd
                                        ON dcd.codigo_cliente  = dck.cliente_codigo_erp
                                       AND dcd.es_actual       = true
            WHERE df.anho = %s AND df.mes_numero = %s
              AND ({ciudad_cond})
              {canal_cond}
              {sup_cond}
            GROUP BY dv.vendedor_sk, COALESCE(dcd.ruta, ''), dp.producto_sk, dp.linea
        ),
        total_vend AS (
            SELECT vendedor_sk, SUM(bs) AS bs_total
            FROM ventas_sku
            GROUP BY vendedor_sk
        ),
        dia_corte AS (
            SELECT COALESCE(MAX(df2.dia_numero), 1) AS d
            FROM dw.dim_fecha df2
            WHERE df2.anho = %s AND df2.mes_numero = %s
              AND df2.fecha_completa <= CURRENT_DATE
        )
        SELECT
            CASE
                WHEN {scz}  THEN 'Santa Cruz'
                WHEN {cbba} THEN 'Cochabamba'
                WHEN {lpz}  THEN 'La Paz'
                ELSE             'Nacional'
            END                                                                  AS "Regional",
            dv.canal_rrhh                                                        AS "Canal",
            INITCAP(dv.supervisor)                                               AS "Supervisor",
            dv.vendedor_nombre                                                   AS "Vendedor",
            NULLIF(vs.ruta, '')                                                  AS "Ruta",
            CASE dp.linea
                WHEN 'ALIMENTOS'            THEN 'Alimentos'
                WHEN 'APEGO'                THEN 'Apego'
                WHEN 'BEBIDAS ALC'          THEN 'Licores'
                WHEN 'HOME Y PERSONAL CARE' THEN 'Home & Personal Care'
                ELSE                             'Sin Clasificar'
            END                                                                  AS "CategorÃ­a",
            dp.producto_nombre                                                   AS "SKU",
            TO_CHAR(MAKE_DATE({anho}, {mes}, 1), 'Mon YYYY')                     AS "PerÃ­odo",
            ROUND(vs.bs::NUMERIC,        0)                                      AS "Bs Vendidos",
            ROUND(vs.uds::NUMERIC,       0)                                      AS "Unidades",
            ROUND(vs.cajas_9l::NUMERIC,  2)                                      AS "Cajas 9L",
            ROUND(CASE WHEN tv.bs_total > 0
                  THEN COALESCE(p.presupuesto, 0) * vs.bs / tv.bs_total
                  ELSE 0 END::NUMERIC, 0)                                        AS "Presupuesto",
            ROUND(CASE WHEN dc.d > 0
                  THEN vs.bs / dc.d * {dias_mes}
                  ELSE 0 END::NUMERIC, 0)                                        AS "Proyectado Cierre"
        FROM ventas_sku vs
        JOIN  dw.dim_vendedor  dv ON dv.vendedor_sk  = vs.vendedor_sk
        JOIN  dw.dim_producto  dp ON dp.producto_sk  = vs.producto_sk
        LEFT JOIN ppto          p  ON p.vendedor_sk  = vs.vendedor_sk
        LEFT JOIN total_vend   tv  ON tv.vendedor_sk = vs.vendedor_sk
        CROSS JOIN dia_corte   dc
        ORDER BY "Regional", "Canal", "Supervisor", "Vendedor", "Ruta", "CategorÃ­a", "SKU"
        """

        _, rows = _run_dw_query(sql, params)

        # GAP y DesviaciÃ³n % calculados en Python
        result = []
        for r in rows:
            bs   = float(r.get('Bs Vendidos') or 0)
            ppto = float(r.get('Presupuesto') or 0)
            gap  = round(bs - ppto)
            dev  = round((gap / ppto * 100), 1) if ppto else None
            result.append({
                **{k: (float(v) if isinstance(v, (int, float)) else v) for k, v in r.items()},
                'GAP':          gap,
                'DesviaciÃ³n %': dev,
            })

        return JsonResponse({'success': True, 'data': result, 'total_filas': len(result)})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


def exportar_clientes_sin_compra(request):
    """
    Exporta a XLSX los clientes de las rutas que no compraron en el mes indicado.
    Auth: header Authorization: Token xxx
    Params: regional, canal, dia, supervisor, anho, mes
    """
    token_str = ""
    auth_header = request.META.get("HTTP_AUTHORIZATION", "")
    if auth_header.startswith("Token "):
        token_str = auth_header[6:]
    else:
        token_str = request.GET.get("_t", "")
    if not token_str:
        return JsonResponse({"error": "No autorizado"}, status=401)
    try:
        from rest_framework.authtoken.models import Token as DRFToken
        token_obj = DRFToken.objects.select_related("user").get(key=token_str)
        if not token_obj.user.is_active:
            return JsonResponse({"error": "Usuario inactivo"}, status=401)
    except Exception:
        return JsonResponse({"error": "Token invÃ¡lido"}, status=401)

    if not _has_dashboard_perm(token_obj.user, 'informacion-rutas'):
        return JsonResponse({"error": "Sin acceso a este dashboard"}, status=403)

    regional_raw = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
    regional     = regional_raw if regional_raw in REGIONALES_VALID else 'nacional'
    canal        = _safe_str(request.GET.get('canal',      'Todos'), 30)
    dia          = _safe_str(request.GET.get('dia',        'Todos'), 10)
    supervisor   = _safe_str(request.GET.get('supervisor', 'Todos'), 100)
    marca        = _safe_str(request.GET.get('marca',      ''),      80)
    anho         = _safe_int(request.GET.get('anho'), datetime.now().year)
    mes          = _safe_int(request.GET.get('mes'),  datetime.now().month)

    ciudad_cond    = _regional_filter(regional)
    regional_cond  = "" if regional == 'nacional' else f"AND ({ciudad_cond})"
    canal_cond     = "AND dv.canal_rrhh = %s" if canal      != 'Todos' else ""
    dia_cond       = "AND dp.dia = %s"        if dia        != 'Todos' else ""
    sup_cond       = "AND dv.supervisor = %s" if supervisor != 'Todos' else ""
    canal_cte_cond = "AND dv2.canal_rrhh = %s" if canal != 'Todos' else ""
    marca_cte_cond = "AND dprod.marca = %s"     if marca else ""

    params = [anho, mes]
    if canal != 'Todos': params.append(canal)   # CTE ventas_mes canal
    if marca:            params.append(marca)   # CTE ventas_mes marca
    if canal != 'Todos': params.append(canal)   # WHERE canal_cond
    if dia   != 'Todos': params.append(dia)
    if supervisor != 'Todos': params.append(supervisor)

    sql = f"""
        WITH ventas_mes AS (
            SELECT DISTINCT fv.cliente_sk
            FROM dw.fact_ventas    fv
            JOIN dw.dim_fecha      df    ON df.fecha_sk     = fv.fecha_sk
            JOIN dw.dim_vendedor   dv2   ON dv2.vendedor_sk = fv.vendedor_sk
            JOIN dw.dim_producto   dprod ON dprod.producto_sk = fv.producto_sk
            WHERE df.anho = %s AND df.mes_numero = %s
              {canal_cte_cond}
              {marca_cte_cond}
        ),
        ultima_compra AS (
            SELECT fv.cliente_sk, MAX(df.fecha_completa) AS ultima_fecha
            FROM dw.fact_ventas fv
            JOIN dw.dim_fecha   df ON df.fecha_sk = fv.fecha_sk
            GROUP BY fv.cliente_sk
        )
        SELECT
            dc.ruta,
            COALESCE(dp.vendedor,   'Sin asignar') AS vendedor,
            COALESCE(dv.supervisor, 'â€"')           AS supervisor,
            COALESCE(dp.dia,        'â€"')           AS dia,
            dc.codigo_cliente,
            dc.nombre_compania,
            COALESCE(dv.canal_rrhh, dc.canal)      AS canal,
            TO_CHAR(uc.ultima_fecha, 'DD/MM/YYYY') AS ultima_compra
        FROM dual.dim_cliente_dual dc
        LEFT JOIN dual.dim_planificacion dp
               ON dp.ruta = dc.ruta AND dp.es_actual = true
        LEFT JOIN dw.dim_vendedor dv
               ON dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
              AND dv.es_vendedor_actual = true
        LEFT JOIN dw.dim_cliente dck
               ON dck.cliente_codigo_erp = dc.codigo_cliente
              AND dck.es_cliente_actual = true
        LEFT JOIN ventas_mes  vm ON vm.cliente_sk  = dck.cliente_sk
        LEFT JOIN ultima_compra uc ON uc.cliente_sk = dck.cliente_sk
        WHERE dc.es_actual = true
          AND vm.cliente_sk IS NULL
          {regional_cond}
          {canal_cond}
          {dia_cond}
          {sup_cond}
        ORDER BY dc.ruta, dc.nombre_compania
    """

    headers = ["Ruta", "Vendedor", "Supervisor", "DÃ­a", "CÃ³d. Cliente", "Nombre", "Canal", "Ãšltima Compra"]

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Clientes Sin Compra")
    bold = Font(bold=True)
    header_row = []
    for h in headers:
        c = openpyxl.cell.WriteOnlyCell(ws, value=h)
        c.font = bold
        header_row.append(c)
    ws.append(header_row)

    try:
        with connections['dw'].cursor() as cursor:
            cursor.execute(sql, params)
            while True:
                batch = cursor.fetchmany(2000)
                if not batch: break
                for row in batch:
                    ws.append(list(row))
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({"success": False, "error": "Error interno del servidor"}, status=500)

    buffer = io.BytesIO()
    wb.save(buffer)
    file_size = buffer.tell()
    buffer.seek(0)

    def _chunks(buf, size=65536):
        chunk = buf.read(size)
        while chunk:
            yield chunk
            chunk = buf.read(size)

    mes_str   = str(mes).zfill(2)
    sup_slug  = supervisor.replace(' ', '_') if supervisor != 'Todos' else 'todos'
    marca_slug = marca.replace(' ', '_') if marca else 'todas'
    filename  = f"clientes_sin_compra_{sup_slug}_{marca_slug}_{anho}_{mes_str}.xlsx"
    response = StreamingHttpResponse(
        _chunks(buffer),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Encoding"]    = "identity"
    response["X-File-Size"]         = str(file_size)
    return response


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD TENDENCIA ESTACIONAL
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('tendencia-estacional')
def dashboard_tendencia_estacional(request):
    """
    ComparaciÃ³n estacional (mismo mes entre gestiones) o Ãºltimos 6 meses.
    Params:
      regional  : Nacional | Santa Cruz | Cochabamba | La Paz
      canal     : Todos | WHS | DTS | PROV | SPM
      anho      : int
      mes       : int  (1-12)
      modo      : estacional | ultimos6
      dia_corte : 0 = mes completo, N = primeros N dÃ­as del mes
    """
    is_admin = _is_admin(request.user)
    profile  = _get_or_create_profile(request.user)
    cargo    = (profile.cargo or '').strip()

    anho      = _safe_int(request.GET.get('anho'),  datetime.now().year)
    mes       = _safe_int(request.GET.get('mes'),   datetime.now().month)
    modo      = request.GET.get('modo', 'estacional')
    dia_corte = _safe_int(request.GET.get('dia_corte'), 0)

    if is_admin:
        regional_raw = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        regional     = regional_raw if regional_raw in REGIONALES_VALID else 'nacional'
        canal        = _safe_str(request.GET.get('canal',      'Todos'), 20)
        supervisor   = _safe_str(request.GET.get('supervisor', ''),      100)
    elif cargo == 'Gerente Regional':
        regional     = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'nacional')
        canal        = _safe_str(request.GET.get('canal', 'Todos'), 20)
        supervisor   = _safe_str(request.GET.get('supervisor', ''), 100)
    elif 'supervisor' in cargo.lower():
        regional     = _REGIONAL_NAME_TO_KEY.get(profile.regional, 'nacional')
        canal        = (profile.canal or '').strip()
        supervisor   = f"{profile.user.first_name} {profile.user.last_name}".strip()
    else:
        regional_raw = request.GET.get('regional', 'nacional').lower().replace(' ', '_')
        regional     = regional_raw if regional_raw in REGIONALES_VALID else 'nacional'
        canal        = _safe_str(request.GET.get('canal',      'Todos'), 20)
        supervisor   = _safe_str(request.GET.get('supervisor', ''),      100)

    joins = """
        FROM dw.fact_ventas fv
        JOIN dw.dim_fecha        df ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor     dv ON dv.vendedor_sk = fv.vendedor_sk
        LEFT JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
    """

    ciudad_cond = _regional_filter(regional)
    extra_conds  = [f"({ciudad_cond})"]
    extra_params = []
    if canal and canal != 'Todos':
        extra_conds.append("dv.canal_rrhh = %s")
        extra_params.append(canal)
    if supervisor:
        extra_conds.append("UPPER(dv.supervisor) = UPPER(%s)")
        extra_params.append(supervisor)
    extra_where = " AND " + " AND ".join(extra_conds)

    # Helper: ejecuta queries de desglose por categorÃ­a y canal para un WHERE+params dado
    def _desgloses(where_conds, params_base):
        # CategorÃ­as
        sql_cat = f"""
            SELECT {_CATEGORIA_CASE} AS categoria,
                   df.anho, df.mes_numero,
                   ROUND(COALESCE(SUM(fv.venta_neta), 0)::NUMERIC, 2) AS total,
                   COALESCE(SUM(fv.cantidad), 0)                       AS cantidad
            {joins}
            WHERE {where_conds}
            GROUP BY categoria, df.anho, df.mes_numero
            ORDER BY df.anho, df.mes_numero, categoria
        """
        _, cat_rows = _run_dw_query(sql_cat, params_base)

        # Canales (solo cuando no hay filtro de canal)
        canal_rows = []
        if canal == 'Todos':
            sql_canal = f"""
                SELECT dv.canal_rrhh AS canal,
                       df.anho, df.mes_numero,
                       ROUND(COALESCE(SUM(fv.venta_neta), 0)::NUMERIC, 2) AS total,
                       COALESCE(SUM(fv.cantidad), 0)                       AS cantidad
                {joins}
                WHERE {where_conds} AND dv.canal_rrhh IS NOT NULL
                GROUP BY dv.canal_rrhh, df.anho, df.mes_numero
                ORDER BY df.anho, df.mes_numero, dv.canal_rrhh
            """
            _, canal_rows = _run_dw_query(sql_canal, params_base)

        return cat_rows, canal_rows

    try:
        if modo == 'estacional':
            conds = ["df.mes_numero = %s", "df.anho BETWEEN %s AND %s"]
            params = [mes, anho - 2, anho]
            if dia_corte > 0:
                conds.append("df.dia_numero <= %s")
                params.append(dia_corte)
            params.extend(extra_params)
            where_str = " AND ".join(conds) + extra_where

            sql = f"""
                SELECT df.anho,
                       ROUND(COALESCE(SUM(fv.venta_neta), 0)::NUMERIC, 2) AS total,
                       COALESCE(SUM(fv.cantidad), 0)                       AS cantidad
                {joins}
                WHERE {where_str}
                GROUP BY df.anho
                ORDER BY df.anho
            """
            _, rows = _run_dw_query(sql, params)
            cat_rows, canal_rows = _desgloses(where_str, params)

            return JsonResponse({
                'success':        True,
                'modo':           'estacional',
                'mes_numero':     mes,
                'anho_ref':       anho,
                'dia_corte':      dia_corte,
                'data':           rows,
                'data_categoria': cat_rows,
                'data_canal':     canal_rows,
            })

        else:  # ultimos6
            start_m, start_y = mes - 5, anho
            while start_m <= 0:
                start_m += 12
                start_y -= 1

            u6_conds = ["df.anho * 100 + df.mes_numero BETWEEN %s AND %s"]
            u6_params = [start_y * 100 + start_m, anho * 100 + mes]
            if dia_corte > 0:
                u6_conds.append("df.dia_numero <= %s")
                u6_params.append(dia_corte)
            u6_params.extend(extra_params)
            where_str = " AND ".join(u6_conds) + extra_where

            sql = f"""
                SELECT df.anho, df.mes_numero,
                       ROUND(COALESCE(SUM(fv.venta_neta), 0)::NUMERIC, 2) AS total,
                       COALESCE(SUM(fv.cantidad), 0)                       AS cantidad
                {joins}
                WHERE {where_str}
                GROUP BY df.anho, df.mes_numero
                ORDER BY df.anho, df.mes_numero
            """
            _, rows = _run_dw_query(sql, u6_params)
            cat_rows, canal_rows = _desgloses(where_str, u6_params)

            return JsonResponse({
                'success':        True,
                'modo':           'ultimos6',
                'data':           rows,
                'data_categoria': cat_rows,
                'data_canal':     canal_rows,
            })

    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  DASHBOARD FICHA DE SKU
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

_TRIM_MESES = {
    '1': (1,  3),
    '2': (4,  6),
    '3': (7,  9),
    '4': (10, 12),
}

_CAT_LINEA_FICHA = {
    'Alimentos':            'ALIMENTOS',
    'Apego':                'APEGO',
    'Licores':              'BEBIDAS ALC',
    'Home & Personal Care': 'HOME Y PERSONAL CARE',
}


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard_almacenes_lista(request):
    """Lista de almacenes activos, opcionalmente filtrados por regional."""
    regional_key = _safe_str(request.GET.get('regional', ''), 50).lower()

    CIUDADES_ALMACEN = {
        'santa_cruz': ['SANTA CRUZ'],
        'cochabamba': ['COCHABAMBA'],
        'la_paz':     ['LA PAZ'],
    }

    params = []

    ciudad_cond = ''
    if regional_key and regional_key != 'nacional' and regional_key in CIUDADES_ALMACEN:
        ciudades = CIUDADES_ALMACEN[regional_key]
        if len(ciudades) == 1:
            ciudad_cond = 'AND da.ciudad = %s'
            params.append(ciudades[0])
        else:
            ph = ', '.join(['%s'] * len(ciudades))
            ciudad_cond = f'AND da.ciudad IN ({ph})'
            params.extend(ciudades)

    sql = f"""
        SELECT da.almacen_codigo_erp AS codigo,
               da.almacen_nombre     AS nombre,
               da.ciudad
        FROM dw.dim_almacen da
        WHERE 1=1
          {ciudad_cond}
        ORDER BY da.ciudad, da.almacen_nombre
    """
    try:
        _, rows = _run_dw_query(sql, params)
        return JsonResponse({'success': True, 'data': [
            {'codigo': r['codigo'], 'nombre': r['nombre'], 'ciudad': r['ciudad']}
            for r in rows
        ]})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('ficha-sku')
def dashboard_ficha_sku_marcas(request):
    """Marcas (proveedor) disponibles, opcionalmente filtradas por categoria."""
    categoria = _safe_str(request.GET.get('categoria', ''), 50)

    cache_key = f'ficha_sku_marcas:{categoria}'
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse({'success': True, 'data': cached})

    params = []
    cat_cond = ""
    if categoria in _CAT_LINEA_FICHA:
        cat_cond = "AND dp.linea = %s"
        params.append(_CAT_LINEA_FICHA[categoria])
    elif categoria == 'Sin Clasificar':
        cat_cond = "AND (dp.linea IS NULL OR dp.linea = 'SIN LINEA')"

    sql = f"""
        SELECT DISTINCT dp.proveedor AS marca
        FROM dw.dim_producto dp
        WHERE dp.es_producto_actual = true
          AND dp.proveedor IS NOT NULL
          AND dp.proveedor <> ''
          {cat_cond}
        ORDER BY marca
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [r['marca'] for r in rows]
        cache.set(cache_key, data, 1800)   # 30 min
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('ficha-sku')
def dashboard_ficha_sku_buscar(request):
    """
    BÃºsqueda de productos por texto (nombre o cÃ³digo).
    Params: q, categoria, marca
    """
    q         = _safe_str(request.GET.get('q', ''), 100).strip()
    categoria = _safe_str(request.GET.get('categoria', ''), 50)
    marca     = _safe_str(request.GET.get('marca', ''), 100)

    if len(q) < 2 and not categoria and not marca:
        return JsonResponse({'success': True, 'data': []})

    params = []
    search_cond = ""
    if len(q) >= 2:
        params = [f'%{q.upper()}%', f'%{q.upper()}%']
        search_cond = "AND (UPPER(dp.producto_nombre) LIKE %s OR UPPER(dp.producto_codigo_erp) LIKE %s)"

    cat_cond = ""
    if categoria in _CAT_LINEA_FICHA:
        cat_cond = "AND dp.linea = %s"
        params.append(_CAT_LINEA_FICHA[categoria])
    elif categoria == 'Sin Clasificar':
        cat_cond = "AND (dp.linea IS NULL OR dp.linea = 'SIN LINEA')"

    marca_cond = ""
    if marca:
        marca_cond = "AND dp.proveedor = %s"
        params.append(marca)

    limit = 50 if len(q) >= 2 else 250

    sql = f"""
        SELECT DISTINCT
            dp.producto_codigo_erp                      AS codigo,
            INITCAP(dp.producto_nombre)                 AS nombre,
            COALESCE(dp.linea, 'SIN LINEA')             AS linea,
            COALESCE(dp.proveedor, '')                  AS marca,
            COALESCE(dp.u_L, 0)                         AS ul
        FROM dw.dim_producto dp
        WHERE dp.es_producto_actual = true
          {search_cond} {cat_cond} {marca_cond}
        ORDER BY INITCAP(dp.producto_nombre)
        LIMIT {limit}
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'linea':  r['linea'],
                'marca':  r['marca'],
                'ul':     float(r['ul'] or 0),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('ficha-sku')
def dashboard_ficha_sku_ventas(request):
    """
    Ventas diarias de un SKU en un trimestre.
    Params: codigo, anho, trimestre (1-4), regional, canal
    """
    codigo       = _safe_str(request.GET.get('codigo', ''), 50)
    anho         = _safe_int(request.GET.get('anho'), datetime.now().year)
    trimestre    = request.GET.get('trimestre', '').strip()
    mes          = _safe_int(request.GET.get('mes'), 0)
    regional_key = request.GET.get('regional', 'santa_cruz').lower().replace(' ', '_')
    canal        = _safe_str(request.GET.get('canal', ''), 30)

    if not codigo:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro codigo requerido'}, status=400)
    if regional_key not in REGIONALES_VALID:
        return JsonResponse({'success': False, 'error': 'Regional invÃ¡lida'}, status=400)

    if trimestre in _TRIM_MESES:
        mes_desde, mes_hasta = _TRIM_MESES[trimestre]
    elif 1 <= mes <= 12:
        mes_desde = mes_hasta = mes
    else:
        mes_desde = mes_hasta = datetime.now().month
    ciudad_cond = _regional_filter(regional_key, campo='dv.ciudad')
    canal_cond  = "AND dv.canal_rrhh = %s" if canal else ""
    params      = [anho, mes_desde, mes_hasta, codigo] + ([canal] if canal else [])

    sql = f"""
        SELECT
            df.fecha_completa::TEXT                                          AS fecha,
            COALESCE(SUM(fv.cantidad), 0)                                    AS unidades,
            COALESCE(SUM(fv.venta_neta), 0)                                  AS bs,
            COALESCE(SUM(
                CASE WHEN dp.linea = 'BEBIDAS ALC'
                THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                ELSE fv.cantidad::NUMERIC END
            ), 0)                                                             AS vol,
            MAX(CASE WHEN dp.linea = 'BEBIDAS ALC' THEN 1 ELSE 0 END)       AS es_licor
        FROM dw.fact_ventas       fv
        JOIN dw.dim_fecha         df  ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor      dv  ON dv.vendedor_sk = fv.vendedor_sk
        JOIN dw.dim_producto      dp  ON dp.producto_sk = fv.producto_sk
        WHERE df.anho = %s
          AND df.mes_numero BETWEEN %s AND %s
          AND dp.producto_codigo_erp = %s
          AND ({ciudad_cond}) {canal_cond}
        GROUP BY df.fecha_completa
        ORDER BY df.fecha_completa
    """
    try:
        _, rows = _run_dw_query(sql, params)
        es_licor = any(r.get('es_licor') for r in rows)
        data = [
            {
                'fecha':    r['fecha'],
                'unidades': float(r['unidades']),
                'bs':       float(r['bs']),
                'vol':      float(r['vol']),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data, 'es_licor': es_licor})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('ficha-sku')
def dashboard_ficha_sku_precios(request):
    """
    Historial de precios de un SKU desde fact_precio_producto.
    Params: codigo
    """
    codigo = _safe_str(request.GET.get('codigo', ''), 50)
    if not codigo:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro codigo requerido'}, status=400)

    sql = """
        SELECT
            lp.lista_nombre                         AS lista,
            fp.precio_venta                         AS precio,
            fp.precio_con_ice                       AS precio_ice,
            fp.fecha_desde_precio::TEXT             AS fecha_desde,
            fp.fecha_hasta_precio::TEXT             AS fecha_hasta,
            fp.es_precio_actual                     AS es_actual
        FROM dw.fact_precio_producto fp
        JOIN dw.dim_lista_precios    lp  ON lp.lista_precios_sk = fp.lista_precios_sk
        JOIN dw.dim_producto         dp  ON dp.producto_sk      = fp.producto_sk
        WHERE dp.producto_codigo_erp = %s
        ORDER BY fp.fecha_desde_precio, lp.lista_nombre
    """
    try:
        _, rows = _run_dw_query(sql, [codigo])
        data = [
            {
                'lista':      r['lista'],
                'precio':     float(r['precio']),
                'precio_ice': float(r['precio_ice']) if r.get('precio_ice') is not None else None,
                'fecha_desde':r['fecha_desde'],
                'fecha_hasta':r['fecha_hasta'],
                'es_actual':  r['es_actual'],
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('ficha-sku')
def dashboard_ficha_sku_inventario(request):
    """
    Stock real desde fact_inventario para un SKU.
    Params: codigo, anho, mes (o trimestre), almacen (codigo_erp opcional)
    Devuelve: stock_actual (Ãºltimo registro), fecha_stock, y snapshots del perÃ­odo.
    """
    codigo    = _safe_str(request.GET.get('codigo', ''), 50)
    anho      = _safe_int(request.GET.get('anho'), datetime.now().year)
    trimestre = request.GET.get('trimestre', '').strip()
    mes       = _safe_int(request.GET.get('mes'), 0)
    almacen   = _safe_str(request.GET.get('almacen', ''), 50)

    if not codigo:
        return JsonResponse({'success': False, 'error': 'cÃ³digo requerido'}, status=400)

    if trimestre in _TRIM_MESES:
        mes_desde, mes_hasta = _TRIM_MESES[trimestre]
    elif 1 <= mes <= 12:
        mes_desde = mes_hasta = mes
    else:
        mes_desde = mes_hasta = datetime.now().month

    last_day    = _cal_mod.monthrange(anho, mes_hasta)[1]
    fecha_desde = f"{anho}-{mes_desde:02d}-01"
    fecha_hasta = f"{anho}-{mes_hasta:02d}-{last_day:02d}"

    almacen_join = "JOIN dw.dim_almacen da ON da.almacen_sk = fi.almacen_sk" if almacen else ""
    almacen_cond = "AND da.almacen_codigo_erp = %s"                          if almacen else ""

    # Snapshots dentro del perÃ­odo
    sql_period = f"""
        SELECT
            fi.fecha_inventario::TEXT               AS fecha,
            COALESCE(SUM(fi.stock_buenos), 0)        AS stock_buenos,
            COALESCE(SUM(fi.stock_total),  0)        AS stock_total
        FROM dw.fact_inventario fi
        JOIN dw.dim_producto dp ON dp.producto_sk = fi.producto_sk
        {almacen_join}
        WHERE dp.producto_codigo_erp = %s
          AND fi.fecha_inventario BETWEEN %s AND %s
          {almacen_cond}
        GROUP BY fi.fecha_inventario
        ORDER BY fi.fecha_inventario
    """

    # Stock mÃ¡s reciente (puede ser fuera del perÃ­odo)
    sql_latest = f"""
        SELECT
            fi.fecha_inventario::TEXT               AS fecha,
            COALESCE(SUM(fi.stock_buenos), 0)        AS stock_buenos,
            COALESCE(SUM(fi.stock_total),  0)        AS stock_total
        FROM dw.fact_inventario fi
        JOIN dw.dim_producto dp ON dp.producto_sk = fi.producto_sk
        {almacen_join}
        WHERE dp.producto_codigo_erp = %s
          {almacen_cond}
        GROUP BY fi.fecha_inventario
        ORDER BY fi.fecha_inventario DESC
        LIMIT 1
    """

    period_params = [codigo, fecha_desde, fecha_hasta] + ([almacen] if almacen else [])
    latest_params = [codigo] + ([almacen] if almacen else [])

    try:
        _, period_rows = _run_dw_query(sql_period, period_params)
        _, latest_rows = _run_dw_query(sql_latest, latest_params)

        data = [
            {
                'fecha':        r['fecha'],
                'stock_buenos': float(r['stock_buenos'] or 0),
                'stock_total':  float(r['stock_total']  or 0),
            }
            for r in period_rows
        ]

        latest = latest_rows[0] if latest_rows else {}
        return JsonResponse({
            'success':      True,
            'stock_actual': float(latest.get('stock_buenos') or 0),
            'fecha_stock':  latest.get('fecha'),
            'data':         data,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# â"€â"€â"€ DistribuciÃ³n de Rutas â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('inventario-almacen')
def dashboard_inventario_almacen(request):
    """
    Stock por producto y almacÃ©n para una fecha dada.
    Params: fecha (YYYY-MM-DD), regional, almacen (codigo_erp)
    """
    fecha        = _safe_str(request.GET.get('fecha', ''), 10)
    regional_key = _safe_str(request.GET.get('regional', ''), 20).lower().replace(' ', '_')
    almacen      = _safe_str(request.GET.get('almacen', ''), 50)

    if not fecha:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro fecha requerido'}, status=400)

    CIUDADES_ALMACEN = {
        'santa_cruz': ['SANTA CRUZ'],
        'cochabamba': ['COCHABAMBA'],
        'la_paz':     ['LA PAZ'],
    }

    params = [fecha]
    almacen_cond  = ''
    regional_cond = ''

    if almacen:
        almacen_cond = 'AND da.almacen_codigo_erp = %s'
        params.append(almacen)
    elif regional_key and regional_key != 'nacional' and regional_key in CIUDADES_ALMACEN:
        ciudades = CIUDADES_ALMACEN[regional_key]
        ph = ', '.join(['%s'] * len(ciudades))
        regional_cond = f'AND da.ciudad IN ({ph})'
        params.extend(ciudades)

    sql = f"""
        SELECT
            da.almacen_nombre                           AS almacen,
            dp.producto_codigo_erp                      AS cod_interno,
            INITCAP(dp.producto_nombre)                 AS producto,
            fi.u_medida                                 AS u_medida,
            COALESCE(SUM(fi.stock_buenos),   0)         AS stock_buenos,
            COALESCE(SUM(fi.stock_danhados), 0)         AS stock_danhados,
            COALESCE(SUM(fi.stock_vencidos), 0)         AS stock_vencidos,
            COALESCE(SUM(fi.stock_total),    0)         AS stock_total
        FROM dw.fact_inventario fi
        JOIN dw.dim_producto dp ON dp.producto_sk = fi.producto_sk
        JOIN dw.dim_almacen  da ON da.almacen_sk  = fi.almacen_sk
        WHERE fi.fecha_inventario = %s
          {almacen_cond}
          {regional_cond}
        GROUP BY da.almacen_nombre, dp.producto_codigo_erp, dp.producto_nombre, fi.u_medida
        ORDER BY da.almacen_nombre, dp.producto_nombre
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'almacen':        r['almacen'],
                'cod_interno':    r['cod_interno'],
                'producto':       r['producto'],
                'u_medida':       r['u_medida'] or '',
                'stock_buenos':   float(r['stock_buenos']   or 0),
                'stock_danhados': float(r['stock_danhados'] or 0),
                'stock_vencidos': float(r['stock_vencidos'] or 0),
                'stock_total':    float(r['stock_total']    or 0),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('distribucion-rutas')
def dashboard_rutas_opciones(request):
    """Supervisores y dÃ­as disponibles, filtrados por canal y/o regional."""
    canal        = _safe_str(request.GET.get('canal', ''), 30).strip()
    regional_key = _safe_str(request.GET.get('regional', ''), 20).lower().replace(' ', '_')
    conds  = ["dv.es_vendedor_actual = true", "dv.supervisor IS NOT NULL", "TRIM(dv.supervisor) <> ''"]
    params = []
    if canal:
        conds.append('dv.canal_rrhh = %s')
        params.append(canal)
    if regional_key and regional_key in REGIONALES_VALID and regional_key != 'nacional':
        ciudad_cond = _regional_filter(regional_key, campo='dv.ciudad')
        conds.append(f'({ciudad_cond})')
    where = 'WHERE ' + ' AND '.join(conds)
    sql_sups = f"""
        SELECT DISTINCT INITCAP(dv.supervisor) AS supervisor
        FROM dw.dim_vendedor dv
        {where}
        ORDER BY supervisor
    """
    sql_dias = """
        SELECT DISTINCT dp.dia
        FROM dual.dim_planificacion dp
        WHERE dp.es_actual = true
          AND dp.dia IS NOT NULL AND TRIM(dp.dia) <> ''
        ORDER BY dp.dia
    """
    try:
        _, sup_rows  = _run_dw_query(sql_sups, params)
        _, dia_rows  = _run_dw_query(sql_dias, [])
        return JsonResponse({
            'success':     True,
            'supervisores': [r['supervisor'] for r in sup_rows],
            'dias':         [r['dia'] for r in dia_rows],
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('distribucion-rutas')
def dashboard_rutas_buscar(request):
    """
    BÃºsqueda dinÃ¡mica de rutas filtradas por supervisor, canal, dÃ­a y texto.
    Params: q, canal, supervisor, dia
    """
    q            = _safe_str(request.GET.get('q', ''),            100).strip()
    regional_key = _safe_str(request.GET.get('regional', ''),     20).lower().replace(' ', '_')
    canal        = _safe_str(request.GET.get('canal', ''),         30).strip()
    supervisor   = _safe_str(request.GET.get('supervisor', ''),   100).strip()
    dia          = _safe_str(request.GET.get('dia', ''),           20).strip()
    vendedor     = _safe_str(request.GET.get('vendedor', ''),     100).strip()

    conds  = ['dp.es_actual = true']
    params = []

    if regional_key and regional_key in REGIONALES_VALID and regional_key != 'nacional':
        ciudad_cond = _regional_filter(regional_key, campo='dv.ciudad')
        conds.append(f'({ciudad_cond})')
    if canal:
        conds.append('dv.canal_rrhh = %s')
        params.append(canal)
    if supervisor:
        conds.append('dv.supervisor ILIKE %s')
        params.append(supervisor)
    if dia:
        conds.append('dp.dia = %s')
        params.append(dia)
    if vendedor:
        conds.append('dp.vendedor ILIKE %s')
        params.append(f'%{vendedor}%')
    if q:
        conds.append('UPPER(dp.ruta) LIKE %s')
        params.append(f'%{q.upper()}%')

    where = 'WHERE ' + ' AND '.join(conds)

    sql = f"""
        SELECT DISTINCT ON (dp.ruta)
            dp.ruta,
            INITCAP(dp.vendedor)     AS vendedor,
            dv.canal_rrhh            AS canal,
            INITCAP(dv.supervisor)   AS supervisor
        FROM dual.dim_planificacion dp
        JOIN dw.dim_vendedor dv
            ON  dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
            AND dv.es_vendedor_actual  = true
        {where}
        ORDER BY dp.ruta
        LIMIT 60
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'ruta':       r['ruta'],
                'vendedor':   r['vendedor']   or '',
                'canal':      r['canal']      or '',
                'supervisor': r['supervisor'] or '',
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('distribucion-rutas')
def dashboard_rutas_info(request):
    """
    PolÃ­gono (coordenadas en secuencia) + estadÃ­sticas de una ruta.
    Params: ruta
    """
    ruta = _safe_str(request.GET.get('ruta', ''), 100).strip()
    if not ruta:
        return JsonResponse({'success': False, 'error': 'ParÃ¡metro ruta requerido'}, status=400)

    try:
        # Tomar solo la versiÃ³n mÃ¡s reciente; id_zona se repite por sucursal
        sql_poly = """
            SELECT dzp.latitud, dzp.longitud
            FROM dual.dim_zona_posicion dzp
            JOIN (
                SELECT id_ruta::integer AS id_zona_int, sucursal_origen
                FROM dual.dim_ruta
                WHERE nombre = %s
                  AND es_ruta_actual = true
                ORDER BY version_ruta DESC NULLS LAST
                LIMIT 1
            ) dr ON dr.id_zona_int = dzp.id_zona
                AND dr.sucursal_origen = dzp.sucursal_origen
            ORDER BY dzp.secuencia
        """
        _, poly_rows = _run_dw_query(sql_poly, [ruta])
        pts = [{'lat': float(r['latitud']), 'lng': float(r['longitud'])} for r in poly_rows]
        polygons = [pts] if len(pts) >= 3 else []

        # Clientes activos en la ruta
        sql_cli = """
            SELECT COUNT(*) AS total
            FROM dual.dim_cliente_dual
            WHERE ruta = %s AND es_actual = true
        """
        _, cli_rows = _run_dw_query(sql_cli, [ruta])
        clientes = int(cli_rows[0]['total']) if cli_rows else 0

        # Vendedor y datos de la ruta (desde dim_planificacion)
        sql_vend = """
            SELECT
                INITCAP(dp.vendedor)   AS vendedor,
                dp.dia,
                dv.canal_rrhh          AS canal,
                INITCAP(dv.supervisor) AS supervisor
            FROM dual.dim_planificacion dp
            JOIN dw.dim_vendedor dv
                ON  dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
                AND dv.es_vendedor_actual  = true
            WHERE dp.ruta = %s AND dp.es_actual = true
            LIMIT 1
        """
        _, vend_rows = _run_dw_query(sql_vend, [ruta])
        vi = vend_rows[0] if vend_rows else {}

        vendedor_full = vi.get('vendedor', '') or ''
        parts = vendedor_full.strip().split()
        if len(parts) >= 4:
            vendedor_corto = f"{parts[0]} {parts[2]}"
        elif len(parts) >= 2:
            vendedor_corto = f"{parts[0]} {parts[-1]}"
        else:
            vendedor_corto = vendedor_full

        # Detectar columnas de coordenadas y clasificaciÃ³n
        sql_cols = """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'dual' AND table_name = 'dim_cliente_dual'
            ORDER BY ordinal_position
        """
        _, col_rows    = _run_dw_query(sql_cols, [])
        cols_disp      = [r['column_name'] for r in col_rows]
        lat_col        = next((c for c in cols_disp if 'lat'   in c.lower()), None)
        lng_col        = next((c for c in cols_disp if 'lon'   in c.lower() or 'lng' in c.lower()), None)
        cls_col        = next((c for c in cols_disp if any(x in c.lower() for x in ['clasif','tipo_cli','segmento','actividad'])), 'canal')

        # Ubicaciones + datos de clientes
        clientes_geo = []
        if lat_col and lng_col:
            try:
                sql_geo = f"""
                    SELECT {lat_col}           AS lat,
                           {lng_col}           AS lng,
                           COALESCE(nombre_compania, '')    AS nombre,
                           COALESCE(codigo_cliente::text, '') AS codigo,
                           COALESCE({cls_col}::text, '')    AS clasificacion
                    FROM dual.dim_cliente_dual
                    WHERE ruta = %s AND es_actual = true
                      AND {lat_col} IS NOT NULL AND {lng_col} IS NOT NULL
                """
                _, geo_rows  = _run_dw_query(sql_geo, [ruta])
                clientes_geo = [
                    {'lat': float(r['lat']), 'lng': float(r['lng']),
                     'nombre': r['nombre'] or '', 'codigo': r['codigo'] or '',
                     'clasificacion': r['clasificacion'] or ''}
                    for r in geo_rows
                ]
            except Exception:
                clientes_geo = []

        return JsonResponse({
            'success':        True,
            'polygons':       polygons,
            'clientes':       clientes,
            'vendedor':       vendedor_full,
            'vendedorCorto':  vendedor_corto,
            'dia':            vi.get('dia',         '') or '',
            'canal':          vi.get('canal',       '') or '',
            'supervisor':     vi.get('supervisor',  '') or '',
            'clientesGeo':    clientes_geo,
        })
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('distribucion-rutas')
def dashboard_rutas_todos_poligonos(request):
    """PolÃ­gonos de todas las rutas que coinciden con los filtros."""
    regional_key = _safe_str(request.GET.get('regional', ''), 20).lower().replace(' ', '_')
    canal        = _safe_str(request.GET.get('canal', ''), 30).strip()
    supervisor   = _safe_str(request.GET.get('supervisor', ''), 100).strip()
    dia          = _safe_str(request.GET.get('dia', ''), 20).strip()
    vendedor     = _safe_str(request.GET.get('vendedor', ''), 100).strip()

    if not (canal or supervisor or vendedor or (regional_key and regional_key != 'nacional')):
        return JsonResponse({'success': False, 'error': 'Se requiere canal, supervisor, vendedor o regional'})

    conds  = ['dp.es_actual = true']
    params = []
    if regional_key and regional_key in REGIONALES_VALID and regional_key != 'nacional':
        ciudad_cond = _regional_filter(regional_key, campo='dv.ciudad')
        conds.append(f'({ciudad_cond})')
    if canal:
        conds.append('dv.canal_rrhh = %s')
        params.append(canal)
    if supervisor:
        conds.append('dv.supervisor ILIKE %s')
        params.append(supervisor)
    if dia:
        conds.append('dp.dia = %s')
        params.append(dia)
    if vendedor:
        conds.append('dp.vendedor ILIKE %s')
        params.append(vendedor)

    where = 'WHERE ' + ' AND '.join(conds)

    sql = f"""
        WITH rutas_filtradas AS (
            SELECT DISTINCT ON (dp.ruta)
                dp.ruta,
                INITCAP(dp.vendedor) AS vendedor_full
            FROM dual.dim_planificacion dp
            JOIN dw.dim_vendedor dv
                ON  dv.vendedor_codigo_erp = SPLIT_PART(dp.codigo_erp, '.', 1)
                AND dv.es_vendedor_actual  = true
            {where}
            ORDER BY dp.ruta
            LIMIT 150
        ),
        ruta_zona AS (
            SELECT DISTINCT ON (dr.nombre)
                dr.nombre,
                dr.id_ruta::integer AS zona_id,
                dr.sucursal_origen
            FROM dual.dim_ruta dr
            WHERE dr.es_ruta_actual = true
              AND dr.nombre IN (SELECT ruta FROM rutas_filtradas)
            ORDER BY dr.nombre, dr.version_ruta DESC NULLS LAST
        )
        SELECT rz.nombre   AS ruta,
               rf.vendedor_full,
               dzp.latitud,
               dzp.longitud
        FROM ruta_zona rz
        JOIN rutas_filtradas rf ON rf.ruta = rz.nombre
        JOIN dual.dim_zona_posicion dzp
            ON  dzp.id_zona         = rz.zona_id
            AND dzp.sucursal_origen = rz.sucursal_origen
        ORDER BY rz.nombre, dzp.secuencia
    """
    try:
        _, rows = _run_dw_query(sql, params)
        rutas: dict = {}
        for r in rows:
            nm = r['ruta']
            if nm not in rutas:
                parts = (r['vendedor_full'] or '').strip().split()
                if len(parts) >= 4:
                    short = f"{parts[0]} {parts[2]}"
                elif len(parts) >= 2:
                    short = f"{parts[0]} {parts[-1]}"
                else:
                    short = r['vendedor_full'] or ''
                rutas[nm] = {'ruta': nm, 'vendedor': short, 'polygon': []}
            rutas[nm]['polygon'].append({'lat': float(r['latitud']), 'lng': float(r['longitud'])})

        # Ubicaciones + datos de clientes para todas las rutas
        clientes_geo = []
        ruta_names   = list(rutas.keys())
        if ruta_names:
            try:
                sql_cols2 = """
                    SELECT column_name FROM information_schema.columns
                    WHERE table_schema = 'dual' AND table_name = 'dim_cliente_dual'
                    ORDER BY ordinal_position
                """
                _, col_rows2  = _run_dw_query(sql_cols2, [])
                cols2         = [r['column_name'] for r in col_rows2]
                lat2          = next((c for c in cols2 if 'lat' in c.lower()), None)
                lng2          = next((c for c in cols2 if 'lon' in c.lower() or 'lng' in c.lower()), None)
                cls2          = next((c for c in cols2 if any(x in c.lower() for x in ['clasif','tipo_cli','segmento','actividad'])), 'canal')
                if lat2 and lng2:
                    placeholders = ','.join(['%s'] * len(ruta_names))
                    sql_cli = f"""
                        SELECT {lat2}                               AS lat,
                               {lng2}                               AS lng,
                               COALESCE(nombre_compania, '')        AS nombre,
                               COALESCE(codigo_cliente::text, '')   AS codigo,
                               COALESCE({cls2}::text, '')           AS clasificacion
                        FROM dual.dim_cliente_dual
                        WHERE ruta IN ({placeholders})
                          AND es_actual = true
                          AND {lat2} IS NOT NULL AND {lng2} IS NOT NULL
                        LIMIT 3000
                    """
                    _, cli_rows  = _run_dw_query(sql_cli, ruta_names)
                    clientes_geo = [
                        {'lat': float(r['lat']), 'lng': float(r['lng']),
                         'nombre': r['nombre'] or '', 'codigo': r['codigo'] or '',
                         'clasificacion': r['clasificacion'] or ''}
                        for r in cli_rows
                    ]
            except Exception:
                clientes_geo = []

        return JsonResponse({'success': True, 'rutas': list(rutas.values()), 'clientesGeo': clientes_geo})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


# ---------------------------------------------------------------------------
# Dashboard Comportamiento Productos
# ---------------------------------------------------------------------------

@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('comportamiento-productos')
def dashboard_comportamiento_opciones(request):
    regional_key = _safe_str(request.GET.get('regional', ''), 50).lower()
    canal        = _safe_str(request.GET.get('canal', ''), 100)
    proveedor    = _safe_str(request.GET.get('proveedor', ''), 150)

    params_v = []
    conds_v  = ['dv.es_vendedor_actual = true']

    regional_cond = _regional_filter(regional_key, 'dv.ciudad')
    if regional_cond:
        conds_v.append(regional_cond)

    if canal:
        conds_v.append('dv.canal_rrhh = %s')
        params_v.append(canal)

    where_v = ' AND '.join(conds_v)
    sql_v = f"""
        SELECT DISTINCT INITCAP(dv.vendedor_nombre) AS vendedor
        FROM dw.dim_vendedor dv
        WHERE {where_v}
        ORDER BY vendedor
        LIMIT 200
    """
    try:
        _, rows_v = _run_dw_query(sql_v, params_v)
        vendedores = [r['vendedor'] for r in rows_v if r['vendedor']]
    except Exception:
        logger.exception("Error interno - vendedores opciones")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)

    sql_p = """
        SELECT DISTINCT dp.proveedor
        FROM dw.dim_producto dp
        WHERE dp.es_producto_actual = true
          AND dp.proveedor IS NOT NULL
          AND TRIM(dp.proveedor) <> ''
        ORDER BY dp.proveedor
    """
    try:
        _, rows_p = _run_dw_query(sql_p, [])
        proveedores = [r['proveedor'] for r in rows_p if r['proveedor']]
    except Exception:
        logger.exception("Error interno - proveedores opciones")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)

    params_m = []
    conds_m  = [
        'dp.es_producto_actual = true',
        'dp.marca IS NOT NULL',
        "TRIM(dp.marca) <> ''",
    ]
    if proveedor:
        conds_m.append('dp.proveedor = %s')
        params_m.append(proveedor)
    where_m = ' AND '.join(conds_m)
    sql_m = f"""
        SELECT DISTINCT dp.marca
        FROM dw.dim_producto dp
        WHERE {where_m}
        ORDER BY dp.marca
    """
    try:
        _, rows_m = _run_dw_query(sql_m, params_m)
        marcas = [r['marca'] for r in rows_m if r['marca']]
    except Exception:
        logger.exception("Error interno - marcas opciones")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)

    return JsonResponse({'success': True, 'vendedores': vendedores, 'proveedores': proveedores, 'marcas': marcas})


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('comportamiento-productos')
def dashboard_comportamiento_productos_buscar(request):
    proveedor = _safe_str(request.GET.get('proveedor', ''), 150)
    marca     = _safe_str(request.GET.get('marca', ''), 150)
    q         = _safe_str(request.GET.get('q', ''), 150)

    params = []
    conds  = ['dp.es_producto_actual = true']

    if proveedor:
        conds.append('dp.proveedor = %s')
        params.append(proveedor)

    if marca:
        conds.append('dp.marca = %s')
        params.append(marca)

    if q:
        q_upper = q.upper()
        conds.append("(UPPER(dp.producto_nombre) LIKE %s OR UPPER(dp.producto_codigo_erp) LIKE %s)")
        params.append(f'%{q_upper}%')
        params.append(f'%{q_upper}%')

    where = ' AND '.join(conds)
    sql = f"""
        SELECT DISTINCT
            dp.producto_codigo_erp                                          AS codigo,
            INITCAP(dp.producto_nombre)                                     AS nombre,
            COALESCE(dp.marca, '')                                          AS marca,
            COALESCE(dp.linea, '')                                          AS linea,
            CASE WHEN dp.linea = 'BEBIDAS ALC' THEN true ELSE false END     AS es_licor
        FROM dw.dim_producto dp
        WHERE {where}
        ORDER BY nombre
        LIMIT 50
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'codigo':   r['codigo'],
                'nombre':   r['nombre'],
                'marca':    r['marca'],
                'linea':    r['linea'],
                'es_licor': r['es_licor'],
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('comportamiento-productos')
def dashboard_comportamiento_grafico1(request):
    regional_key = _safe_str(request.GET.get('regional', ''), 50).lower()
    canal        = _safe_str(request.GET.get('canal', ''), 100)
    vendedor     = _safe_str(request.GET.get('vendedor', ''), 150)
    anho         = _safe_int(request.GET.get('anho', ''), 0)
    mes          = _safe_int(request.GET.get('mes', ''), 0)
    dia_corte    = _safe_int(request.GET.get('dia_corte', ''), 0)
    codigos_raw  = request.GET.get('codigos', '') or ''
    proveedor    = _safe_str(request.GET.get('proveedor', ''), 150)
    marca        = _safe_str(request.GET.get('marca', ''), 150)

    if not anho or not mes:
        return JsonResponse({'success': False, 'error': 'Parametros anho y mes son requeridos'})

    if canal:
        dim_field = 'INITCAP(dv.vendedor_nombre)'
        modo      = 'vendedor'
    else:
        dim_field = 'dv.canal_rrhh'
        modo      = 'canal'

    params = [anho, mes]
    conds  = [
        'df.anho = %s',
        'df.mes_numero = %s',
        'dv.es_vendedor_actual = true',
        'dp.es_producto_actual = true',
    ]

    if dia_corte:
        conds.append('df.dia_numero <= %s')
        params.append(dia_corte)

    regional_cond = _regional_filter(regional_key, 'dv.ciudad')
    if regional_cond:
        conds.append(regional_cond)

    if canal:
        conds.append('dv.canal_rrhh = %s')
        params.append(canal)

    if vendedor:
        conds.append('INITCAP(dv.vendedor_nombre) = %s')
        params.append(vendedor)

    codigos = [_safe_str(c.strip(), 100) for c in codigos_raw.split(',') if c.strip()]
    if codigos:
        placeholders = ', '.join(['%s'] * len(codigos))
        conds.append(f'dp.producto_codigo_erp IN ({placeholders})')
        params.extend(codigos)
    else:
        if marca:
            conds.append('dp.marca = %s')
            params.append(marca)
        if proveedor:
            conds.append('dp.proveedor = %s')
            params.append(proveedor)

    where = ' AND '.join(conds)
    sql = f"""
        SELECT {dim_field} AS dimension,
               COALESCE(SUM(fv.venta_neta), 0)   AS bs,
               COALESCE(SUM(fv.cantidad), 0)      AS uds,
               FLOOR(COALESCE(SUM(
                   CASE WHEN dp.linea = 'BEBIDAS ALC'
                        THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                        ELSE 0 END
               ), 0))                              AS cajas9l
        FROM dw.fact_ventas fv
        JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
        JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
        WHERE {where}
        GROUP BY {dim_field}
        ORDER BY bs DESC
        LIMIT 30
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'dimension': r['dimension'] or '',
                'bs':        float(r['bs']),
                'uds':       int(r['uds']),
                'cajas9l':   int(r['cajas9l']),
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'modo': modo, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('comportamiento-productos')
def dashboard_comportamiento_grafico2(request):
    regional_key = _safe_str(request.GET.get('regional', ''), 50).lower()
    canal        = _safe_str(request.GET.get('canal', ''), 100)
    vendedor     = _safe_str(request.GET.get('vendedor', ''), 150)
    anho         = _safe_int(request.GET.get('anho', ''), 0)
    mes          = _safe_int(request.GET.get('mes', ''), 0)
    dia_corte    = _safe_int(request.GET.get('dia_corte', ''), 0)
    codigos_raw  = request.GET.get('codigos', '') or ''
    proveedor    = _safe_str(request.GET.get('proveedor', ''), 150)
    marca        = _safe_str(request.GET.get('marca', ''), 150)

    if not anho or not mes:
        return JsonResponse({'success': False, 'error': 'Parametros anho y mes son requeridos'})

    codigos = [_safe_str(c.strip(), 100) for c in codigos_raw.split(',') if c.strip()]
    if not codigos and not marca and not proveedor:
        return JsonResponse({'success': False, 'error': 'Se requiere proveedor, marca o productos'})

    params = [anho, mes]
    conds  = [
        'df.anho = %s',
        'df.mes_numero = %s',
        'dv.es_vendedor_actual = true',
        'dp.es_producto_actual = true',
    ]

    if dia_corte:
        conds.append('df.dia_numero <= %s')
        params.append(dia_corte)

    regional_cond = _regional_filter(regional_key, 'dv.ciudad')
    if regional_cond:
        conds.append(regional_cond)

    if canal:
        conds.append('dv.canal_rrhh = %s')
        params.append(canal)

    if vendedor:
        conds.append('INITCAP(dv.vendedor_nombre) = %s')
        params.append(vendedor)

    if codigos:
        placeholders = ', '.join(['%s'] * len(codigos))
        conds.append(f'dp.producto_codigo_erp IN ({placeholders})')
        params.extend(codigos)
    else:
        if marca:
            conds.append('dp.marca = %s')
            params.append(marca)
        if proveedor:
            conds.append('dp.proveedor = %s')
            params.append(proveedor)

    where = ' AND '.join(conds)
    sql = f"""
        SELECT dp.producto_codigo_erp                                           AS codigo,
               INITCAP(dp.producto_nombre)                                      AS nombre,
               COALESCE(dp.marca, '')                                           AS marca,
               COALESCE(SUM(fv.venta_neta), 0)                                 AS bs,
               COALESCE(SUM(fv.cantidad), 0)                                    AS uds,
               FLOOR(COALESCE(SUM(
                   CASE WHEN dp.linea = 'BEBIDAS ALC'
                        THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                        ELSE 0 END
               ), 0))                                                            AS cajas9l,
               CASE WHEN MAX(dp.linea) = 'BEBIDAS ALC' THEN true ELSE false END AS es_licor
        FROM dw.fact_ventas fv
        JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
        JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
        WHERE {where}
        GROUP BY dp.producto_codigo_erp, dp.producto_nombre, dp.marca
        ORDER BY bs DESC
        LIMIT 50
    """
    try:
        _, rows = _run_dw_query(sql, params)
        data = [
            {
                'codigo':   r['codigo'],
                'nombre':   r['nombre'],
                'marca':    r['marca'],
                'bs':       float(r['bs']),
                'uds':      int(r['uds']),
                'cajas9l':  int(r['cajas9l']),
                'es_licor': r['es_licor'],
            }
            for r in rows
        ]
        return JsonResponse({'success': True, 'data': data})
    except Exception:
        logger.exception("Error interno")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
@_require_perm('comportamiento-productos')
def dashboard_comportamiento_tabla(request):
    regional_key = _safe_str(request.GET.get('regional', ''), 50).lower()
    canal        = _safe_str(request.GET.get('canal', ''), 100)
    vendedor     = _safe_str(request.GET.get('vendedor', ''), 150)
    codigo       = _safe_str(request.GET.get('codigo', ''), 100)

    if not codigo:
        return JsonResponse({'success': False, 'error': 'Parametro codigo es requerido'})

    base_conds  = ['dp.producto_codigo_erp = %s']
    base_params = [codigo]

    regional_cond = _regional_filter(regional_key, 'dv.ciudad')
    if regional_cond:
        base_conds.append(regional_cond)

    if canal:
        base_conds.append('dv.canal_rrhh = %s')
        base_params.append(canal)

    if vendedor:
        base_conds.append('INITCAP(dv.vendedor_nombre) = %s')
        base_params.append(vendedor)

    base_where = ' AND '.join(base_conds)

    sql_anho_mes = f"""
        SELECT df.anho,
               df.mes_numero                                AS mes,
               COUNT(DISTINCT fv.cliente_sk)                AS clientes,
               COALESCE(SUM(fv.cantidad), 0)                AS uds,
               FLOOR(COALESCE(SUM(
                   CASE WHEN dp.linea = 'BEBIDAS ALC'
                        THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                        ELSE 0 END
               ), 0))                                        AS cajas9l
        FROM dw.fact_ventas fv
        JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
        JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
        WHERE {base_where}
        GROUP BY df.anho, df.mes_numero
        ORDER BY df.anho, df.mes_numero
    """

    sql_canal_anho_mes = f"""
        SELECT dv.canal_rrhh                                AS canal,
               df.anho,
               df.mes_numero                                AS mes,
               COUNT(DISTINCT fv.cliente_sk)                AS clientes,
               COALESCE(SUM(fv.cantidad), 0)                AS uds,
               FLOOR(COALESCE(SUM(
                   CASE WHEN dp.linea = 'BEBIDAS ALC'
                        THEN fv.cantidad::NUMERIC * COALESCE(dp.u_L, 0) / 9000.0
                        ELSE 0 END
               ), 0))                                        AS cajas9l
        FROM dw.fact_ventas fv
        JOIN dw.dim_fecha    df ON df.fecha_sk    = fv.fecha_sk
        JOIN dw.dim_vendedor dv ON dv.vendedor_sk = fv.vendedor_sk
        JOIN dw.dim_producto dp ON dp.producto_sk = fv.producto_sk
        WHERE {base_where}
        GROUP BY dv.canal_rrhh, df.anho, df.mes_numero
        ORDER BY dv.canal_rrhh, df.anho, df.mes_numero
    """

    try:
        _, rows1 = _run_dw_query(sql_anho_mes, base_params)
        por_anho_mes = [
            {'anho': r['anho'], 'mes': r['mes'], 'clientes': int(r['clientes']), 'uds': int(r['uds']), 'cajas9l': int(r['cajas9l'])}
            for r in rows1
        ]
    except Exception:
        logger.exception("Error interno - por_anho_mes")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)

    try:
        _, rows2 = _run_dw_query(sql_canal_anho_mes, base_params)
        por_canal_anho_mes = [
            {
                'canal':    r['canal'] or '',
                'anho':     r['anho'],
                'mes':      r['mes'],
                'clientes': int(r['clientes']),
                'uds':      int(r['uds']),
                'cajas9l':  int(r['cajas9l']),
            }
            for r in rows2
        ]
    except Exception:
        logger.exception("Error interno - por_canal_anho_mes")
        return JsonResponse({'success': False, 'error': 'Error interno del servidor'}, status=500)

    return JsonResponse({
        'success':            True,
        'por_anho_mes':       por_anho_mes,
        'por_canal_anho_mes': por_canal_anho_mes,
    })


# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
#  REPORTES / TICKETS
# â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def _serialize_reporte(r):
    return {
        'id':          r.id,
        'tipo':        r.tipo,
        'subtipo':     r.subtipo,
        'descripcion': r.descripcion,
        'estado':      r.estado,
        'prioridad':   r.prioridad,
        'context':     r.context,
        'created_at':  r.created_at.isoformat(),
        'updated_at':  r.updated_at.isoformat(),
        'user': {
            'id':        r.user.id        if r.user else None,
            'username':  r.user.username  if r.user else '(eliminado)',
            'full_name': r.user.get_full_name() if r.user else '(eliminado)',
        },
    }


@api_view(['POST'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def reporte_create(request):
    """Cualquier usuario autenticado puede crear un reporte."""
    data = request.data
    tipo        = data.get('tipo', '').upper()
    subtipo     = data.get('subtipo', '')
    descripcion = (data.get('descripcion') or '').strip()
    context     = data.get('context', {})

    if tipo not in ('BUG', 'ERROR', 'SOLICITUD'):
        return JsonResponse({'success': False, 'error': 'Tipo invÃ¡lido'}, status=400)
    if not descripcion:
        return JsonResponse({'success': False, 'error': 'La descripciÃ³n es requerida'}, status=400)

    r = Reporte.objects.create(
        user=request.user,
        tipo=tipo,
        subtipo=subtipo,
        descripcion=descripcion,
        context=context if isinstance(context, dict) else {},
    )
    return JsonResponse({'success': True, 'id': r.id}, status=201)


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def reporte_list(request):
    """Lista todos los reportes. Solo admins/subadmins."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    # Marcar como revisado ahora
    UserProfile.objects.filter(user=request.user).update(
        reports_last_checked=timezone.now()
    )

    qs = Reporte.objects.select_related('user').all()

    tipo     = request.GET.get('tipo', '')
    estado   = request.GET.get('estado', '')
    prioridad = request.GET.get('prioridad', '')
    search   = request.GET.get('search', '').strip()

    if tipo:
        qs = qs.filter(tipo=tipo.upper())
    if estado:
        qs = qs.filter(estado=estado.upper())
    if prioridad:
        qs = qs.filter(prioridad=prioridad.upper())
    if search:
        qs = qs.filter(
            models.Q(user__username__icontains=search)  |
            models.Q(user__first_name__icontains=search) |
            models.Q(user__last_name__icontains=search)  |
            models.Q(descripcion__icontains=search)
        )

    return JsonResponse([_serialize_reporte(r) for r in qs], safe=False)


@api_view(['PATCH'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def reporte_update(request, reporte_id):
    """Actualiza estado y/o prioridad de un reporte. Solo admins/subadmins."""
    if not _is_user_manager(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)

    try:
        r = Reporte.objects.get(id=reporte_id)
    except Reporte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No encontrado'}, status=404)

    ESTADOS   = {c[0] for c in Reporte.ESTADO_CHOICES}
    PRIORIDADES = {c[0] for c in Reporte.PRIORIDAD_CHOICES}

    nuevo_estado    = (request.data.get('estado') or '').upper()
    nueva_prioridad = (request.data.get('prioridad') or '').upper()

    if nuevo_estado and nuevo_estado in ESTADOS:
        r.estado = nuevo_estado
    if nueva_prioridad and nueva_prioridad in PRIORIDADES:
        r.prioridad = nueva_prioridad
    r.save(update_fields=['estado', 'prioridad', 'updated_at'])

    return JsonResponse({'success': True, 'reporte': _serialize_reporte(r)})


@api_view(['GET'])
@authentication_classes([ExpiringTokenAuthentication])
@permission_classes([IsAuthenticated])
def reporte_unread_count(request):
    """Devuelve cuÃ¡ntos reportes nuevos hay desde la Ãºltima vez que el admin revisÃ³."""
    if not _is_user_manager(request.user):
        return JsonResponse({'count': 0})

    profile = _get_or_create_profile(request.user)
    since   = profile.reports_last_checked

    count = Reporte.objects.filter(
        created_at__gt=since
    ).count() if since else Reporte.objects.count()

    return JsonResponse({'count': count})
